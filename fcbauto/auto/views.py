import os
import logging
import pandas as pd
import re

logger = logging.getLogger(__name__)
import dateparser
import numpy as np
from datetime import datetime, timedelta
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .forms import ExcelUploadForm
from .map import consu_mapping, comm_mapping, guar_mapping, credit_mapping, prin_mapping,Gender_dict,Country_dict,state_dict,Marital_dict,Borrower_dict,Employer_dict,Title_dict,Occu_dict,AccountStatus_dict,Loan_dict,Repayment_dict,Currency_dict,Classification_dict,Collateraltype_dict,Positioninbusiness_dict,ConsuToComm,CommToConsu, commercial_keywords,consumer_merged_mapping,commercial_merged_mapping,guarantor_columns_to_clear,principal_officer_columns_to_clear,sheet_name_mappings,TIER_1_SUFFIXES,TIER_3_AMBIGUOUS
from .filename_utils import generate_filename, generate_fallback_filename, generate_filename_from_user
from rapidfuzz import fuzz, process
from typing import Union, Optional
from word2number import w2n
from datetime import datetime
import traceback
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from .models import UploadSession
from .exceptions import (
    DataValidationError, FileProcessingError, MergeError,
    OutputGenerationError, VerificationError
)
from django.db.models import Count, Avg
from django.utils import timezone
from .file_archival_utils import save_original_file, delete_original_file
from datetime import timedelta
# from reportlab.lib.pagesizes import letter, A4
# from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
# from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
# from reportlab.lib.units import inch
# from reportlab.lib import colors
# from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.http import HttpResponse
from django.conf import settings
from io import BytesIO
import os

# Rate limiting for upload protection
try:
    from django_ratelimit.decorators import ratelimit
    RATELIMIT_AVAILABLE = True
except ImportError:
    RATELIMIT_AVAILABLE = False
    logger.warning("[SECURITY] django-ratelimit not installed. Rate limiting disabled.")




def extract_subscriber_alias_from_filename(filename):
    """
    Extract subscriber alias from filename by removing date patterns
    
    Args:
        filename (str): The filename to extract subscriber alias from
        
    Returns:
        str: The subscriber alias without date information
    """
    if not filename:
        return filename
    
    # Remove file extension
    base_filename = filename
    if '.' in base_filename:
        base_filename = base_filename.rsplit('.', 1)[0]
    
    # Remove common date patterns
    # Pattern 1: Remove YYYY_MM_DD format
    base_filename = re.sub(r'[_\s]*\d{4}[_\s]*\d{1,2}[_\s]*\d{1,2}[_\s]*', '', base_filename)
    
    # Pattern 2: Remove Month_Year or Year_Month patterns
    month_names = ['january', 'february', 'march', 'april', 'may', 'june',
                   'july', 'august', 'september', 'october', 'november', 'december',
                   'jan', 'feb', 'mar', 'apr', 'jun', 'jul', 'aug', 'sep', 'sept', 'oct', 'nov', 'dec']
    
    for month_name in month_names:
        # Remove month_year pattern (e.g., "may 2024", "may_2024")
        pattern = rf'[_\s]*{month_name}[_\s]*\d{{4}}[_\s]*'
        base_filename = re.sub(pattern, '', base_filename, flags=re.IGNORECASE)
        
        # Remove year_month pattern (e.g., "2024 may", "2024_may")
        pattern = rf'[_\s]*\d{{4}}[_\s]*{month_name}[_\s]*'
        base_filename = re.sub(pattern, '', base_filename, flags=re.IGNORECASE)
    
    # Clean up any trailing/leading spaces or underscores
    base_filename = base_filename.strip(' _')
    
    # If nothing left, return original filename without extension
    if not base_filename:
        return filename.rsplit('.', 1)[0] if '.' in filename else filename
    
    return base_filename


def extract_date_from_filename(filename):
    """
    Extract month and year from filename in various formats
    
    Args:
        filename (str): The filename to extract date from
        
    Returns:
        tuple: (month, year) as integers, or (None, None) if no date found
    """
    if not filename:
        return None, None
    
    # Remove file extension
    base_filename = filename.lower()
    if '.' in base_filename:
        base_filename = base_filename.rsplit('.', 1)[0]
    
    # Pattern 1: YYYY_MM_DD format (e.g., alekun_2024_03_31)
    pattern1 = r'(\d{4})_(\d{1,2})_(\d{1,2})'
    match1 = re.search(pattern1, base_filename)
    if match1:
        year, month, day = match1.groups()
        return int(month), int(year)
    
    # Pattern 2: Month_Year format (e.g., alekun_may_2024, alekun_march_2024)
    month_names = {
        'january': 1, 'jan': 1,
        'february': 2, 'feb': 2,
        'march': 3, 'mar': 3,
        'april': 4, 'apr': 4,
        'may': 5,
        'june': 6, 'jun': 6,
        'july': 7, 'jul': 7,
        'august': 8, 'aug': 8,
        'september': 9, 'sep': 9, 'sept': 9,
        'october': 10, 'oct': 10,
        'november': 11, 'nov': 11,
        'december': 12, 'dec': 12
    }
    
    # Look for month name followed by year
    for month_name, month_num in month_names.items():
        pattern2 = rf'{month_name}[_\s]*?(\d{{4}})'
        match2 = re.search(pattern2, base_filename)
        if match2:
            year = match2.group(1)
            return month_num, int(year)
    
    # Pattern 3: Year_Month format (e.g., alekun_2024_may)
    for month_name, month_num in month_names.items():
        pattern3 = rf'(\d{{4}})[_\s]*?{month_name}'
        match3 = re.search(pattern3, base_filename)
        if match3:
            year = match3.group(1)
            return month_num, int(year)
    
    # No date pattern found
    return None, None


def create_empty_sheet(mapping_dict):
    """
    Create an empty DataFrame with columns from the mapping dictionary
    """
    columns = list(mapping_dict.keys())
    return pd.DataFrame(columns=columns)


def validate_required_sheets_present(xds, cleaned_names_override=None):
    """
    Validate that an uploaded workbook contains the minimum required sheets
    before any processing begins.

    Rules:
    - If any merged sheet (consumermerged / commercialmerged) is present,
      skip the check entirely — merged sheets are self-contained.
    - Otherwise BOTH of the following must be present:
        1. At least one borrower sheet:
               IndividualBorrowerTemplate  OR  CorporateBorrowerTemplate
        2. CreditInformation
    - PrincipalOfficersTemplate and GuarantorsInformation are optional.

    Args:
        xds: dict of sheet-name → DataFrame (used to derive cleaned names when
             cleaned_names_override is not provided)
        cleaned_names_override: optional set of already-canonical sheet-type names.
             Pass this when the xds keys are raw filenames (e.g. for CSV uploads)
             rather than Excel sheet tab names.

    Returns:
        tuple: (is_valid: bool, error_message: str)
    """
    if cleaned_names_override is not None:
        cleaned_names = set(cleaned_names_override)
    else:
        cleaned_names = {clean_sheet_name(name) for name in xds.keys()}

    # Merged sheets bypass the check
    if 'consumermerged' in cleaned_names or 'commercialmerged' in cleaned_names:
        return True, ""

    has_borrower = (
        'individualborrowertemplate' in cleaned_names
        or 'corporateborrowertemplate' in cleaned_names
    )
    has_credit = 'creditinformation' in cleaned_names

    if has_borrower and has_credit:
        return True, ""

    missing = []
    if not has_borrower:
        missing.append("a Borrower sheet (Individual Borrower Template or Corporate Borrower Template)")
    if not has_credit:
        missing.append("Credit Information sheet")

    error_message = (
        "❌ Validation Failed - Required Sheets Are Missing.\n"
        + "\n".join(f"- Missing: {m}" for m in missing)
        + "\nPlease include all required sheets and re-upload your file."
    )
    return False, error_message


def ensure_all_sheets_exist(xds):
    """
    Check for missing sheets and create them with appropriate headers if needed
    """
    # Define expected sheets and their corresponding mappings
    expected_sheets = {
        'individualborrowertemplate': consu_mapping,
        'corporateborrowertemplate': comm_mapping,
        'creditinformation': credit_mapping,
        'guarantorsinformation': guar_mapping,
        'principalofficerstemplate': prin_mapping
    }
    
    processed_sheets = {}
    missing_sheets = []
    existing_sheets = []
    
    logger.info("\n=== SHEET PROCESSING REPORT ===")
    logger.info("Checking for required sheets...")
    
    # First check if we have merged sheets
    has_merged_sheets = False
    for original_name in xds.keys():
        cleaned_name = clean_sheet_name(original_name)
        if cleaned_name in ['consumermerged', 'commercialmerged']:
            has_merged_sheets = True
            logger.info(f"? Found merged sheet: {original_name}")
            processed_sheets[cleaned_name] = xds[original_name]
            existing_sheets.append(original_name)
    
    # If we have merged sheets, skip generating missing sheets
    if has_merged_sheets:
        logger.info("\n=== MERGED SHEETS DETECTED ===")
        logger.info("Skipping generation of missing sheets as merged sheets are present")
        return processed_sheets
    
    # Regular sheet processing if no merged sheets found
    for sheet_name, mapping in expected_sheets.items():
        # Clean the sheet name for comparison
        cleaned_name = clean_sheet_name(sheet_name)
        
        # Check if sheet exists in uploaded file
        sheet_exists = False
        for original_name in xds.keys():
            if clean_sheet_name(original_name) == cleaned_name:
                logger.info(f"? Found existing sheet: {original_name}")
                processed_sheets[cleaned_name] = xds[original_name]
                sheet_exists = True
                existing_sheets.append(sheet_name)
                break
        
        # If sheet doesn't exist, create it
        if not sheet_exists:
            logger.info(f"? Missing sheet detected: {sheet_name}")
            logger.info(f"? Generating new sheet: {sheet_name}")
            logger.info(f"  - Adding {len(mapping)} columns based on template")
            processed_sheets[cleaned_name] = create_empty_sheet(mapping)
            missing_sheets.append(sheet_name)
    
    # Summary report
    logger.info("\n=== SHEET GENERATION SUMMARY ===")
    logger.info(f"Total sheets required: {len(expected_sheets)}")
    logger.info(f"Sheets found in upload: {len(existing_sheets)}")
    logger.info(f"Sheets auto-generated: {len(missing_sheets)}")
    
    if missing_sheets:
        logger.info("\nAuto-generated sheets:")
        for sheet in missing_sheets:
            logger.info(f"- {sheet}")
    
    logger.info("\n=============================")
    
    return processed_sheets

def preprocess_tenor_from_headers(df):
    """
    Checks column headers for time units (e.g., 'Loan Tenor (Months)')
    and converts the data in that column to days. This version handles
    days, weeks, months, and years.
    """
    df_copy = df.copy()
    
    # --- UPDATED: Comprehensive dictionary for all units ---
    header_unit_multipliers = {
        # Weeks
        'weeks': 7, 'week': 7, 'w': 7,
        # Months
        'months': 30, 'month': 30, 'mnth': 30, 'mth': 30, 
        'mths': 30, 'mnths': 30, 'mons': 30, 'm': 30,
        # Years
        'years': 365, 'year': 365, 'y': 365, 'yr': 365, 'yrs': 365,
    }

    # Regex to find any of the units in the dictionary, ignoring case
    # This looks for the unit as a whole word
    pattern = r'\b(' + '|'.join(header_unit_multipliers.keys()) + r')\b'

    for col in df_copy.columns:
        # Skip arrears columns - they are handled by preprocess_arrears_from_headers
        col_lower = str(col).lower().replace(' ', '').replace('_', '')
        if 'arrears' in col_lower or 'dpd' in col_lower or 'overdue' in col_lower or 'pastdue' in col_lower:
            continue
        
        # Search for a unit in the column name (case-insensitive)
        match = re.search(pattern, col, re.IGNORECASE)
        
        if match:
            unit_found = match.group(0).lower()
            multiplier = header_unit_multipliers[unit_found]
            
            logger.info(f"Found unit '{unit_found}' in header '{col}'. Applying multiplier: {multiplier}")
            
            # Apply the multiplier to the column.
            # pd.to_numeric converts numbers; errors='coerce' handles non-numbers gracefully.
            # Convert to numeric, preserve empty values as empty strings
            numeric_series = pd.to_numeric(df_copy[col], errors='coerce')
            # Replace NaN with empty string, otherwise multiply by multiplier and convert to int then string
            df_copy[col] = numeric_series.apply(lambda x: '' if pd.isna(x) else str(int(x * multiplier)))
    
    return df_copy

def preprocess_arrears_from_headers(df):
    """
    Preprocess arrears columns (DAYSINARREARS, MONTHSINARREARS, etc.):
    1. Extract only numeric values from text (e.g., 'over 500 days' → 500)
    2. If column name indicates months, multiply by 30 to convert to days
    
    This runs BEFORE column mapping, so works on original column names.
    """
    df_copy = df.copy()
    
    # Patterns that indicate months (need to multiply by 30)
    months_patterns = ['monthsinarrears', 'monthinarrears', 'montharrears', 
                       'monthsarrears', 'dpdinmonths', 'arrearsinmonths']
    
    # Patterns that indicate arrears columns we should process
    arrears_patterns = months_patterns + [
        'daysinarrears', 'daysoverdue', 'dayspastdue', 'dpd',
        'numberofdayspastdue', 'numberofdaysoverdue', 'noofdaysoverdue',
        'overduedaysprin', 'daysareas'
    ]
    
    for col in df_copy.columns:
        col_lower = str(col).lower().replace(' ', '').replace('_', '')
        
        # Check if this column matches any arrears pattern
        is_arrears_col = any(pattern in col_lower for pattern in arrears_patterns)
        
        if not is_arrears_col:
            continue
        
        # Check if this is a months column (needs × 30)
        # Check both: full patterns AND standalone months in parentheses
        is_months = any(pattern in col_lower for pattern in months_patterns)
        
        # Also check for months unit in parentheses like "(months)" or "(month)" or "(mths)"
        if not is_months:
            months_unit_pattern = re.search(r'\((?:months?|mths?|mnths?)\)', col_lower)
            is_months = months_unit_pattern is not None
        
        multiplier = 30 if is_months else 1
        
        if is_months:
            logger.info(f"Found months arrears column '{col}'. Will multiply by 30.")
        else:
            logger.info(f"Processing arrears column '{col}' - extracting numbers.")
        
        def extract_and_convert(value):
            if pd.isna(value) or value == '':
                return ''
            
            value_str = str(value).strip()

            value_str = re.sub(r'(\d)[,\s](\d)', r'\1\2', value_str)
            
            # Find all numbers in the string
            numbers = re.findall(r'\d+', value_str)
            
            if not numbers:
                return ''
            
            # Use the first number found
            num = int(numbers[0])
            
            # Apply multiplier if months column
            result = num * multiplier
            
            return str(result)
        
        df_copy[col] = df_copy[col].apply(extract_and_convert)
    
    return df_copy

def clean_sheet_name(sheet_name):
    """
    Clean sheet names by removing special characters and normalize common variations
    to canonical sheet type names.
    
    This allows users to name their sheets flexibly while still getting
    proper column mapping and processing applied.
    
    Uses sheet_name_mappings from map.py for the normalization.
    """
    # First, clean the sheet name (remove special chars, lowercase)
    cleaned_name = re.sub(r'[^a-zA-Z0-9]', '', sheet_name).lower()
    
    # Check if cleaned name matches any known variation (from map.py)
    if cleaned_name in sheet_name_mappings:
        return sheet_name_mappings[cleaned_name]
    
    # If no exact match, return the cleaned name as-is
    return cleaned_name


def make_column_names_unique(df):
    """
    Checks for duplicate column names and makes them unique by adding a suffix.
    e.g., [NAME, EMAIL, NAME] -> [NAME, EMAIL, NAME_1]
    """
    cols = list(df.columns)
    seen = set()
    for i, col in enumerate(cols):
        if col in seen:
            suffix = 1
            while f"{col}_{suffix}" in seen:
                suffix += 1
            new_name = f"{col}_{suffix}"
            cols[i] = new_name
            seen.add(new_name)
        else:
            seen.add(col)
    df.columns = cols
    return df

def remove_special_characters(column_name):
    """Remove special characters and all spaces from column names"""
    # Remove non-alphanumeric characters but allow spaces
    pattern = r'[^a-zA-Z0-9]'  # Remove special characters
    cleaned_name = re.sub(pattern, '', column_name)  # Remove special characters
    
    # Remove all spaces
    cleaned_name = cleaned_name.replace(' ', '')  # Remove all spaces
    
    return cleaned_name

def remove_special_chars(text):
    """Remove special characters from text while preserving spaces"""
    if pd.isna(text) or text is None:  # Handle NaN and None values
        return ''
    if not isinstance(text, str):
        text = str(text)
    
    # Remove special characters but keep spaces
    cleaned = re.sub(r'[^a-zA-Z\s]', '', text)
    
    # Replace multiple spaces with single space and trim
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    
    return cleaned

def remove_titles(name):
    if not isinstance(name, str):
        return name
    
    titles = [
        'Miss', 'Mrs', 'Rev', 'Dr', 'Mr', 'MS', 'CAPT','pastor','doctor',
        'COL', 'LADY', 'MAJ', 'PST', 'PROF', 'REV', 'SGT',
        'SIR', 'HE', 'JUDG', 'CHF', 'ALHJ', 'APOS', 'CDR', 'ALH', 'Alh',
        'BISH', 'FLT', 'BARR', 'MGEN', 'GEN', 'HON', 'ENGR', 'LT', 'AND', 'and',
        'PASTOR', 'PAST', 'PST', 'ALHAJI', 'ALH', 'ALH.', 'ALHAJ', 'ALHADJI', 'ALHAJJI', 'ALHAJ.', 'ALHADJ', 'ALHADJ.',
        'PASTOR.', 'PASTOR', 'PAST.', 'PST.', 'REV.', 'REV', 'DR.', 'MR.', 'MRS.', 'MS.'
    ]
    
    pattern = r'\b(?:' + '|'.join(re.escape(title) for title in titles) + r')\b'
    cleaned_name = re.sub(pattern, '', name, flags=re.IGNORECASE)
    return ' '.join(cleaned_name.split())


def remove_duplicate_columns(df):
    """
    Remove duplicate columns, keeping first occurrence
    
    Args:
        df (pd.DataFrame): Input DataFrame
    
    Returns:
        pd.DataFrame: DataFrame with unique columns
    """
    if df is None or df.empty:
        return df
    
    # Identify unique columns
    unique_columns = []
    duplicate_columns = []  # To keep track of duplicates
    for col in df.columns:
        if col not in unique_columns:
            unique_columns.append(col)
        else:
            duplicate_columns.append(col)  # Track duplicates
    
    # Create DataFrame with unique columns
    df_cleaned = df[unique_columns]
    
    # Log column removals
    columns_removed = len(df.columns) - len(unique_columns)
    if columns_removed > 0:
        logger.info(f"Removed {columns_removed} duplicate columns: {duplicate_columns}")
    
    return df_cleaned


def convert_date(date_string):
    """Converts a date string or Excel serial number to the specified format (YYYYMMDD), 
    or returns None for empty/invalid rows.
    
    Args:
        date_string: A string or number representing a date.
        
    Returns:
        A string representing the date in the specified format (YYYYMMDD), or None for empty or invalid dates.
    """
    # Check if the cell is empty or None
    if date_string is None or (isinstance(date_string, float) and np.isnan(date_string)):
        return None

    # Define common missing value representations
    missing_values = ["", "None", "NaN", "null", "N/A", "n/a", "na", "NA", "#N/A", "?", "missing",'N.A']
    
    # Check if the cell is a missing value
    if isinstance(date_string, str) and date_string.strip() in missing_values:
        return None
        
    # Check if the date is already in YYYYMMDD format
    if isinstance(date_string, str):
        # Remove any whitespace
        clean_date = date_string.strip()
        # Check if it's already in YYYYMMDD format (8 digits with no separators)
        if re.match(r'^\d{8}$', clean_date):
            # Validate that it's a valid date
            try:
                year = int(clean_date[:4])
                month = int(clean_date[4:6])
                day = int(clean_date[6:8])
                # Basic validation
                if 1 <= month <= 12 and 1 <= day <= 31 and 1900 <= year <= 2100:
                    return clean_date  # Already in the correct format
            except (ValueError, IndexError):
                pass  # Not a valid YYYYMMDD date, continue with conversion

    # Check if the input is a number (e.g., Excel serial number)
    try:
        serial_number = float(date_string)
        
        # Check if the serial number is within the valid Excel date range
        if serial_number <= 0 or serial_number > 2958465:
            return None  # Invalid range for Excel date serial numbers
        
        # Excel serial date base is 1899-12-30
        base_date = datetime(1899, 12, 30)
        calculated_date = base_date + timedelta(days=int(serial_number))
        if calculated_date.year < 1900:
            return None
        return f"{calculated_date.year:04d}{calculated_date.month:02d}{calculated_date.day:02d}"
    except (ValueError, TypeError):
        # If not a valid number, proceed with parsing as a string
        pass

    # Define date formats with explicit separation between 2-digit and 4-digit year formats
    two_digit_year_formats = [
        '%d/%m/%y', '%m/%d/%y', '%y/%m/%d',  # Two-digit year formats with slashes
        '%d-%m-%y', '%m-%d-%y', '%y-%m-%d',  # Two-digit year formats with hyphens
        '%d.%m.%y', '%m.%d.%y', '%y.%m.%d',  # Two-digit year formats with dots
        '%d-%b-%y', '%b-%d-%y', '%y-%b-%d',  # Two-digit year formats with short month names (Jan, Feb, Mar)
        '%d/%b/%y', '%b/%d/%y', '%y/%b/%d',  # Two-digit year formats with short month names and slashes
        '%d.%b.%y', '%b.%d.%y', '%y.%b.%d',  # Two-digit year formats with short month names and dots
    ]
    
    four_digit_year_formats = [
        '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d',  # Four-digit year formats with slashes
        '%d-%m-%Y', '%m-%d-%Y', '%Y-%m-%d',  # Four-digit year formats with hyphens
        '%d.%m.%Y', '%m.%d.%Y', '%Y.%m.%d',  # Four-digit year formats with dots
        '%Y%m%d', '%d%m%Y', '%m%d%Y'         # Four-digit year formats without separators
    ]
    
    # First try with four-digit year formats
    for fmt in four_digit_year_formats:
        try:
            date = datetime.strptime(str(date_string).strip(), fmt)
            if date.year < 1900:
                return None
            return f"{date.year:04d}{date.month:02d}{date.day:02d}"
        except ValueError:
            continue
    
    # Then try with two-digit year formats and apply the sliding window
    for fmt in two_digit_year_formats:
        try:
            date = datetime.strptime(str(date_string).strip(), fmt)
            
            # Apply Excel's sliding window logic for two-digit years
            two_digit_year = date.year % 100
            if 0 <= two_digit_year <= 29:
                adjusted_year = 2000 + two_digit_year
            else:
                adjusted_year = 1900 + two_digit_year
            
            # Replace the year while keeping month/day the same
            date = date.replace(year=adjusted_year)
            if date.year < 1900:
                return None
            return f"{date.year:04d}{date.month:02d}{date.day:02d}"
        except ValueError:
            continue
    
    # If all explicit formats fail, try dateparser as a fallback
    try:
        date = dateparser.parse(str(date_string))
        if date:
            if date.year < 1900:
                return None
            return f"{date.year:04d}{date.month:02d}{date.day:02d}"
    except Exception:
        pass
        
    return None


# ============================================================================
# OPTIMIZED DATE PROCESSING (Vectorized + Cached)
# ============================================================================

from functools import lru_cache

# Create cached version of convert_date for fallback cases
@lru_cache(maxsize=20000)
def convert_date_cached(date_string):
    """Cached wrapper around convert_date for edge cases"""
    return convert_date(date_string)


def process_dates_vectorized(df, cutoff_date=None):
    """
    Optimized vectorized date processing - 10-50x faster than row-by-row .apply()
    
    Maintains exact same behavior as convert_date():
    - Returns YYYYMMDD format strings
    - Returns None for invalid/missing dates
    - Handles Excel serials, multiple text formats, 2-digit year sliding window
    
    For LOANEFFECTIVEDATE and ACCOUNTSTATUSDATE columns with cutoff_date:
    - If MDY interpretation gives date > cutoff, tries DMY interpretation
    - This handles ambiguous dates like 03/05/2024 that could be March 5 or May 3
    """
    date_columns = [
        'DATEOFBIRTH', 'DATEOFINCORPORATION', 
        'PRINCIPALOFFICER1DATEOFBIRTH', 'PRINCIPALOFFICER2DATEOFBIRTH',
        'SPOUSEDATEOFBIRTH', 'GUARANTORDATEOFBIRTHINCORPORATION',
        'LOANEFFECTIVEDATE', 'MATURITYDATE', 'LASTPAYMENTDATE',
        'DEFEREDPAYMENTDATE', 'LITIGATIONDATE', 'ACCOUNTSTATUSDATE'
    ]
    
    # Columns that need cutoff validation for MDY/DMY switching
    cutoff_columns = ['LOANEFFECTIVEDATE']
    
    missing_values = ["", "None", "NaN", "null", "N/A", "n/a", "na", "NA", "#N/A", "?", "missing", 'N.A', 'nan']
    
    for col in df.columns:
        if 'date' not in col.lower() and col not in date_columns:
            continue
            
        logger.info(f"[OPTIMIZED] Processing date column: {col}")
        original_count = len(df)
        
        # Step 1: Convert to string, strip whitespace
        series = df[col].astype(str).str.strip()
        
        # Step 2: Replace missing values with pd.NA
        series = series.replace(missing_values, pd.NA)
        
        # Initialize result series
        result = pd.Series([None] * len(df), index=df.index, dtype=object)
        
        # Step 3: Detect already-formatted YYYYMMDD dates (vectorized regex)
        mask_yyyymmdd = series.str.match(r'^\d{8}$', na=False)
        if mask_yyyymmdd.any():
            # Validate year/month/day ranges
            valid_yyyymmdd = series[mask_yyyymmdd].copy()
            years = valid_yyyymmdd.str[:4].astype(int)
            months = valid_yyyymmdd.str[4:6].astype(int)
            days = valid_yyyymmdd.str[6:8].astype(int)
            
            valid_mask = (years >= 1900) & (years <= 2100) & (months >= 1) & (months <= 12) & (days >= 1) & (days <= 31)
            result[mask_yyyymmdd] = valid_yyyymmdd.where(valid_mask, None)
            logger.info(f"  - Found {mask_yyyymmdd.sum()} already-formatted YYYYMMDD dates")
        
        # Step 4: Try numeric (Excel serial) conversion (vectorized)
        remaining_mask = result.isna() & series.notna()
        if remaining_mask.any():
            numeric_series = pd.to_numeric(series[remaining_mask], errors='coerce')
            is_excel_serial = (numeric_series > 0) & (numeric_series <= 2958465)
            
            if is_excel_serial.any():
                try:
                    # Convert Excel serials (base date: 1899-12-30)
                    excel_dates = pd.to_datetime(
                        numeric_series[is_excel_serial],
                        unit='D',
                        origin='1899-12-30',
                        errors='coerce'
                    )
                    # Filter out dates before 1900
                    valid_excel = excel_dates[excel_dates.dt.year >= 1900]
                    result.loc[valid_excel.index] = valid_excel.dt.strftime('%Y%m%d')
                    logger.info(f"  - Converted {is_excel_serial.sum()} Excel serial dates")
                except Exception as e:
                    logger.info(f"  - Excel serial conversion error: {e}")
        
        # Step 5: Try pandas datetime parsing (vectorized - handles most text formats)
        # IMPORTANT: Must try 4-digit year formats BEFORE 2-digit to match original logic
        remaining_mask = result.isna() & series.notna()
        if remaining_mask.any():
            remaining_series = series[remaining_mask]
            
            # First, try common 4-digit year formats (matches original precedence)
            # Include formats WITH time components (time will be ignored, only date extracted)
            four_digit_formats = [
                # Formats with slashes
                '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d',
                # Formats with hyphens
                '%d-%m-%Y', '%m-%d-%Y', '%Y-%m-%d',
                # Formats with dots
                '%d.%m.%Y', '%m.%d.%Y', '%Y.%m.%d',
                # Formats without separators
                '%Y%m%d', '%d%m%Y', '%m%d%Y',
                # Datetime formats with time components (slashes)
                '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S', '%Y/%m/%d %H:%M:%S',
                '%d/%m/%Y %H:%M', '%m/%d/%Y %H:%M', '%Y/%m/%d %H:%M',
                # Datetime formats with time components (hyphens)
                '%d-%m-%Y %H:%M:%S', '%m-%d-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S',
                '%d-%m-%Y %H:%M', '%m-%d-%Y %H:%M', '%Y-%m-%d %H:%M'
            ]
            
            for fmt in four_digit_formats:
                still_remaining = result[remaining_mask].isna()
                if not still_remaining.any():
                    break
                    
                try:
                    parsed = pd.to_datetime(
                        remaining_series[still_remaining],
                        format=fmt,
                        errors='coerce'
                    )
                    valid = parsed[parsed.dt.year >= 1900]
                    result.loc[valid.index] = valid.dt.strftime('%Y%m%d')
                except Exception:
                    continue
            
            # Update remaining mask after 4-digit attempts
            remaining_mask = result.isna() & series.notna()
            
            # Then try 2-digit year formats with sliding window (00-29 -> 2000s, 30-99 -> 1900s)
            if remaining_mask.any():
                remaining_series = series[remaining_mask]
                two_digit_formats = [
                    # Formats with slashes
                    '%d/%m/%y', '%m/%d/%y', '%y/%m/%d',
                    # Formats with hyphens
                    '%d-%m-%y', '%m-%d-%y', '%y-%m-%d',
                    # Formats with dots
                    '%d.%m.%y', '%m.%d.%y', '%y.%m.%d',
                    # Formats with short month names (Jan, Feb, Mar) - hyphens
                    '%d-%b-%y', '%b-%d-%y', '%y-%b-%d',
                    # Formats with short month names - slashes
                    '%d/%b/%y', '%b/%d/%y', '%y/%b/%d',
                    # Formats with short month names - dots
                    '%d.%b.%y', '%b.%d.%y', '%y.%b.%d'
                ]
                
                for fmt in two_digit_formats:
                    still_remaining = result[remaining_mask].isna()
                    if not still_remaining.any():
                        break
                        
                    try:
                        parsed = pd.to_datetime(
                            remaining_series[still_remaining],
                            format=fmt,
                            errors='coerce'
                        )
                        
                        # Apply sliding window logic to 2-digit years
                        if parsed.notna().any():
                            def apply_sliding_window(dt):
                                if pd.isna(dt):
                                    return dt
                                two_digit_year = dt.year % 100
                                
                                # Sliding window: 00-29 -> 2000s, 30-99 -> 1900s
                                if 0 <= two_digit_year <= 29:
                                    adjusted_year = 2000 + two_digit_year
                                else:
                                    adjusted_year = 1900 + two_digit_year
                                
                                return dt.replace(year=adjusted_year)
                            
                            adjusted = parsed.dropna().apply(apply_sliding_window)
                            valid = adjusted[adjusted.dt.year >= 1900]
                            result.loc[valid.index] = valid.dt.strftime('%Y%m%d')
                    except Exception:
                        continue
            
            parsed_count = result.notna().sum() - mask_yyyymmdd.sum() - (is_excel_serial.sum() if 'is_excel_serial' in locals() else 0)
            if parsed_count > 0:
                logger.info(f"  - Parsed {parsed_count} dates using explicit formats")
        
        # Step 6: Fallback to cached convert_date for remaining edge cases
        remaining_mask = result.isna() & series.notna()
        remaining_count = remaining_mask.sum()
        
        if remaining_count > 0:
            logger.info(f"  - Using cached fallback for {remaining_count} edge cases")
            # Use cached version to avoid re-processing same values
            result[remaining_mask] = series[remaining_mask].apply(convert_date_cached)
        
        # Step 7: For LOANEFFECTIVEDATE and ACCOUNTSTATUSDATE, apply cutoff validation
        # If a date is beyond cutoff and could be interpreted as DMY instead of MDY, swap it
        if col in cutoff_columns and cutoff_date is not None:
            logger.info(f"  - Applying cutoff validation for {col} (cutoff: {cutoff_date.strftime('%Y-%m-%d')})")
            
            def validate_with_cutoff(row_idx):
                """Check if date needs MDY->DMY swap based on cutoff"""
                date_str = result[row_idx]
                original_str = series[row_idx]
                
                if pd.isna(date_str) or pd.isna(original_str):
                    return date_str
                
                try:
                    # Parse the result date
                    parsed_date = datetime.strptime(str(date_str), '%Y%m%d')
                    
                    # If date is beyond cutoff, check if original string is ambiguous
                    if parsed_date > cutoff_date:
                        original_clean = str(original_str).strip()
                        
                        # Check if it's an ambiguous date pattern (like 03/05/2024)
                        import re
                        if re.match(r'^\d{1,2}[/\-\.]\d{1,2}[/\-\.]\d{2,4}$', original_clean):
                            # Try DMY interpretation
                            dmy_formats = ['%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%d/%m/%y', '%d-%m-%y', '%d.%m.%y']
                            for fmt in dmy_formats:
                                try:
                                    dmy_parsed = datetime.strptime(original_clean, fmt)
                                    # Apply sliding window for 2-digit years
                                    if dmy_parsed.year < 100:
                                        two_digit = dmy_parsed.year % 100
                                        if 0 <= two_digit <= 29:
                                            dmy_parsed = dmy_parsed.replace(year=2000 + two_digit)
                                        else:
                                            dmy_parsed = dmy_parsed.replace(year=1900 + two_digit)
                                    
                                    # If DMY is valid and <= cutoff, use it
                                    if dmy_parsed >= datetime(1900, 1, 1) and dmy_parsed <= cutoff_date:
                                        logger.info(f"    - Swapped MDY->DMY for '{original_clean}': {date_str} -> {dmy_parsed.strftime('%Y%m%d')}")
                                        return dmy_parsed.strftime('%Y%m%d')
                                except ValueError:
                                    continue
                    
                    return date_str
                except Exception:
                    return date_str
            
            # Apply cutoff validation row by row (only for this specific column)
            for idx in result.index:
                result[idx] = validate_with_cutoff(idx)
            
            # Count how many dates are still beyond cutoff (for logging)
            future_count = 0
            for idx in result.index:
                if pd.notna(result[idx]):
                    try:
                        d = datetime.strptime(str(result[idx]), '%Y%m%d')
                        if d > cutoff_date:
                            future_count += 1
                    except:
                        pass
            if future_count > 0:
                logger.warning(f"  - {future_count} dates still beyond cutoff after MDY/DMY swap attempt")
        # Step 7b: For ACCOUNTSTATUSDATE, apply an independent auto-computed cutoff.
        # Cutoff = last day of the current calendar month at runtime.
        # Ambiguous dates beyond the cutoff are swapped (DMY); if still unresolvable,
        # the original parsed value is kept — nothing is blanked or removed.
        if col == 'ACCOUNTSTATUSDATE':
            import calendar as _calendar
            _now = datetime.now()
            _last_day = _calendar.monthrange(_now.year, _now.month)[1]
            _asd_cutoff = datetime(_now.year, _now.month, _last_day, 23, 59, 59)
            logger.info(f"  - Applying independent cutoff validation for ACCOUNTSTATUSDATE (cutoff: {_asd_cutoff.strftime('%Y-%m-%d')})")

            def validate_asd(row_idx):
                date_str = result[row_idx]

                if pd.isna(date_str):
                    return date_str

                try:
                    parsed_date = datetime.strptime(str(date_str), '%Y%m%d')

                    if parsed_date > _asd_cutoff:
                        # Swap month and day directly on the parsed result.
                        # This is format-agnostic — works regardless of whether the
                        # original was DD/MM/YYYY, YYYY-MM-DD, or a pandas datetime.
                        m, d = parsed_date.month, parsed_date.day
                        if m <= 12 and d <= 12 and m != d:  # Only ambiguous if both ≤ 12
                            try:
                                swapped = parsed_date.replace(month=d, day=m)
                                if swapped >= datetime(1900, 1, 1) and swapped <= _asd_cutoff:
                                    logger.info(f"    - [ASD] Reinterpreted {date_str} -> {swapped.strftime('%Y%m%d')} (swapped month/day)")
                                    return swapped.strftime('%Y%m%d')
                            except ValueError:
                                pass  # Swapped date is invalid (e.g. month=13) — keep original
                        # Not ambiguous or swap didn't help — keep original parsed value
                    return date_str
                except Exception:
                    return date_str

            for idx in result.index:
                result[idx] = validate_asd(idx)

            asd_future_count = sum(
                1 for idx in result.index
                if pd.notna(result[idx]) and
                datetime.strptime(str(result[idx]), '%Y%m%d') > _asd_cutoff
                if result[idx] and str(result[idx]).isdigit()
            )
            if asd_future_count > 0:
                logger.info(f"  - [ASD] {asd_future_count} dates remain beyond cutoff (kept as-is — unresolvable ambiguity)")

        # Assign back to dataframe
        df[col] = result
        
        # Print summary
        valid_count = df[col].notna().sum()
        if original_count > 0:
            logger.info(f"  - Result: {valid_count}/{original_count} valid dates ({valid_count/original_count*100:.1f}%)")
        else:
            logger.info(f"  - Result: {valid_count}/{original_count} valid dates (N/A - no records)")
    
    return df


def process_dates(df, cutoff_date=None):
    """Process date fields in the DataFrame with optimized vectorized operations
    
    Args:
        df: DataFrame to process
        cutoff_date: Optional datetime object for LOANEFFECTIVEDATE/ACCOUNTSTATUSDATE validation
    """
    # Use the optimized vectorized version
    return process_dates_vectorized(df, cutoff_date=cutoff_date)


def remove_special_chars(text):
    """Remove special characters from text while preserving spaces"""
    if not text:
        return ''
    
    # Convert to string if not already
    text = str(text)
    # Remove carriage returns and line feeds
    text = text.replace('\r', '').replace('\n', '')
    # Replace common punctuation with spaces
    text = re.sub(r'[.,\'"\-_/\\|&]', ' ', text)
    # Remove any remaining special characters but keep spaces
    text = re.sub(r'[^a-zA-Z\s]', '', text)
    # Replace multiple spaces with single space and strip
    text = ' '.join(text.split())
    
    return text.strip()

def clean_name_preserving_special_chars(text):
    """Clean names by replacing hyphens with spaces and removing all other special characters"""
    if not text:
        return ''
    
    # Convert to string if not already
    text = str(text)
    
    # Remove carriage returns and line feeds
    text = text.replace('\r', '').replace('\n', '')
    
    # First replace hyphens with spaces
    text = text.replace('-', ' ').replace("'", '')
    
    # Remove all other special characters
    text = re.sub(r'[^a-zA-Z0-9&]', ' ', text)
    
    # Replace multiple spaces with single space and strip
    text = ' '.join(text.split())
    
    return text.strip()

def clean_business_name(df):
    """
    Clean business name columns by removing purely numeric values.
    Values like '0', '000', '123', '0.00', '11111' are cleared.
    Business names should contain letters.
    
    Args:
        df: DataFrame to process
    Returns:
        DataFrame with cleaned business name columns
    """
    business_name_columns = ['BUSINESSNAME', 'SURNAME']
    
    def is_purely_numeric(value):
        """Check if value is purely numeric with 6 or fewer digits"""
        if not value or not isinstance(value, str):
            return False
        value = str(value).strip()
        # Remove decimal points and commas, check if remaining is all digits
        cleaned = value.replace('.', '').replace(',', '')
        if not cleaned:
            return False
        # If all remaining characters are digits AND 6 or fewer, clear it
        # Keep values with 7+ digits (could be registration/CAC numbers)
        return cleaned.isdigit() and len(cleaned) <= 6
    
    for col in business_name_columns:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: '' if is_purely_numeric(x) else x)
            logger.info(f"Cleaned purely numeric values from column: {col}")
    
    return df

def process_names(df):
    """Process names before column mapping"""
    if df is None or df.empty:
        return df
        
    name_groups = {
        'primary': ['SURNAME', 'FIRSTNAME', 'MIDDLENAME'],
        'spouse': ['SPOUSESURNAME', 'SPOUSEFIRSTNAME', 'SPOUSEMIDDLENAME'],
        'principal1': ['PRINCIPALOFFICER1SURNAME', 'PRINCIPALOFFICER1FIRSTNAME', 'PRINCIPALOFFICER1MIDDLENAME'],
        'principal2': ['PRINCIPALOFFICER2SURNAME', 'PRINCIPALOFFICER2FIRSTNAME', 'PRINCIPALOFFICER2MIDDLENAME'],
        'guarantor': ['INDIVIDUALGUARANTORSURNAME', 'INDIVIDUALGUARANTORFIRSTNAME', 'INDIVIDUALGUARNTORMIDDLENAME']
    }
    
    for group_name, name_columns in name_groups.items():
        if all(col in df.columns for col in name_columns):
            # Debug print
            logger.info(f"\nProcessing group: {group_name}")
            logger.info(f"Original columns:\n{df[name_columns].head()}")
            
            # Explicitly clean columns
            for col in name_columns:
                # Convert to string, replace NaN with empty string
                df[col] = df[col].apply(lambda x: '' if x is None or (isinstance(x, float) and pd.isna(x)) else str(x).strip())
            
            # Print after initial cleaning
            logger.info(f"After initial cleaning:\n{df[name_columns].head()}")
            
            # Remove titles and clean names while preserving special characters
            for col in name_columns:
                df[col] = df[col].apply(remove_titles).apply(clean_name_preserving_special_chars)
            
            # Clear purely numeric values from name columns (no letters = not a name)
            for col in name_columns:
                df[col] = df[col].apply(
                    lambda x: '' if x and re.sub(r'[^a-zA-Z]', '', str(x)) == '' else x
                )
            
            # Combine non-empty name components
            def combine_names(row):
                # Get the three name components
                surname = row[name_columns[0]]
                firstname = row[name_columns[1]]
                middlename = row[name_columns[2]]
                
                # Check if all 3 columns have the same value (duplicate full name scenario)
                # If so, only use the first one to avoid concatenation issues
                if surname == firstname == middlename and surname:
                    return surname  # Use only the first occurrence
                
                # Check if 2 columns are the same
                if surname == firstname and surname:
                    # Surname and firstname are same, use surname + middlename
                    name_components = [surname, middlename] if middlename else [surname]
                elif firstname == middlename and firstname:
                    # Firstname and middlename are same, use surname + firstname
                    name_components = [surname, firstname] if surname else [firstname]
                elif surname == middlename and surname:
                    # Surname and middlename are same, use surname + firstname
                    name_components = [surname, firstname] if firstname else [surname]
                else:
                    # Normal case: all different or some empty
                    name_components = [surname, firstname, middlename]
                
                # Remove empty strings and join
                name_components = [comp for comp in name_components if comp]
                return ' '.join(name_components)
            
            temp_full_name = f'FULL_NAME_{group_name}'
            df[temp_full_name] = df.apply(combine_names, axis=1)
            
            # Print combined names
            logger.info(f"Combined names:\n{df[temp_full_name].head()}")
            
            # Split the full name back into components
            name_parts = df[temp_full_name].apply(lambda x: pd.Series(x.split(maxsplit=2) + ['', '', ''])[:3])
            
            # Update original columns with processed parts
            df[name_columns[0]] = name_parts[0]
            df[name_columns[1]] = name_parts[1]
            df[name_columns[2]] = name_parts[2]
            
            # Print final processed columns
            logger.info(f"Final processed columns:\n{df[name_columns].head()}")
            
            # Drop the temporary column
            df = df.drop(temp_full_name, axis=1)
        else:
            # Process individual columns if the full group is not present
            for col in name_columns:
                if col in df.columns:
                    df[col] = df[col].apply(lambda x: '' if x is None or (isinstance(x, float) and pd.isna(x)) else str(x).strip())
                    df[col] = df[col].apply(remove_titles).apply(clean_name_preserving_special_chars)
                    # Clear purely numeric values (no letters = not a name)
                    df[col] = df[col].apply(
                        lambda x: '' if x and re.sub(r'[^a-zA-Z]', '', str(x)) == '' else x
                    )
    return df

def rename_columns_with_fuzzy_rapidfuzz(df, mapping, threshold=90):
    def fuzzy_match(column, alt_names):
        result = process.extractOne(column, alt_names, scorer=fuzz.token_set_ratio)
        if result and result[1] >= threshold:
            return result[0]
        return None

    # Track renamed columns to avoid conflicts
    renamed_columns = set()

    # Create a mapping to track which keys have been used
    used_keys_mapping = {key: None for key in mapping}

    # Collect columns to drop due to conflicts
    columns_to_drop = []

    # Iterate over the columns and rename them
    for column in list(df.columns):  # Use list to avoid issues when dropping columns
        found_match = False
        for mapped_column, alt_names in mapping.items():
            # Check if the key has been used
            if used_keys_mapping[mapped_column] is not None:
                continue

            # Check if the column name is in alt_names
            if column.lower() in alt_names or column.upper() in alt_names or column == mapped_column:
                # Check for key conflict: if mapped_column already exists in df.columns (and is not the current column)
                if mapped_column in df.columns and column != mapped_column:
                    columns_to_drop.append(column)
                    logger.info(f"Column {column} dropped due to key conflict with {mapped_column}.")
                else:
                    df.rename(columns={column: mapped_column}, inplace=True)
                    renamed_columns.add(mapped_column)
                    used_keys_mapping[mapped_column] = column
                    logger.info(f"Renamed {column} to {mapped_column}")
                found_match = True
                break

        # If no exact match found, try fuzzy matching
        if not found_match:
            fuzzy_match_result = fuzzy_match(column, mapping.keys())
            if fuzzy_match_result:
                # Check for key conflict: if fuzzy_match_result already exists in df.columns (and is not the current column)
                if fuzzy_match_result in df.columns and column != fuzzy_match_result:
                    columns_to_drop.append(column)
                    logger.info(f"Column {column} dropped due to key conflict with {fuzzy_match_result} (fuzzy match).")
                elif used_keys_mapping[fuzzy_match_result] is None:
                    df.rename(columns={column: fuzzy_match_result}, inplace=True)
                    renamed_columns.add(fuzzy_match_result)
                    used_keys_mapping[fuzzy_match_result] = column
                    logger.info(f"Fuzzy matched {column} to {fuzzy_match_result}")
                else:
                    columns_to_drop.append(column)
                    logger.info(f"Column {column} dropped due to key conflict (fuzzy match already used).")

    # Drop all columns that were marked for dropping
    if columns_to_drop:
        df.drop(columns=columns_to_drop, inplace=True, errors='ignore')

    # Drop duplicate columns
    df = df.loc[:, ~df.columns.duplicated()]

    # Add columns for keys that were not mapped
    new_columns = {key: None for key, used_column in used_keys_mapping.items() if used_column is None}
    df = pd.concat([df, pd.DataFrame(new_columns, index=df.index)], axis=1)

    # Ensure all mapping keys are present as columns
    for key in mapping.keys():
        if key not in df.columns:
            df[key] = None

    # Reorder the columns based on the keys in the dictionary
    df = df[list(mapping.keys())]

    return df

def fill_data_column(df):
    """
    Fill the 'DATA' column with 'D' after column renaming
    """
    if 'DATA' in df.columns:
        df['DATA'] = 'D'
    else:
        logger.info("===========================")
    return df

def fill_depend_column(df):
    """
    Fill the 'DEPENDANTS' column with '00' after column renaming
    """
    if 'DEPENDANTS' in df.columns:
        df['DEPENDANTS'] = df['DEPENDANTS'].apply(lambda x: '00' if pd.isna(x) or str(x).strip() in ['', 'None', 'nan', 'null', 'nill', 'nil', 'na', 'n/a'] else x)
    else:
        logger.info("\n=== DEPENDANTS COLUMN NOT FOUND ===") 
    return df

def process_gender(df):
    """Process gender fields in the DataFrame"""
    gender_columns = [
        'GENDER',
        'SPOUSEGENDER',
        'PRINCIPALOFFICER1GENDER',
        'PRINCIPALOFFICER2GENDER',
        'GUARANTORGENDER',
        'INDIVIDUALGUARANTORGENDER'
    ]
    
    for col in gender_columns:
        if col in df.columns:
            try:
                # Check if the column has any non-null values before processing
                if df[col].notna().any():
                    df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
                    df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
                    df[col] = df[col].apply(map_gender)
                else:
                    logger.info(f"No non-null values found in column '{col}'.")
            except Exception as e:
                logger.info(f"Error processing column '{col}': {e}")
    return df

def map_gender(value):
    """Maps gender values to standardized format"""
    if isinstance(value, pd.Series):  # Handle Series input
        return value.apply(map_gender)
    
    if pd.isna(value) or value is None:
        return None

    if not isinstance(value, str):
        value = str(value)

    value = value.lower().strip()
    
    if value in ['', 'none', 'nan', 'null', 'n/a']:
        return None

    for category, values in Gender_dict.items():
        if value in values:
            return category
    
    return None
def process_nationality(df):
    """Enhanced nationality processing with comprehensive error handling and .any() ambiguity resolution"""
    if df is None or df.empty:
        return df
    
    nationality_columns = [
        'NATIONALITY',
        'PRIMARYADDRESSCOUNTRY',
        'EMPLOYERCOUNTRY',
        'SECONDARYADDRESSCOUNTRY',
        'BUSINESSOFFICEADDRESSCOUNTRY',
        'PRINCIPALOFFICER1COUNTRY',
        'PRINCIPALOFFICER2COUNTRY',
        'GUARANTORPRIMARYADDRESSCOUNTRY',
    ]
    
    def clean_country_value(value):
        """Robust country value cleaning with detailed logging"""
        try:
            # Handle NaN or None values first
            if pd.isna(value) or value is None:
                return None        
            # Convert to string safely
            value = str(value).strip()          
            # Convert to lowercase and remove special characters
            value = value.lower()
            value = re.sub(r'[^a-zA-Z0-9\s]', '', value)           
            # Check for empty or invalid values
            if not value or value in ['none', 'nan', 'null', 'na']:
                return None          
            return value        
        except Exception as e:
            logger.info(f"Error cleaning country value '{value}': {e}")
            return None
    def standardize_country(value):
        """Enhanced country standardization with detailed logging"""
        if value is None:
            return None       
        try:
            for standard_name, variations in Country_dict.items():
                if value in [v.lower() for v in variations]:
                    return standard_name           
            return None       
        except Exception as e:
            logger.info(f"Error standardizing country '{value}': {e}")
            return None    
    # Find columns that exist in the DataFrame
    found_columns = [col for col in nationality_columns if col in df.columns]   
    for column in found_columns:        
        try:
            # Check if the column has any non-null values using .any()
            if df[column].notna().any():
                df[column] = df[column].apply(clean_country_value)
                df[column] = df[column].apply(standardize_country)              
            else:
                logger.info(f"SKIP: No non-null values in column {column}")  
        except Exception as column_e:
            logger.info(f"? FAILED to process column {column}: {column_e}")
            logger.info(traceback.format_exc())
    return df


def remove_spaces(text):
    """Remove spaces from the input string."""
    if text is None:
        return ""
    return str(text).replace(" ", "")

    
def process_special_characters(df):
    """Remove special characters from all columns except specified ones, preserving '&' in address columns"""
    if df is None or df.empty:
        return df
    
    # List of columns to exclude from special character removal
    excluded_columns = [
        'DATEOFBIRTH',
        'DATEOFINCORPORATION',
        'PRINCIPALOFFICER1DATEOFBIRTH',
        'PRINCIPALOFFICER2DATEOFBIRTH',
        'SPOUSEDATEOFBIRTH',
        'GUARANTORDATEOFBIRTHINCORPORATION',
        'LOANEFFECTIVEDATE',
        'MATURITYDATE',
        'LASTPAYMENTDATE',
        'DEFEREDPAYMENTDATE',
        'LITIGATIONDATE',
        'FACILITYTYPE',
        'BRANCHCODE',
        'BRANCH CODE',
        'CUSTOMERBRANCHUCODE',
        'CUSTOMERBRANCHCODE',
        'EMAIL',
        'EMAILADDRESS',
        'PRINCIPALOFFICER1EMAILADDRESS',
        'PRINCIPALOFFICER2EMAILADDRESS',
        'GUARANTOREMAIL',
        'OUTSTANDINGBALANCE',
        'MONTHLYREPAYMENT',
        'TOTALREPAYMENT',
        'CREDITLIMIT',
        'AVAILEDLIMIT',
        'OUTSTANDINGBALANCE',
        'CURRENTBALANCEDEBT',
        'INSTALMENTAMOUNT',
        'OVERDUEAMOUNT',
        'LASTPAYMENTAMOUNT',
        'ACCOUNTSTATUSDATE',
        'SURNAME',
        'FIRSTNAME',
        'MIDDLENAME',
        'INDIVIDUALGUARANTORSURNAME',
        'INDIVIDUALGUARANTORFIRSTNAME',
        'INDIVIDUALGUARANTORMIDDLENAME',
        'PRINCIPALOFFICERSURNAME',
        'PRINCIPALOFFICERFIRSTNAME',
        'PRINCIPALOFFICERMIDDLENAME',
        'CURRENCY',
    ]

    # List of columns that should preserve '&'
    address_columns = [
        'PRIMARYADDRESSLINE1',
        'PRIMARYADDRESSLINE2',
        'SECONDARYADDRESSLINE1',
        'SECONDARYADDRESSLINE2',
        'BUSINESSOFFICEADDRESSLINE1',
        'BUSINESSOFFICEADDRESSLINE2',
        'GUARANTORPRIMARYADDRESSLINE1',
        'GUARANTORPRIMARYADDRESSLINE2',
        'PRINCIPALOFFICER1PRIMARYADDRESSLINE1',
        'PRINCIPALOFFICER1PRIMARYADDRESSLINE2',
        'PRINCIPALOFFICER2PRIMARYADDRESSLINE1',
        'PRINCIPALOFFICER2PRIMARYADDRESSLINE2',
        'SECONDARYADDRESSCITYLGA',
        'BUSINESSOFFICEADDRESSCITYLGA',
        'GUARANTORPRIMARYADDRESSCITYLGA',
        'PRINCIPALOFFICER1CITY',
        'PRINCIPALOFFICER2CITY',
        'PRIMARYADDRESSCITY',
        'COLLATERALDETAILS',
        'BUSINESSNAME',
        'BUSINESSCATEGORY'
    ]
    
    # Account number columns that should preserve '/' and '-'
    account_number_columns = [
        'ACCOUNTNUMBER',
        'CUSTOMERSACCOUNTNUMBER'
    ]

    # Find processable columns (those not in excluded list)
    processable_columns = [col for col in df.columns if col not in excluded_columns]
    
    for column in processable_columns:
        # Safely apply the transformation
        try:
            # Check if the column has any non-null values before processing
            if df[column].notna().any():
                if column in account_number_columns:
                    # Special handling for account numbers - keep '/' and '-'
                    df[column] = df[column].apply(
                        lambda x: re.sub(r'[^a-zA-Z0-9/\-]', '', str(x)) if pd.notnull(x) else x
                    )
                elif column in address_columns:
                    # Keep '&' in address columns
                    df[column] = df[column].apply(
                        lambda x: re.sub(r'[^a-zA-Z0-9&]', ' ', str(x)) if pd.notnull(x) else x
                    )
                else:
                    df[column] = df[column].apply(
                        lambda x: re.sub(r'[^a-zA-Z0-9]', ' ', str(x)) if pd.notnull(x) else x
                    )
                # Remove double spaces
                df[column] = df[column].apply(lambda x: re.sub(r'\s+', ' ', x).strip() if isinstance(x, str) else x)
        except Exception as e:
            logger.info(f"Error processing column {column}: {e}")

    # Now handle specific columns to remove spaces
    # ------------------------------------------------# take notr of this.-------------------------------------------------------
    for col in ['CUSTOMERID', 'TAXID', 'OTHERID','LEGALCHALLENGESTATUS','LOANSECURITYSTATUS','ACCOUNTSTATUS']:
        if col in df.columns:
            df[col] = df[col].apply(remove_spaces)

    # Updated email processing logic
    email_columns = [
        'EMAILADDRESS', 
        'PRINCIPALOFFICER1EMAILADDRESS',
        'PRINCIPALOFFICER2EMAILADDRESS', 
        'GUARANTOREMAIL'
    ]
    
    for col in email_columns:
        if col in df.columns:
            try:
                # Convert to lowercase and filter valid emails
                df[col] = df[col].str.lower()
                df[col] = df[col].apply(
                    lambda x: x if pd.notnull(x) and (
                        x.endswith('@gmail.com') or 
                        x.endswith('@yahoo.com')) or 
                        # x.endswith('.co.uk')) or 
                        x.endswith('.com') or
                        x.endswith('.ng') or
                        x.endswith('.net') or
                        x.endswith('.org') or
                        x.endswith('.biz') or
                        x.endswith('.info') or
                        # x.endswith('.co.uk') or
                        x.endswith('.us')
                else ''
                )
            except Exception as e:
                logger.info(f"Error processing email column {col}: {e}")
    
    return df

def replace_ampersands(df):
    """
    Replace all instances of '&' with 'And' across all string columns in the DataFrame
    """
    # Remove duplicate columns to avoid DataFrame return from df[column]
    df = df.loc[:, ~df.columns.duplicated()]
    for column in df.columns:
        # Only process object (string) columns
        if df[column].dtype == 'object':
            df[column] = df[column].apply(
                lambda x: str(x).replace('&', ' And ') if pd.notna(x) else x
            )
    logger.info("Replaced '&' with 'And' across all string columns")
    return df

def process_passport_number(df):
    """
    Cleans the Passport Number column based on specified criteria.
    
    Parameters:
    df (pd.DataFrame): The input DataFrame.
    
    Returns:
    pd.DataFrame: The updated DataFrame with valid Passport Numbers retained.
    """
    # List of Passport Number columns to process
    passport_columns = ['PASSPORTNUMBER',
                        'PRINCIPALOFFICER1PASSPORTNUMBER',
                        'PRINCIPALOFFICER2PASSPORTNUMBER',
                        'GUARNATORINTLPASSPORTNUMBER']  # You can add more columns to this list if needed
    
    for column_name in passport_columns:
        if column_name in df.columns:
            # Function to clean Passport Number-
            def clean_passport(passport):
                # Convert to string
                passport_str = str(passport)
                passport_str = re.sub(r'[^a-zA-Z0-9]', '', passport_str)
                # Check if the value is numeric
                if passport_str.isdigit():
                    return ''  # Remove if purely numeric
                # Discard if the cleaned ID is not exactly 11 or 10 characters
                if len(passport_str) not in [9,10,11]:
                    return ''
                return passport_str  # Keep alphanumeric values

            # Apply the cleaning function to the PASSPORT_NUMBER column
            df[column_name] = df[column_name].apply(clean_passport)

    return df


def process_identity_numbers(df):
    """
    Cleans the National Identity Number columns based on specified criteria.
    
    Updated Criteria:
    - If an ID starts with "NIN" and is followed by numbers, the "NIN" is removed and the numbers are kept.
    - An ID that is exactly 11 numeric digits (and not a repetitive pattern) is kept.
    - An alphanumeric ID must be 10 or 11 characters long and start with two letters followed by a digit.
    
    Parameters:
        df (pd.DataFrame): The input DataFrame.
    
    Returns:
        pd.DataFrame: The updated DataFrame with valid National Identity Numbers retained.
    """
    
    # List of National Identity Number columns to process
    identity_columns = [
        'NATIONALIDENTITYNUMBER',  
        'PRINCIPALOFFICER1NATIONALID',
        'PRINCIPALOFFICER2NATIONALID',
        'GUARANTORNATIONALIDNUMBER',
    ]
    
    for column_name in identity_columns:
        if column_name in df.columns:
            def clean_identity(identity):
                # Convert the value to a string
                identity_str = str(identity)
                # Remove all non-alphanumeric characters
                identity_str = re.sub(r'[^a-zA-Z0-9]', '', identity_str)

                # NEW Check 1: Check for 'NIN' followed only by digits.
                if re.match(r'^NIN\d+$', identity_str, re.IGNORECASE):
                    # If it matches, remove the first 3 characters ('NIN') and return the rest.
                    return identity_str[3:] 

                if len(identity_str) not in [10, 11]:
                    return ''

                if identity_str.isdigit() and len(identity_str) == 11:
                    if len(set(identity_str)) > 1:
                        return identity_str
                    return ''
                
                
                if not re.match(r'^[a-zA-Z]{2}\d', identity_str):
                    return ''
                
                return identity_str
            
            df[column_name] = df[column_name].apply(clean_identity)
    
    return df

def process_tax_numbers(df):

    
    # List of National Identity Number columns to process
    identity_columns = [
            'TAXID'
    ]
    
    for column_name in identity_columns:
        if column_name in df.columns:
            def clean_identity(identity):
                # Convert the value to a string
                identity_str = str(identity)
                # Remove all non-alphanumeric characters (i.e., spaces and special characters)
                identity_str = re.sub(r'[^a-zA-Z0-9]', '', identity_str)
                
                if identity_str.isdigit():
                    # Check if it's not a repetitive pattern (all same digit)
                    if len(set(identity_str)) > 1:  # More than one unique digit
                        return identity_str
                    return ''  # Repetitive numeric pattern, discard
                
                # Discard if the cleaned ID is not exactly 10 or 11 characters
                if len(identity_str) not in [9, 10, 11]:
                    return ''
                
                #  Check that the ID starts with two letters followed immediately by a digit
                # if not re.match(r'^[a-zA-Z]{1}\d', identity_str):
                #     return ''
                
                return identity_str
            
            df[column_name] = df[column_name].apply(clean_identity)
    
    return df


def process_DriversLicense(df):
    """
    Cleans the Driver's License columns based on specified criteria.
    
    Updated Criteria:
    - A license number starting with "LNO" (case-insensitive) is always kept, regardless of length.
    - For all other formats, the value must be 10, 11, or 12 characters long after cleaning.
    - Other formats must also start with three letters (case insensitive) immediately followed by a digit.
    
    Parameters:
        df (pd.DataFrame): The input DataFrame.
    
    Returns:
        pd.DataFrame: The updated DataFrame with valid Driver's License values retained.
    """
    
    # List of Driver's License columns to process
    dLicense = [ 'DRIVERSLICENSENUMBER',
            'PRINCIPALOFFICER1DRIVERSLISCENCENUMBER',
            'PRINCIPALOFFICER2DRIVERSLISCENCENUMBER',
            'GUARANTORDRIVERSLICENCENUMBER']  # You can add more columns to this list if needed
    
    for column_name in dLicense:
        if column_name in df.columns:
            def clean_driversLicense(value):
                # Convert the value to a string
                value_str = str(value)
                # Remove all non-alphanumeric characters (i.e., spaces and special characters)
                value_str = re.sub(r'[^a-zA-Z0-9]', '', value_str)
                
                # NEW: Add special check for license numbers starting with 'LNO'.
                # This check is case-insensitive. If it matches, we keep the value and ignore other rules.
                if value_str.upper().startswith('LNO'):
                    return value_str
                
                # Check 1: Length must be 10, 11, or 12 characters.
                if len(value_str) not in [10, 11, 12]:
                    return ''
                
                # Check 2: Must start with three letters followed by a digit.
                if not re.match(r'^[a-zA-Z]{3}\d', value_str):
                    return ''
                
                return value_str
            
            # Apply the cleaning function to the Driver's License column
            df[column_name] = df[column_name].apply(clean_driversLicense)
    
    return df


def process_business_id(df):
    """
    Clears the values in the specified column where the values are not alphanumeric
    (containing both letters and numbers).
    
    Parameters:
    df (pd.DataFrame): The input DataFrame.
    column_name (str): The name of the column to process.
    
    Returns:
    pd.DataFrame: The updated DataFrame with non-alphanumeric values cleared in the specified column.
    """
    column_name = [
        'BUSINESSREGISTRATIONNUMBER',
        # Add any other relevant column names that may appear
    ]
    for col in column_name:
        if col in df.columns:
            # Convert to string and remove spaces and special characters
            df[col] = df[col].astype(str).apply(
                lambda x: ''.join(char for char in x if char.isalnum())
            )
            
            # Keep only values that start with "RN", "BC", or "BN" (case-insensitive)
            df[col] = df[col].where(
                # df[col].str.contains(r'(?=.*[a-zA-Z])(?=.*\d)', regex=True), 
                df[col].str.match(r'^(rn|bc|bn|rc)', case=False).fillna(False),
                ''
            )
            
            # Replace 'nan' or 'None' with empty string
            df[col] = df[col].replace({'nan': '', 'None': ''})
    
    return df

def process_bvn_number(df):
    """
    Cleans the BVN number columns based on specified criteria.
    
    Parameters:
    df (pd.DataFrame): The input DataFrame.
    
    Returns:
    pd.DataFrame: The updated DataFrame with valid BVN values retained.
    """
    # List of BVN columns to process
    bvn_columns = ['BVNNUMBER',
                   'PRINCIPALOFFFICER1BVNNUMBER',
                   'PRINCIPALOFFICER2BVNNUMBER',
                   'GUARANTORBVN']  # You can add more columns to this list if needed
    
    for column_name in bvn_columns:
        if column_name in df.columns:
            # Function to clean BVN number
            def clean_bvn(bvn):
                # Convert to string
                bvn_str = str(bvn)
                # Check if the length is 11 and if it's numeric
                if len(bvn_str) == 11 and bvn_str.isdigit():
                    # Check 1: Remove if all characters are identical (e.g., "22222222222")
                    # This is still a useful check for junk data.
                    if bvn_str == bvn_str[0] * 11:
                        return ''
                    
                    # UPDATED Check 2: Enforce the rule that valid BVNs must be 20,000,000,000 or higher.
                    # This will remove any BVN starting with '0' or '1'.
                    if int(bvn_str) < 20000000000:
                        return ''

                    return bvn_str  # Keep the valid BVN
                return ''  # Remove if not 11 digits or not numeric
            
            # Apply the cleaning function to the BVNNUMBER column
            df[column_name] = df[column_name].apply(clean_bvn)

    return df
# ---------------------------------------------------------------REMODIFY THIS---------------------------------------------------------------------
def process_otherid(df):
    """
    Cleans the National Identity Number columns based on specified criteria.
    
    Updated Criteria:
    - Each ID must be exactly 11 characters long. If the cleaned ID is not exactly 11 characters, it is discarded.
    - Each ID must start with two letters (case insensitive) immediately followed by a digit.
      If the starting pattern is not met, the ID is discarded.
    
    Parameters:
        df (pd.DataFrame): The input DataFrame.
    
    Returns:
        pd.DataFrame: The updated DataFrame with valid National Identity Numbers retained.
    """
    
    # List of Other Identity Number columns to process
    otherid_columns = [
       'OTHERID',
       'PRINCIPALOFFICER1OTHERID',
       'PRINCIPALOFFICER2OTHERID',
       'GUARANTOROTHERID'
    ]
    
    for column_name in otherid_columns:
        if column_name in df.columns:
            def clean_otherid(other):
                # Convert the value to a string
                other_str = str(other)
                # Remove all non-alphanumeric characters (i.e., spaces and special characters)
                other_str = re.sub(r'[^a-zA-Z0-9]', '', other_str)
                
                # Discard if the cleaned ID is not exactly 10 or 11 characters
                if len(other_str) not in [10, 11]:
                    return ''
                
                # Check that the ID starts with one letters followed immediately by a digit.
                if not re.match(r'^[a-zA-Z]{1}\d', other_str):
                    return ''
                
                return other_str
            
            df[column_name] = df[column_name].apply(clean_otherid)
    
    return df



# Define the state columns
state_columns = [
    'STATE', 
    'PRIMARYADDRESSSTATE', 
    'SECONDARYADDRESSSTATE', 
    'EMPLOYERSTATE', 
    'BUSINESSOFFICEADDRESSSTATE', 
    'GUARANTORPRIMARYADDRESSSTATE', 
    'PRINCIPALOFFICER1STATE', 
    'PRINCIPALOFFICER2STATE'
]
# Define a function to perform fuzzy mapping
def fuzzy_map_state(state_name, state_dict):
    # Check if the state_name is empty or contains only whitespace
    if not state_name.strip():
        return None

    max_score = -1
    matched_state = None

    # Iterate through the state_dict and calculate fuzz ratio
    for state_code, names in state_dict.items():
        for name in names:
            score = fuzz.ratio(state_name.lower(), name.lower())
            if score > max_score:
                max_score = score
                matched_state = state_code

    # Define a threshold score (you can adjust this based on your requirements)
    threshold_score = 98

    # If the similarity score is above the threshold, return the corresponding state code
    if max_score >= threshold_score:
        return matched_state
    else:
        return None  # Return None if no good match is found

# Function to process state columns in the DataFrame
def process_states(consu):
    """Process state fields in the DataFrame"""
    for column in state_columns:
        if column in consu.columns and consu[column].apply(lambda x: not pd.isna(x) and str(x).strip() != '').any():
            # Clean and preprocess the column
            consu[column] = consu[column].apply(lambda x: str(x) if not pd.isna(x) else None)
            # Apply the fuzzy mapping function to non-empty values
            consu[column] = consu[column].apply(lambda x: fuzzy_map_state(x, state_dict) if not pd.isna(x) and str(x).strip() != '' else None)
        else:
            # No non-empty values found in the column, no action required
            pass
    return consu

def map_marital(value):
    if isinstance(value, str):
        for category, values in Marital_dict.items():
            if value in values:
                return category
    return None

def process_marital_status(df):
    """Process marital status fields in the DataFrame"""
    # Define the marital status columns to look for
    marital_columns = [
        'MARITALSTATUS',
        # Add any other relevant column names that may appear
    ]
    
    # Iterate through the list of potential marital status columns
    for col in marital_columns:
        if col in df.columns:
            # Clean the marital status values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(map_marital)
    
    return df

def map_borrowert(value):
    if isinstance(value, str):
        for category, values in Borrower_dict.items():
            if value in values:
                return category
    return None

def process_borrower_type(df):
    """Process borrower type fields in the DataFrame"""
    # Define the borrower type columns to look for
    borrower_columns = [
        'BORROWERTYPE'
        # Add any other relevant column names that may appear
    ]
    
    # Iterate through the list of potential borrower type columns
    for col in borrower_columns:
        if col in df.columns:
            # Clean the borrower type values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(map_borrowert)
    
    return df

def map_employers(value):
    if isinstance(value, str):
        for category, values in Employer_dict.items():
            if value in values:
                return category
    return None

def process_employment_status(df):
    """Process employment status fields in the DataFrame"""
    # Define the employment status columns to look for
    employment_columns = [
        'EMPLOYMENTSTATUS'
        # Add any other relevant column names that may appear
    ]
    
    # Iterate through the list of potential employment status columns
    for col in employment_columns:
        if col in df.columns:
            # Clean the employment status values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(map_employers)
    
    return df

def map_title(value):
    if isinstance(value, str):
        for category, values in Title_dict.items():
            if value in values:
                return category
    return None

def process_title(df):
    """Process title fields in the DataFrame"""
    # Define the title columns to look for
    title_columns = [
        'TITLE'
        # Add any other relevant column names that may appear
    ]
    
    # Iterate through the list of potential title columns
    for col in title_columns:
        if col in df.columns:
            # Clean the title values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(map_title)
    
    return df

def occu_title(value):
    if isinstance(value, str):
        for category, values in Occu_dict.items():
            if value in values:
                return category
        # If no match, check if the value is numeric
        if value.isdigit():
            return None  # Return None for numeric values
        # If the value is alphabetic, return it unchanged
        if value.isalpha():
            return value
    return None  # Return None for non-string types or unmatched cases

def process_occu(df):
    """Process title fields in the DataFrame"""
    # Define the title columns to look for
    occu_columns = [
        'OCCUPATION',
    ]
    
    # Iterate through the list of potential title columns
    for col in occu_columns:
        if col in df.columns:
            # Clean the title values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(occu_title)
    
    return df

def map_poistioninBusiness(value):
    """Maps account status values to standardized format."""
    if pd.isna(value) or value is None:
        return None
    
    # Convert to string and clean
    value = str(value).lower()
    value = re.sub(r'[^a-zA-Z0-9]', '', value)
    
    for category, values in Positioninbusiness_dict.items():
        # Convert dictionary values to lowercase and remove special characters for comparison
        dict_values = [str(v).lower().replace(r'[^a-zA-Z0-9]', '') for v in values]
        if value in dict_values:
            return category
    return None  # Return None if no match is found

def positioninBusiness(df):
    """Process account status fields in the DataFrame."""
    # Define the account status columns to look for
    status_columns = [
        'PRINCIPALOFFICER1POSITIONINBUSINESS',
        'PRINCIPALOFFICER2POSITIONINBUSINESS', 

    ]

    # Iterate through the list of potential account status columns
    for col in status_columns:
        if col in df.columns:
            logger.info(f"Processing account status column: {col}")
            
            # Clean the account status values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(map_poistioninBusiness)
            
            # Print unique values after processing
            logger.info(f"Unique values in {col} after processing: {df[col].unique()}")
    
    return df

def clear_previous_info_columns(df):
    """
    Clear the contents of previous information columns while keeping headers
    """
    columns_to_clear = [
        'PREVIOUSACCOUNTNUMBER',
        'PREVIOUSNAME',
        'PREVIOUSCUSTOMERID',
        'PREVIOUSBRANCHCODE',
        'BUSINESSSECTOR',
        'PICTUREFILEPATH'
    ]
    
    logger.info("\n=== CLEARING PREVIOUS INFO COLUMNS ===")
    for col in columns_to_clear:
        if col in df.columns:
            df[col] = ''
    logger.info("Previous info columns cleared")  
    return df
def map_accountStatus(value):
    """Maps account status values to standardized format."""
    if pd.isna(value) or value is None:
        return None
    
    # Convert to string and clean
    value = str(value).lower()
    value = re.sub(r'[^a-zA-Z0-9]', '', value)
    
    for category, values in AccountStatus_dict.items():
        # Convert dictionary values to lowercase and remove special characters for comparison
        dict_values = [re.sub(r'[^a-zA-Z0-9]', '', str(v).lower()) for v in values]
        if value in dict_values:
            return category
    return None  # Return None if no match is found

def process_account_status(df):
    """Process account status fields in the DataFrame."""
    # Define the account status columns to look for
    status_columns = [
        'ACCOUNTSTATUS',
        'STATUS', 

    ]

    # Iterate through the list of potential account status columns
    for col in status_columns:
        if col in df.columns:
            logger.info(f"Processing account status column: {col}")
            
            # Clean the account status values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(map_accountStatus)
            
            # Print unique values after processing
            logger.info(f"Unique values in {col} after processing: {df[col].unique()}")
    
    return df

def exact_map_loan(loan_name):
    # Clean the input
    loan_name_clean = re.sub(r'[^a-zA-Z0-9]', '', str(loan_name)).lower()
    # Clean and compare each dictionary value
    for loan_code, names in Loan_dict.items():
        for name in names:
            name_clean = re.sub(r'[^a-zA-Z0-9]', '', str(name)).lower()
            if loan_name_clean == name_clean:
                return loan_code
    return None

def process_loan_type(df):
    """Process business sector fields in the DataFrame"""
    # Define the business sector columns to look for
    loan_columns = [
        'FACILITYTYPE',
        # Add any other relevant column names that may appear
    ]
    
    # Iterate through the list of potential business sector columns
    for col in loan_columns:
        if col in df.columns:
            # Clean the business sector values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)
            df[col] = df[col].apply(exact_map_loan)
    
    return df
def map_currency(value):
    """Maps currency values to standardized format.
    
    Replaces currency symbols with their codes before cleaning,
    then matches against Currency_dict.
    """
    original_value = value  # For logging
    
    if pd.isna(value) or value is None:
        return None
    
    value_str = str(value).strip()
    
    if not value_str:
        return None
    
    # FIRST: Replace currency symbols with their codes BEFORE any cleaning
    # This ensures symbols like ₦, $, £, € are converted to matchable text
    symbol_replacements = {'₦': 'NGN', '$': 'USD', '£': 'GBP', '€': 'EUR'}
    for symbol, code in symbol_replacements.items():
        if symbol in value_str:
            value_str = code  # Replace entire value with the code
            logger.info(f"CURRENCY: '{original_value}' -> symbol '{symbol}' found, replaced with '{code}'")
            break
    
    # SECOND: Clean and match against Currency_dict
    value_clean = re.sub(r'[^a-zA-Z0-9]', '', value_str.lower())
    
    if not value_clean:
        logger.info(f"CURRENCY: '{original_value}' -> None (empty after cleaning)")
        return None
    
    for category, values in Currency_dict.items():
        # Convert dictionary values to lowercase and remove special characters for comparison
        dict_values = [re.sub(r'[^a-zA-Z0-9]', '', str(v).lower()) for v in values]
        if value_clean in dict_values:
            logger.info(f"CURRENCY: '{original_value}' -> '{category}' (matched '{value_clean}')")
            return category
    
    logger.info(f"CURRENCY: '{original_value}' -> None (no match for '{value_clean}')")
    return None

def process_currency(df):
    """Process currency fields in the DataFrame."""
    currency_columns = ['CURRENCY']
    
    for col in currency_columns:
        if col in df.columns:
            # Log RAW values BEFORE any processing
            raw_unique = df[col].dropna().unique()[:10]  # First 10 unique values
            logger.info(f"CURRENCY RAW VALUES BEFORE processing: {list(raw_unique)}")
            # Also log the repr to see hidden Unicode
            for val in raw_unique[:5]:
                logger.info(f"CURRENCY RAW repr: {repr(val)}")
            
            logger.info(f"Processing currency column: {col}")
            
            # Apply map_currency directly - it handles symbol replacement internally
            df[col] = df[col].apply(map_currency)
            
            # Print unique values after processing
            logger.info(f"Unique values in {col} after processing: {df[col].unique()}")
    
    return df

def map_repayment(value):
    """Maps repayment values to standardized format."""
    for category, values in Repayment_dict.items():
        if value in values:
            return category
    return None  # Return None if no match is found

def process_repayment(df):
    """Process repayment fields in the DataFrame."""
    repayment_columns = ['REPAYMENTFREQUENCY']  # Define the repayment columns to look for
    
    for col in repayment_columns:
        if col in df.columns:
            # Clean the repayment values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9\s]', '', x) if isinstance(x, str) else x)  # Allow spaces
            df[col] = df[col].apply(map_repayment)
    
    return df
def map_collateraltype(value):
    for category, values in Collateraltype_dict.items():
        if value in values:
            return category
    return None

def process_collateral_type(df):
    """Process collateral type fields in the DataFrame."""
    collateral_columns = ['COLLATERALTYPE']  # Define the collateral type columns to look for
    
    for col in collateral_columns:
        if col in df.columns:
            # Clean the collateral type values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9]', '', x) if isinstance(x, str) else x)  # Allow spaces
            df[col] = df[col].apply(map_collateraltype)
    
    return df
def map_classification(value):
    """Maps classification values to standardized format."""
    if pd.isna(value) or value is None:
        return None  # Return None for NaN or None values

    if not isinstance(value, str):
        value = str(value)  # Convert to string if not already

    # Check against the Classification_dict
    for category, values in Classification_dict.items():
        if value in values:
            return category  # Return the matched category

    return None 
def process_classification(df):
    """Process classification fields in the DataFrame."""
    classification_columns = ['LOANCLASSIFICATION']  # Define the classification columns to look for
    
    for col in classification_columns:
        if col in df.columns:
            # Clean the classification values
            df[col] = df[col].apply(lambda x: x.lower() if isinstance(x, str) else x)
            df[col] = df[col].apply(lambda x: re.sub(r'[^a-zA-Z0-9\s]', '', x) if isinstance(x, str) else x)  # Allow spaces
            df[col] = df[col].apply(map_classification)  # Apply the mapping function
    
    return df

def process_phone_columns(df):
    """
    Process numeric columns including telephone numbers
    """
    # Define columns that need numeric processing
    phone_columns = [
        'MOBILENUMBER', 'WORKTELEPHONE', 'HOMETELEPHONE', 
        'PRIMARYPHONENUMBER', 'SECONDARYPHONENUMBER',
        'PRINCIPALOFFICER1PHONENUMBER', 'PRINCIPALOFFICER2PHONENUMBER',
        'GUARANTORPRIMARYPHONENUMBER'
    ]
    
    try:
        if df is not None and not df.empty:
            # Process phone number columns
            for col in phone_columns:
                if col in df.columns:
                    logger.info(f"Processing phone number column: {col}")
                    df[col] = df[col].astype(str)
                    
                    # First extract only digits from the string, keeping spaces to separate numbers
                    df[col] = df[col].apply(lambda x: ''.join(char if char.isdigit() or char in [',', ';', '/', '|', '-', ' '] else ' ' for char in str(x)))
                    
                    # Split on any non-digit character and take the first non-empty number
                    df[col] = df[col].apply(lambda x: next((num.strip() for num in re.split(r'\D+', x) if num.strip()), ''))
                    
                     # Pad with zeros if less than 11 digits
                    # df[col] = df[col].apply(lambda x: x.zfill(11) if x and len(x) < 11 else x)
                    # Pad with zeros at the BEGINNING if less than 11 digits
                    df[col] = df[col].apply(lambda x: x.rjust(11, '0') if x and len(x) < 11 else x)
                    
                    # New validation: Check if number > 11 digits and doesn't start with 234
                    df[col] = df[col].apply(lambda x: '' if len(x) > 11 and not x.startswith('234') else x)

                    # Remove numbers that are more than 14 characters
                    df[col] = df[col].apply(lambda x: x if len(x) <= 13 else '')

                    # New validation: Check for more than 5 consecutive same digits
                    def has_repeating_sequence(number):
                        if not number:
                            return False
                        count = 1
                        prev_digit = number[0]
                        for digit in number[1:]:
                            if digit == prev_digit:
                                count += 1
                                if count > 5:
                                    return True
                            else:
                                count = 1
                                prev_digit = digit
                        return False
                    
                    df[col] = df[col].apply(lambda x: '' if has_repeating_sequence(x) else x)
                    
                    # Remove repetitive numbers (e.g., 00000000000, 11111111111)
                    df[col] = df[col].apply(lambda x: '' if x and len(set(x)) == 1 else x)
                    # Replace 'nan' with empty string
                    df[col] = df[col].replace({'nan': ''})
    
    except Exception as e:
        logger.info(f"Error in process_phone_columns: {e}")
        traceback.print_exc()
    
    return df

def convert_tenor_to_days(tenor: Union[str, int, float]) -> Optional[int]:
    """Converts a composite tenor string (e.g., '2 month 3 weeks') to a total number of days.

    If the input is a number (without unit), it returns that number as an integer.
    It handles multiple number-unit pairs by summing their respective day conversions.
    
    Supported units (case-insensitive):
        - days/d or day
        - weeks/w or week
        - months/m or month
        - years/y or year
    """
    if tenor is None or tenor == '':
        return None

    # If the input is already numeric, return it as integer
    if isinstance(tenor, (int, float)):
        return int(tenor)

    # Convert to string and normalize to lower-case
    tenor_str = str(tenor).strip().lower()

    # Optional: Convert written-out numbers (like "two", "three") to digits using w2n.
    try:
        tenor_str = re.sub(
            r'\b(one|two|three|four|five|six|seven|eight|nine|ten|eleven|twelve)\b',
            lambda m: str(w2n.word_to_num(m.group())), tenor_str)
    except Exception as ex:
        # If conversion fails, just proceed with the original string.
        pass

    # Define a regex pattern that finds multiple number-unit pairs
    pattern = r'(\d+(?:\.\d+)?)\s*(days?|weeks?|months?|years?|d|w|m|y)'
    matches = re.findall(pattern, tenor_str)

    total_days = 0
    if matches:
        # Define mapping between recognized units and their day multipliers
        unit_mapping = {
            'day': 1,
            'days': 1,
            'd': 1,
            'dys':1,
            'week': 7,
            'weeks': 7,
            'w': 7,
            'wk':7,
            'wks':7,
            'month': 30,
            'mnth':30,
            'mth':30,
            'mths':30,
            'mnths':30,
            'mons':30,
            'months': 30,
            'm': 30,
            'year': 365,
            'years': 365,
            'y': 365,
            'yr': 365,
            'yrs': 365,
        }
        for num_str, unit in matches:
            try:
                number = float(num_str)
            except ValueError:
                continue  # Skip if conversion fails
            multiplier = unit_mapping.get(unit, None)
            if multiplier is not None:
                total_days += number * multiplier
        # Return total days as an integer
        return int(total_days)
    else:
        # Fallback: If no unit-pattern was found, try converting the whole string to a number
        try:
            return int(float(tenor_str))
        except ValueError:
            return None
def process_loan_tenor(df):
    """
    Process loan tenor column in the DataFrame.
    Args:
        df: Input DataFrame
    Returns:
        DataFrame with processed loan tenor
    """
    if df is None:
        logger.info("Input DataFrame is None.")
        return None

    if not isinstance(df, pd.DataFrame):
        logger.info("Input is not a valid DataFrame.")
        return None

    # Columns to process for loan tenor
    tenor_columns = [ 'FACILITYTENOR',
                    #  'DAYSINARREARS'
                     ]

    # Process each potential tenor column
    for col in tenor_columns:
        if col in df.columns:
            logger.info(f"Processing column: {col}")

            # Apply conversion

            df[col] = df[col].apply(convert_tenor_to_days)
            # Convert to numeric, handling any conversion errors
            numeric_series = pd.to_numeric(df[col], errors='coerce')
            # Replace NaN with empty string, otherwise convert to int then string
            df[col] = numeric_series.apply(lambda x: '' if pd.isna(x) else str(int(np.ceil(x))))
        else:
            logger.info(f"Column {col} not found in DataFrame.")

    return df

def try_convert_to_float(x):
    """
    Enhanced numeric conversion function to handle mixed alphanumeric values
    and remove date-like strings.
    
    Args:
        x: Input value to convert
    
    Returns:
        Converted float as string if successful, otherwise returns cleaned value with commas removed
    """
    # If input is None or empty, return empty string
    if pd.isna(x) or x == '':
        return ''
    
    # Convert to string if not already and strip leading/trailing spaces
    x = str(x).strip()

    # --- NEW: Check for common date string formats ---
    # This regex looks for patterns like YYYY-MM-DD, DD/MM/YYYY, MM.DD.YY etc.
    # It's designed to catch strings that are clearly dates and not ambiguous numbers.
    # It uses start (^) and end ($) anchors to match the entire string.
    date_pattern = re.compile(
        r'^(?:\d{4}[-/]\d{1,2}[-/]\d{1,2})|(?:\d{1,2}[-/]\d{1,2}[-/]\d{2,4})$'
    )
    if date_pattern.match(x):
        return ''  # It's a date string, so clear it.
    
    # Remove common currency symbols/words before checking for letters
    # Examples handled: N233,000 | NGN 15,500 | $1,200.50 | £3,000 | €2.75
    x_no_currency = re.sub(r'(?i)\b(ngn|naira|usd|eur|gbp|ghs|zar|kes|cad|aud)\b', '', x)
    x_no_currency = re.sub(r'[?$€£]', '', x_no_currency)
    # Remove a leading standalone 'N' (common local currency prefix) when it precedes digits
    x_no_currency = re.sub(r'^\s*[nN]\s*(?=\d)', '', x_no_currency)
    
    # Remove thousand separators and extra spaces
    x_no_currency = re.sub(r'[ ,]', '', x_no_currency)
    
    # If letters remain after currency removal, we'll strip them later and keep numeric parts
    # (Previously we cleared such values; now we preserve digits.)
    
    # Remove specific special characters (keep decimal point)
    x_no_currency = re.sub(r'[-?]', '', x_no_currency)
    
    # First, check if the string is fully numeric (with a single decimal point)
    if re.match(r'^[0-9]+(\.[0-9]+)?$', x_no_currency):
        try:
            float_value = float(x_no_currency)
            return '{:.2f}'.format(float_value)
        except Exception:
            return ''
    
    try:
        # Remove any non-numeric characters except decimal point
        cleaned_value = re.sub(r'[^0-9.]', '', x_no_currency)
        
        # If nothing remains after cleaning, return original value
        if not cleaned_value:
            return ''
        
        # Count decimal points
        if cleaned_value.count('.') > 1:
            # If multiple decimal points, it's likely a formatting issue
            # Return the cleaned value (with commas removed) but don't try to convert
            return cleaned_value
        
        # Convert to float and format to 2 decimal places
        float_value = float(cleaned_value)
        return '{:.2f}'.format(float_value)
    
    except (ValueError, TypeError) as e:
        # If conversion fails, return the cleaned value with commas removed
        return cleaned_value if 'cleaned_value' in locals() else ''

def process_numeric_columns(df):
    """Process numeric columns to standardize their format"""
    numeric_columns = [
        'AVAILEDLIMIT', 
        'CREDITLIMIT',
        'OVERDUEAMOUNT',
        'LASTPAYMENTAMOUNT',
        'INSTALMENTAMOUNT',
        'INCOME',
        'OUTSTANDINGBALANCE'
    ]
    
    for col in numeric_columns:
        if col in df.columns:
            logger.info(f"Processing numeric column: {col}")
            
            # Apply the enhanced conversion function - this will retain original values that can't be converted
            df[col] = df[col].apply(try_convert_to_float)
            
            # Print sample values after processing for verification
            logger.info(f"Sample values in {col} after processing:")
            logger.info(df[col].head())
    
    return df

def process_collateral_details(df):
    """
    Process the COLLATERALDETAILS column by removing numeric values and special characters.
    Preserves spaces between words for readability.
    
    Args:
        df (pd.DataFrame): Input DataFrame containing COLLATERALDETAILS column
        
    Returns:
        pd.DataFrame: DataFrame with cleaned COLLATERALDETAILS column
    """
    if 'COLLATERALDETAILS' in df.columns:
        def clean_collateral_details(text):
            if pd.isna(text) or not isinstance(text, str):
                return text
            
            # Remove numeric values
            # text = re.sub(r'\d+', '', text)
            
            # Remove special characters but preserve spaces and ampersands
            # text = re.sub(r'[^a-zA-Z\s&]', '', text)
            
            # # Remove multiple spaces and strip
            text = re.sub(r'\s+', ' ', text).strip()
            
            return text
            
        df['COLLATERALDETAILS'] = df['COLLATERALDETAILS'].apply(clean_collateral_details)
    
    return df



def calculate_credit_unmerged_records(borrower_df, credit_df, borrower_type="individual"):
    if borrower_df.empty or credit_df.empty:
        return {
            'unmerged_count': 0,
            'total_valid_credit': 0,
            'matched_count': 0,
            'unmerged_customer_ids': []
        }
    
    # Filter valid credit records (prioritize CUSTOMERID, fallback to ACCOUNTNUMBER)
    valid_credit_records = credit_df.copy()
    
    # Create a matching field for credit records
    valid_credit_records['MATCH_ID'] = ''
    
    # First, use CUSTOMERID where available
    if 'CUSTOMERID' in credit_df.columns:
        customerid_mask = (
            credit_df['CUSTOMERID'].notna() & 
            (credit_df['CUSTOMERID'].str.strip() != '')
        )
        valid_credit_records.loc[customerid_mask, 'MATCH_ID'] = credit_df.loc[customerid_mask, 'CUSTOMERID'].str.strip()
    
    # For records without CUSTOMERID, use ACCOUNTNUMBER as fallback
    if 'ACCOUNTNUMBER' in credit_df.columns:
        accountnumber_mask = (
            (valid_credit_records['MATCH_ID'] == '') &
            credit_df['ACCOUNTNUMBER'].notna() & 
            (credit_df['ACCOUNTNUMBER'].str.strip() != '')
        )
        valid_credit_records.loc[accountnumber_mask, 'MATCH_ID'] = credit_df.loc[accountnumber_mask, 'ACCOUNTNUMBER'].str.strip()
    
    # Filter to only records with valid match IDs
    valid_credit_records = valid_credit_records[valid_credit_records['MATCH_ID'] != '']
    
    if valid_credit_records.empty:
        return {
            'unmerged_count': 0,
            'total_valid_credit': 0,
            'matched_count': 0,
            'unmerged_customer_ids': []
        }
    
    # Get unique customer IDs from borrower template
    borrower_customer_ids = set()
    if 'CUSTOMERID' in borrower_df.columns:
        valid_borrowers = borrower_df[
            borrower_df['CUSTOMERID'].notna() & 
            (borrower_df['CUSTOMERID'].str.strip() != '')
        ]
        if not valid_borrowers.empty:
            borrower_customer_ids = set(valid_borrowers['CUSTOMERID'].str.strip())
    
    # A credit record is considered "matched" if:
    # 1. Its MATCH_ID (CUSTOMERID or ACCOUNTNUMBER) matches a borrower's CUSTOMERID, OR
    # 2. Its ACCOUNTNUMBER matches a borrower's CUSTOMERID (fallback merge scenario)
    
    # Check primary match: MATCH_ID in borrower_customer_ids
    primary_match_mask = valid_credit_records['MATCH_ID'].isin(borrower_customer_ids)
    
    # Check fallback match: ACCOUNTNUMBER in borrower_customer_ids
    # This handles the case where merge_individual_borrowers matched on CUSTOMERID (indi) -> ACCOUNTNUMBER (credit)
    fallback_match_mask = pd.Series(False, index=valid_credit_records.index)
    if 'ACCOUNTNUMBER' in valid_credit_records.columns:
        accountnumber_valid = (
            valid_credit_records['ACCOUNTNUMBER'].notna() & 
            (valid_credit_records['ACCOUNTNUMBER'].str.strip() != '')
        )
        fallback_match_mask = accountnumber_valid & valid_credit_records['ACCOUNTNUMBER'].str.strip().isin(borrower_customer_ids)
    
    # Combined match: either primary OR fallback
    combined_match_mask = primary_match_mask | fallback_match_mask
    
    matched_records = valid_credit_records[combined_match_mask]
    unmerged_records = valid_credit_records[~combined_match_mask]
    
    # Get sample of unmerged customer IDs for debugging
    unmerged_customer_ids = unmerged_records['MATCH_ID'].unique().tolist()[:10]
    
    result = {
        'unmerged_count': len(unmerged_records),
        'total_valid_credit': len(valid_credit_records),
        'matched_count': len(matched_records),
        'unmerged_customer_ids': unmerged_customer_ids
    }
    
    return result


def merge_individual_borrowers(consu, credit, guar):
    """Merge individual borrower DataFrames"""
    # Validate DataFrames
    if consu.empty or credit.empty:
        logger.info("Warning: Individual borrower or credit information DataFrame is empty")
        return pd.DataFrame(), set()
    
    # Filter out rows with empty or blank 'CUSTOMERID'
    consu_cleaned = consu[
        consu['CUSTOMERID'].notna() & 
        (consu['CUSTOMERID'].str.strip() != '')
    ]

    # Also filter credit to prevent empty CUSTOMERID cross-joins
    credit_cleaned = credit[
        credit['CUSTOMERID'].notna() &
        (credit['CUSTOMERID'].astype(str).str.strip() != '')
    ] if 'CUSTOMERID' in credit.columns else credit
    
    # Merge attempts for individual borrowers
    merge_attempted = False
    indi = pd.DataFrame()  # Initialize indi DataFrame
    
    try:
        # First attempt: Merge on CUSTOMERID
        if 'CUSTOMERID' in credit.columns:
            logger.info("Attempting primary merge on CUSTOMERID")
            indi = pd.merge(
                consu_cleaned, 
                credit_cleaned, 
                on='CUSTOMERID', 
                how='inner',
                indicator=True  # Add merge indicator
            )
            logger.info(f"Primary merge matches: {indi.shape[0]} rows")
            logger.info("Merge indicator counts:")
            logger.info(indi['_merge'].value_counts())
            indi = indi.drop(columns=['_merge'])
            merge_attempted = True
    except Exception as e:
        logger.info(f"Primary merge failed: {str(e)}")

    # Fallback if primary merge failed or resulted in empty DataFrame
    if not merge_attempted or indi.empty:
        logger.info("Attempting fallback merge with ACCOUNTNUMBER")
        try:
            if 'ACCOUNTNUMBER' in credit.columns:
                # Filter credit ACCOUNTNUMBER blanks before fallback merge
                credit_acct_cleaned = credit[
                    credit['ACCOUNTNUMBER'].notna() &
                    (credit['ACCOUNTNUMBER'].astype(str).str.strip() != '')
                ] if 'ACCOUNTNUMBER' in credit.columns else credit
                # Use outer join temporarily to analyze matches
                temp_merge = pd.merge(
                    consu_cleaned,
                    credit_acct_cleaned,
                    left_on='CUSTOMERID',
                    right_on='ACCOUNTNUMBER',
                    how='outer',
                    indicator=True
                )
                logger.info("Fallback merge analysis:")
                logger.info(temp_merge['_merge'].value_counts())
                
                # Perform actual inner join
                indi = temp_merge[temp_merge['_merge'] == 'both'].copy()
                if not indi.empty:
                    indi = indi.drop(columns=['_merge'])
                    
                    # Drop CUSTOMERID from credit if it exists
                    if 'CUSTOMERID_y' in indi.columns:
                        indi = indi.drop(columns=['CUSTOMERID_y'], errors='ignore')  # Drop the credit CUSTOMERID if it exists
                        
                    # Rename CUSTOMERID_x to CUSTOMERID
                    if 'CUSTOMERID_x' in indi.columns:
                        indi = indi.rename(columns={'CUSTOMERID_x': 'CUSTOMERID'})
                    
                    logger.info(f"Fallback merge successful: {indi.shape[0]} rows")
                else:
                    logger.info("Warning: Fallback merge resulted in empty DataFrame")
        except Exception as e:
            logger.info(f"Fallback merge failed: {str(e)}")
            return pd.DataFrame(), set()

    if indi.empty:
        logger.info("Error: All merge attempts failed to produce results")
        logger.info(f"Consu shape: {consu_cleaned.shape}")
        logger.info(f"Credit shape: {credit.shape}")
        return pd.DataFrame(), set()
  
    
    # Handle BVNNUMBER column conflicts from individual borrower and credit info merge
    # Priority: Individual Borrower BVNNUMBER > Credit Info BVNNUMBER
    if 'BVNNUMBER_x' in indi.columns and 'BVNNUMBER_y' in indi.columns:
        logger.info("Coalescing BVNNUMBER columns from individual borrower and credit info")
        # Use individual borrower's BVN if available, otherwise use credit info's BVN
        indi['BVNNUMBER'] = indi['BVNNUMBER_x'].fillna('')
        indi.loc[indi['BVNNUMBER'].str.strip() == '', 'BVNNUMBER'] = indi.loc[indi['BVNNUMBER'].str.strip() == '', 'BVNNUMBER_y'].fillna('')
        # Drop the suffixed columns
        indi = indi.drop(columns=['BVNNUMBER_x', 'BVNNUMBER_y'], errors='ignore')
        logger.info("BVNNUMBER columns coalesced successfully")
    elif 'BVNNUMBER_y' in indi.columns and 'BVNNUMBER' not in indi.columns:
        # Only credit info had BVNNUMBER
        indi = indi.rename(columns={'BVNNUMBER_y': 'BVNNUMBER'})
        logger.info("Using BVNNUMBER from credit info (individual borrower template had none)")
    elif 'BVNNUMBER_x' in indi.columns:
        # Only individual borrower had BVNNUMBER 
        indi = indi.rename(columns={'BVNNUMBER_x': 'BVNNUMBER'})
        logger.info("Using BVNNUMBER from individual borrower template")
    
    indi.drop(columns=['NUMBEROFDIRECTORS'], inplace=True)
    
    # Track which credit IDs were successfully matched
    matched_credit_ids = set()
    if not indi.empty and 'CUSTOMERID' in indi.columns:
        matched_credit_ids = set(indi['CUSTOMERID'].dropna().unique())

    try:
        logger.info("Attempting guarantor merge with ACCOUNTNUMBER")
        indi_has_key = (
            indi['ACCOUNTNUMBER'].notna() &
            (indi['ACCOUNTNUMBER'].astype(str).str.strip() != '') &
            (~indi['ACCOUNTNUMBER'].astype(str).str.lower().isin(['nan', 'none']))
        )
        indi_for_merge = indi[indi_has_key]
        indi_no_key = indi[~indi_has_key]

        if not guar.empty and 'CUSTOMERSACCOUNTNUMBER' in guar.columns:
            guar = guar.copy()
            guar_has_key = (
                guar['CUSTOMERSACCOUNTNUMBER'].notna() &
                (guar['CUSTOMERSACCOUNTNUMBER'].astype(str).str.strip() != '') &
                (~guar['CUSTOMERSACCOUNTNUMBER'].astype(str).str.lower().isin(['nan', 'none']))
            )
            guar_clean = guar[guar_has_key]
        else:
            guar_clean = guar

        merged = pd.merge(
            indi_for_merge,
            guar_clean,
            left_on='ACCOUNTNUMBER',
            right_on='CUSTOMERSACCOUNTNUMBER',
            how='left'
        )
        # Re-attach rows that had blank keys (they get no guarantor data)
        indi = pd.concat([merged, indi_no_key], ignore_index=True)
        logger.info(f"Guarantor merge completed. Final shape: {indi.shape}")

    except Exception as e:
        logger.info(f"Guarantor merge failed: {str(e)}")
        logger.info("Continuing with original data")

    # Always ensure every guarantor column header exists in indi,
    # even when guar was empty or the merge produced no matches.
    # This MUST run AFTER the guarantor merge to avoid _x/_y suffix duplication.
    for guar_col in guar_mapping.keys():
        if guar_col not in indi.columns:
            indi[guar_col] = ''
    
    logger.info(f"[INDIVIDUAL MERGE] Matched {len(matched_credit_ids)} unique credit customer IDs")
    
    # Reorder columns based on consumer_merged_mapping to ensure consistent output
    # This fixes BVNNUMBER position when it comes from credit info instead of individual template
    if not indi.empty:
        # Get the canonical column order from consumer_merged_mapping
        canonical_order = list(consumer_merged_mapping.keys())
        
        # Get current columns that exist in the DataFrame
        current_columns = list(indi.columns)
        
        # Build the reordered column list: canonical columns first (in order), then any extras
        ordered_columns = []
        for col in canonical_order:
            if col in current_columns:
                ordered_columns.append(col)
        
        # Add any columns not in the mapping at the end (preserves extra columns)
        for col in current_columns:
            if col not in ordered_columns:
                ordered_columns.append(col)
        
        indi = indi[ordered_columns]
        logger.info(f"[INDIVIDUAL MERGE] Reordered columns to canonical order. BVNNUMBER position: {ordered_columns.index('BVNNUMBER') + 1 if 'BVNNUMBER' in ordered_columns else 'N/A'}")
    
    return indi, matched_credit_ids

def merge_corporate_borrowers(comm, credit, prin):
    """Merge corporate borrower DataFrames"""
    # Validate DataFrames
    if comm.empty or credit.empty:
        logger.info("Warning: Corporate borrower or credit information DataFrame is empty")
        return pd.DataFrame(), set()
    
    # Filter out rows with empty or blank 'CUSTOMERID'
    comm_cleaned = comm[
        comm['CUSTOMERID'].notna() & 
        (comm['CUSTOMERID'].str.strip() != '')
    ]

    # Also filter credit to prevent empty CUSTOMERID cross-joins
    credit_cleaned = credit[
        credit['CUSTOMERID'].notna() &
        (credit['CUSTOMERID'].astype(str).str.strip() != '')
    ] if 'CUSTOMERID' in credit.columns else credit
    
    # Merge attempts for corporate borrowers
    merge_attempted = False
    corpo = pd.DataFrame()  # Initialize corpo DataFrame
    
    try:
         # First attempt: Merge on CUSTOMERID
        if 'CUSTOMERID' in credit.columns:
            logger.info("Attempting primary merge on CUSTOMERID")
            corpo = pd.merge(
                comm_cleaned, 
                credit_cleaned,
                on='CUSTOMERID', 
                how='inner',
                indicator=True
            )
            logger.info(f"Primary merge matches: {corpo.shape[0]} rows")
            logger.info("Merge indicator counts:")
            logger.info(corpo['_merge'].value_counts())
            corpo = corpo.drop(columns=['_merge'])
            merge_attempted = True
    except Exception as e:
        logger.info(f"Primary merge failed: {str(e)}")
# Fallback if primary merge failed or resulted in empty DataFrame
    if not merge_attempted or corpo.empty:
        logger.info("Attempting fallback merge with ACCOUNTNUMBER")
        try:
           if 'ACCOUNTNUMBER' in credit.columns:
                # Filter credit ACCOUNTNUMBER blanks before fallback merge
                credit_acct_cleaned = credit[
                    credit['ACCOUNTNUMBER'].notna() &
                    (credit['ACCOUNTNUMBER'].astype(str).str.strip() != '')
                ] if 'ACCOUNTNUMBER' in credit.columns else credit
                # Use outer join temporarily to analyze matches
                temp_merge = pd.merge(
                    comm_cleaned,
                    credit_acct_cleaned,
                    left_on='CUSTOMERID',
                    right_on='ACCOUNTNUMBER',
                    how='outer',
                    indicator=True
                )
                logger.info("Fallback merge analysis:")
                logger.info(temp_merge['_merge'].value_counts())
 # Perform actual inner join
                corpo = temp_merge[temp_merge['_merge'] == 'both'].copy()
                if not corpo.empty:
                    corpo = corpo.drop(columns=['_merge'])
                    
                    # Drop CUSTOMERID from credit if it exists
                    if 'CUSTOMERID_y' in corpo.columns:
                        corpo = corpo.drop(columns=['CUSTOMERID_y'], errors='ignore')  # Drop the credit CUSTOMERID if it exists
                        
                    # Rename CUSTOMERID_x to CUSTOMERID
                    if 'CUSTOMERID_x' in corpo.columns:
                        corpo = corpo.rename(columns={'CUSTOMERID_x': 'CUSTOMERID'})
                    
                    logger.info(f"Fallback merge successful: {corpo.shape[0]} rows")
                else:
                    logger.info("Warning: Fallback merge resulted in empty DataFrame")
        except Exception as e:
            logger.info(f"Fallback merge failed: {str(e)}")
            return pd.DataFrame(), set()

    if corpo.empty:
        logger.info("Error: All merge attempts failed to produce results")
        logger.info(f"Consu shape: {comm_cleaned.shape}")
        logger.info(f"Credit shape: {credit.shape}")
        return pd.DataFrame(), set()
    logger.info("After merging with credit (inner join):")
    logger.info(f"corpo shape: {corpo.shape}")


    # Merge with principal officers information
    try:
        corpo = pd.merge(
            corpo,
            prin,
            left_on='CUSTOMERID',
            right_on='CUSTOMERID',
            how='left'
        )
        logger.info(f"principal merge successful. Final shape: {corpo.shape}")
    except Exception as e:
        logger.info(f"Principal merge failed: {str(e)}")
    corpo.drop(columns=['FACILITYOWNERSHIPTYPE', 'INCOME', 'INCOMEFREQUENCY', 'OWNERTENANT', 'NUMBEROFPARTICIPANTSINJOINTLOAN', 'DEPENDANTS', 'BVNNUMBER'], inplace=True, errors='ignore')
    
    # Track which credit IDs were successfully matched
    matched_credit_ids = set()
    if not corpo.empty and 'CUSTOMERID' in corpo.columns:
        matched_credit_ids = set(corpo['CUSTOMERID'].dropna().unique())
    
    logger.info(f"[CORPORATE MERGE] Matched {len(matched_credit_ids)} unique credit customer IDs")
    
    # Reorder columns based on commercial_merged_mapping to ensure consistent output
    if not corpo.empty:
        # Get the canonical column order from commercial_merged_mapping
        canonical_order = list(commercial_merged_mapping.keys())
        
        # Get current columns that exist in the DataFrame
        current_columns = list(corpo.columns)
        
        # Build the reordered column list: canonical columns first (in order), then any extras
        ordered_columns = []
        for col in canonical_order:
            if col in current_columns:
                ordered_columns.append(col)
        
        # Add any columns not in the mapping at the end (preserves extra columns)
        for col in current_columns:
            if col not in ordered_columns:
                ordered_columns.append(col)
        
        corpo = corpo[ordered_columns]
        logger.info(f"[CORPORATE MERGE] Reordered columns to canonical order")
    
    return corpo, matched_credit_ids

def remove_duplicates(df, columns_to_check=None):
    """
    Remove duplicates from DataFrame to mimic Excel's Remove Duplicates feature
    
    Args:
        df (pd.DataFrame): Input DataFrame
        columns_to_check (list, optional): Columns to check for duplicates (like Excel's column selection)
                                          If None, all columns are used
    
    Returns:
        pd.DataFrame: Cleaned DataFrame with duplicates removed
    """
    if df is None or df.empty:
        return df
    
    # If no columns specified, use all columns (like Excel default)
    if columns_to_check is None or len(columns_to_check) == 0:
        columns_to_check = df.columns.tolist()
    else:
        # Only use columns that actually exist in the dataframe
        columns_to_check = [col for col in columns_to_check if col in df.columns]
        
        if not columns_to_check:
            logger.info("None of the specified columns found in DataFrame. Using all columns.")
            columns_to_check = df.columns.tolist()
    
    # Create a copy for case-insensitive comparison
    df_clean = df.copy()
    
    # Make string comparisons case-insensitive like Excel
    for col in columns_to_check:
        if df_clean[col].dtype == 'object':  # Only process string columns
            # Convert to lowercase for case-insensitive comparison (like Excel)
            df_clean[col] = df_clean[col].astype(str).str.lower()
            
            # Excel ignores leading/trailing spaces in comparisons
            df_clean[col] = df_clean[col].str.strip()
    
    # Perform duplicate removal (keeping first occurrence like Excel)
    # Get indices of rows to keep
    indices_to_keep = df_clean.drop_duplicates(
        subset=columns_to_check,
        keep='first'
    ).index
    
    # Use original dataframe with these indices to preserve original data
    df_cleaned = df.loc[indices_to_keep].reset_index(drop=True)
    
    # Log removed rows
    rows_removed = len(df) - len(df_cleaned)
    if rows_removed > 0:
        logger.info(f"Removed {rows_removed} duplicate rows")
    
    return df_cleaned
# Pre-computed lowercase keyword sets for confidence classification (module-level for performance)
_COMMERCIAL_KEYWORDS_LOWER = frozenset(k.lower() for k in commercial_keywords)
_TIER_2_STRONG = _COMMERCIAL_KEYWORDS_LOWER - TIER_1_SUFFIXES - TIER_3_AMBIGUOUS


def is_commercial_entity(name, commercial_keywords):
    """
    Check for commercial entities by looking at standalone words
    
    Args:
        name (str): Full name to check
        commercial_keywords (list): List of commercial keywords
    
    Returns:
        bool: True if likely a commercial entity, False otherwise
    """
    if not isinstance(name, str):
        return False
    
    # Convert to lowercase and split into words
    name_words = set(name.lower().split())
    
    # Convert keywords to lowercase for case-insensitive comparison
    commercial_keywords_lower = [keyword.lower() for keyword in commercial_keywords]
    # Check for standalone commercial keywords
    commercial_matches = [
        keyword for keyword in commercial_keywords_lower
        if keyword in name_words
    ]
    
    
    return len(commercial_matches) > 0

def classify_commercial_confidence(name):
    """
    Tiered confidence classifier for commercial entity detection.

    Returns:
        'auto_commercial' - Last word is a Tier 1 suffix, OR 2+ Tier 1/2 combined matches.
        'manual_review'   - Exactly 1 Tier 2 match, OR any Tier 3 ambiguous match.
        None              - No commercial keyword matched (stays in individual).
    """
    if not isinstance(name, str) or not name.strip():
        return None

    words = name.lower().split()
    words_set = set(words)

    tier1_matches = words_set & TIER_1_SUFFIXES
    tier2_matches = words_set & _TIER_2_STRONG
    tier3_matches = words_set & TIER_3_AMBIGUOUS

    # No commercial keyword at all → stay in individual
    if not tier1_matches and not tier2_matches and not tier3_matches:
        return None

    # Any Tier 1 or Tier 2 match → unambiguously commercial, auto-move
    if tier1_matches or tier2_matches:
        return 'auto_commercial'

    # Only Tier 3 ambiguous word(s) matched → manual review
    return 'manual_review'

def split_commercial_entities(indi):
    """
    Split individual borrowers into three groups based on commercial confidence.

    Returns:
        remaining_indi     - Records with no commercial keyword match (stay as individual).
        auto_commercial    - High-confidence commercial entities (auto-moved, skip UI).
        manual_review      - Ambiguous records sent to the verification queue.
    """
    auto_commercial_rows = []
    manual_review_rows = []
    rows_to_remove = []

    for index, row in indi.iterrows():
        name_columns = ['SURNAME', 'FIRSTNAME', 'MIDDLENAME']
        full_name = ' '.join([str(row[col]) for col in name_columns if pd.notna(row[col]) and str(row[col]).lower() not in ('nan', 'none', '')])

        confidence = classify_commercial_confidence(full_name)
        if confidence is None:
            continue

        commercial_row = row.copy()
        original_combined_name = full_name.strip()
        commercial_row['ORIGINAL_BUSINESSNAME'] = original_combined_name
        commercial_row['SURNAME'] = original_combined_name
        commercial_row['DATA'] = 'D'

        rows_to_remove.append(index)
        if confidence == 'auto_commercial':
            auto_commercial_rows.append(commercial_row)
        else:
            manual_review_rows.append(commercial_row)

    indi = indi.drop(rows_to_remove).reset_index(drop=True)

    def _finalise(rows, columns):
        if not rows:
            return pd.DataFrame(columns=columns)
        df = pd.DataFrame(rows)
        if 'DATA' not in df.columns:
            df['DATA'] = 'D'
        df['DATA'] = df['DATA'].fillna('D')
        df = df.where(pd.notnull(df), '')
        return df

    auto_commercial = _finalise(auto_commercial_rows, indi.columns)
    manual_review = _finalise(manual_review_rows, indi.columns)

    return indi, auto_commercial, manual_review

def is_consumer_entity(name, commercial_keywords, threshold=90):
    """
    Check if a business name is likely a consumer entity by confirming it doesn't contain (fuzzy) commercial keywords as standalone words.
    Uses fuzzy matching for standalone words only.
    """
    if name is None or not isinstance(name, str) or not name.strip():
        return False
    
    # Convert to lowercase and split into words for single word matching
    name_words = set(name.lower().split())
    commercial_keywords_lower = [keyword.lower() for keyword in commercial_keywords]
    
    # Fuzzy match: only match if a word in the business name is a fuzzy match to a commercial keyword
    for word in name_words:
        for keyword in commercial_keywords_lower:
            if fuzz.ratio(word, keyword) >= threshold:
                # If the word is a fuzzy match to a commercial keyword, it's not a consumer entity
                return False
    return True

def classify_consumer_confidence(name):
    """
    Tiered confidence classifier for consumer entity detection.

    Returns:
        'auto_consumer'  - Zero commercial keyword matches → clearly a personal name.
        'manual_review'  - ONLY Tier 3 ambiguous keyword(s) matched → needs human check.
        None             - A Tier 1 or Tier 2 keyword matched → stays in corporate.
    """
    if not isinstance(name, str) or not name.strip():
        return None

    words_set = set(name.lower().split())

    if words_set & TIER_1_SUFFIXES:
        return None
    if words_set & _TIER_2_STRONG:
        return None

    tier3_matches = words_set & TIER_3_AMBIGUOUS
    if tier3_matches:
        return 'manual_review'

    return 'auto_consumer'

def split_consumer_entities(corpo):
    """
    Split corporate borrowers into three groups based on consumer confidence.

    Returns:
        remaining_corpo  - Records with a Tier 1 or Tier 2 keyword match (stay as corporate).
        auto_consumer    - Zero keyword matches → clearly personal names (auto-moved, skip UI).
        manual_review    - Only Tier 3 ambiguous keyword(s) matched → sent to verification queue.
    """
    if 'BUSINESSNAME' not in corpo.columns:
        return corpo, pd.DataFrame(), pd.DataFrame()

    auto_consumer_rows = []
    manual_review_rows = []
    rows_to_remove = []

    for index, row in corpo.iterrows():
        if pd.isna(row['BUSINESSNAME']):
            continue

        business_name = str(row['BUSINESSNAME']).strip()
        confidence = classify_consumer_confidence(business_name)

        if confidence is None:
            continue
        if 'DATEOFINCORPORATION' not in row.index:
            continue

        doi_value = row['DATEOFINCORPORATION']
        doi_is_empty = (
            doi_value is None
            or (isinstance(doi_value, float) and pd.isna(doi_value))
            or str(doi_value).strip() == ''
            or str(doi_value).strip().lower() in ('none', 'nan', 'null', 'nill', 'nil')
        )
        if doi_is_empty:
            continue

        try:
            doi_str = str(doi_value).strip()
            parsed_doi = dateparser.parse(doi_str)
            if parsed_doi is None:
                continue  # Unparseable date — keep as commercial
            years_since_incorporation = datetime.now().year - parsed_doi.year
            if years_since_incorporation < 18:
                continue  # Recently incorporated — keep as commercial
        except Exception:
            continue  # Any parsing error — keep as commercial

        consumer_data = row.to_dict()
        consumer_data['ORIGINAL_BUSINESSNAME'] = business_name

        cleaned_business_name = remove_titles(business_name)
        name_parts = cleaned_business_name.split(maxsplit=2)
        consumer_data['SURNAME'] = name_parts[0] if len(name_parts) > 0 else ''
        consumer_data['FIRSTNAME'] = name_parts[1] if len(name_parts) > 1 else ''
        consumer_data['MIDDLENAME'] = name_parts[2] if len(name_parts) > 2 else ''
        consumer_data['DEPENDANTS'] = '00'
        consumer_data['DATA'] = 'D'

        if 'BUSINESSNAME' in consumer_data:
            del consumer_data['BUSINESSNAME']

        rows_to_remove.append(index)
        if confidence == 'auto_consumer':
            auto_consumer_rows.append(consumer_data)
        else:
            manual_review_rows.append(consumer_data)

    if rows_to_remove:
        corpo = corpo.drop(rows_to_remove).reset_index(drop=True)

    def _to_df(rows):
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        return df.where(pd.notnull(df), '')

    return corpo, _to_df(auto_consumer_rows), _to_df(manual_review_rows)

def merge_dataframes(processed_sheets):
    """
    Main merging function with sequential processing
    
    Args:
        processed_sheets (dict): Dictionary of processed DataFrames
    
    Returns:
        tuple: (Individual borrowers DataFrame, Corporate borrowers DataFrame)
    """
    # Check if we have merged sheets
    if 'consumermerged' in processed_sheets or 'commercialmerged' in processed_sheets:
        logger.info("\n=== PROCESSING MERGED SHEETS ===")
        indi = processed_sheets.get('consumermerged', pd.DataFrame())
        corpo = processed_sheets.get('commercialmerged', pd.DataFrame())
    else:
        indi = processed_sheets.get('individualborrowertemplate', pd.DataFrame())
        corpo = processed_sheets.get('corporateborrowertemplate', pd.DataFrame())
        
        # Apply null value cleaning
        indi = indi.applymap(lambda x: None if str(x).strip().lower() in ['none', 'nan', 'null', 'nill', 'nil'] else x)
        corpo = corpo.applymap(lambda x: None if str(x).strip().lower() in ['none', 'nan', 'null', 'nill', 'nil'] else x)
        
        logger.info("\n=== MERGED SHEET DATA (Before Split) ===")
        logger.info("Individual records:", len(indi))
        logger.info("Corporate records:", len(corpo))

        # --- Added Processing for Merged Sheets ---
        # Split commercial entities from the consumer_merged data
        if not indi.empty:
            logger.info("\nSplitting commercial entities from consumer_merged data...")
            indi, auto_commercial_merged, manual_commercial_merged = split_commercial_entities(indi)
            corpo2 = pd.concat([auto_commercial_merged, manual_commercial_merged], ignore_index=True)
            logger.info(f"  - Individual records after split: {len(indi)}")
            logger.info(f"  - Commercial entities extracted: {len(corpo2)}")

            # Rename and concatenate if commercial entities were found
            if not corpo2.empty:
                logger.info("\nRenaming columns for extracted commercial entities...")
                corpo2 = rename_columns(corpo2, ConsuToComm.copy())
                
                # Ensure both dataframes have reset indexes
                if not corpo.empty:
                    corpo = corpo.reset_index(drop=True)
                corpo2 = corpo2.reset_index(drop=True)
                
                logger.info("\nConcatenating extracted commercial entities with corporate data...")
                try:
                    # If corpo is empty, just use corpo2
                    if corpo.empty:
                        corpo = corpo2
                        logger.info(f"  - Using extracted commercial entities as corporate data: {len(corpo)} rows")
                    else:
                        # Use columns parameter to ensure concatenation uses only columns from mapping
                        common_columns = [col for col in corpo2.columns if col in corpo.columns]
                        if not common_columns:
                            # If no common columns, use all columns from corpo2
                            corpo = pd.concat([corpo, corpo2], ignore_index=True, sort=False)
                        else:
                            corpo = pd.concat([corpo[common_columns], corpo2[common_columns]], ignore_index=True)
                        logger.info(f"  - Total corporate records after concatenation: {len(corpo)}")
                except Exception as e:
                    logger.info(f"Error during commercial concatenation: {e}")
                    logger.info(f"corpo columns: {list(corpo.columns)}")
                    logger.info(f"corpo2 columns: {list(corpo2.columns)}")
                    # If concatenation fails, at least ensure corpo2 is preserved
                    if corpo.empty:
                        corpo = corpo2.copy()
                        logger.info("Using only extracted commercial entities as corporate data")
        
        # Split consumer entities from the commercial_merged data
        if not corpo.empty:
            logger.info("\nSplitting consumer entities from commercial_merged data...")
            corpo, auto_consumer_merged, manual_consumer_merged = split_consumer_entities(corpo)
            indi2 = pd.concat([auto_consumer_merged, manual_consumer_merged], ignore_index=True)
            logger.info(f"  - Corporate records after split: {len(corpo)}")
            logger.info(f"  - Consumer entities extracted: {len(indi2)}")

            # Rename and concatenate if consumer entities were found
            if not indi2.empty:
                logger.info("\nRenaming columns for extracted consumer entities...")
                # Apply the CommToConsu mapping to rename columns and strictly order them
                indi2 = rename_columns(indi2, CommToConsu.copy())
                
                # Ensure both dataframes have reset indexes
                indi = indi.reset_index(drop=True)
                indi2 = indi2.reset_index(drop=True)
                
                logger.info("\nConcatenating extracted consumer entities with individual data...")
                try:
                    # If indi is empty, just use indi2
                    if indi.empty:
                        indi = indi2
                        logger.info(f"  - Using extracted consumer entities as individual data: {len(indi)} rows")
                    else:
                        # Ensure both dataframes have the same column ordering by applying the same mapping
                        indi = rename_columns(indi, CommToConsu.copy())
                        indi2 = rename_columns(indi2, CommToConsu.copy())
                        
                        # Direct concatenation without filtering to common columns
                        indi = pd.concat([indi, indi2], ignore_index=True, sort=False)
                        logger.info(f"Total individual borrowers after concatenation: {len(indi)}")
                except Exception as e:
                    logger.info(f"Error during consumer concatenation: {str(e)}")
                    logger.info(f"indi columns: {list(indi.columns)}")
                    logger.info(f"indi2 columns: {list(indi2.columns)}")
                    # If concatenation fails, at least ensure indi2 is preserved
                    if indi.empty:
                        indi = indi2.copy()
                        logger.info("Using only extracted consumer entities as individual data")
        # --- End Added Processing ---
                
        logger.info("\n=== FINAL MERGED SHEET DATA ===")
        logger.info("Final Individual records:", len(indi))
        logger.info("Final Corporate records:", len(corpo))
        
        return indi, corpo

    # Regular processing for non-merged sheets
    # Extract DataFrames from processed sheets
    consu = processed_sheets.get('individualborrowertemplate', pd.DataFrame())
    comm = processed_sheets.get('corporateborrowertemplate', pd.DataFrame())
    credit = processed_sheets.get('creditinformation', pd.DataFrame())
    guar = processed_sheets.get('guarantorsinformation', pd.DataFrame())
    prin = processed_sheets.get('principalofficerstemplate', pd.DataFrame())

    indi = merge_individual_borrowers(consu, credit, guar)
    corpo = merge_corporate_borrowers(comm, credit, prin)

    # Print merged corporate borrowers
    logger.info("\n=== MERGED CORPORATE BORROWERS ===")
    logger.info(corpo.head()) 

    indi = indi.applymap(lambda x: None if str(x).strip().lower() in ['none', 'nan', 'null', 'nill', 'nil'] else x)
    corpo = corpo.applymap(lambda x: None if str(x).strip().lower() in ['none', 'nan', 'null', 'nill', 'nil'] else x)
    
    #Step 3: Split commercial entities from individual borrowers
    indi, auto_commercial_reg, manual_commercial_reg = split_commercial_entities(indi)
    corpo2 = pd.concat([auto_commercial_reg, manual_commercial_reg], ignore_index=True)

    logger.info("\n=== SHEET DATA AFTER MERGING ===")
    logger.info("Number of rows:", len(indi))
    logger.info("First few rows:")
    logger.info(indi.head())

    logger.info("Number of rows:", len(corpo))
    logger.info("First few rows:")
    logger.info(corpo.head())
    logger.info("Original corporate borrowers:", len(corpo))
    logger.info("Commercial entities to add:", len(corpo2))
    
    logger.info("Number of rows:", len(corpo2))
    logger.info("First few rows:")
    logger.info(corpo2.head())
    logger.info("================================")

    # Step 4: Rename commercial entities before combining
    if not corpo2.empty:
        # Rename corpo2 columns to match corporate borrower template
        corpo2 = rename_columns(corpo2, ConsuToComm.copy())
        
        # Debug statement to show corpo2 details before concatenation
        logger.info("Number of commercial entities:", len(corpo2))
        logger.info("First few rows of corpo2:")
        logger.info(corpo2.head())
        
        # Ensure both dataframes have reset indexes
        corpo = corpo.reset_index(drop=True)
        corpo2 = corpo2.reset_index(drop=True)
        
        # Combine commercial entities with existing corporate borrowers
        try:
            # If corpo is empty, just use corpo2
            if corpo.empty:
                corpo = corpo2
                logger.info(f"Using extracted commercial entities as corporate data: {len(corpo)} rows")
            else:
                # Apply ConsuToComm mapping to ensure columns are aligned
                corpo2 = rename_columns(corpo2, ConsuToComm.copy())
                
                # Direct concatenation without filtering to common columns
                corpo = pd.concat([corpo, corpo2], ignore_index=True, sort=False)
                logger.info(f"Total corporate borrowers after concatenation: {len(corpo)}")
        except Exception as e:
            logger.info(f"Error during commercial concatenation: {e}")
            logger.info(f"corpo columns: {list(corpo.columns)}")
            logger.info(f"corpo2 columns: {list(corpo2.columns)}")
            # If concatenation fails, at least ensure corpo2 is preserved
            if corpo.empty:
                corpo = corpo2.copy()
                logger.info("Using only extracted commercial entities as corporate data")

        # Additional check to verify commercial entities were added
        commercial_entities_in_corpo = pd.DataFrame()
        if 'BUSINESSNAME' in corpo.columns:
            commercial_entities_in_corpo = corpo[
                corpo['BUSINESSNAME'].apply(
                    lambda x: any(keyword in str(x).lower() for keyword in commercial_keywords)
                )
            ]
            logger.info("\nCommercial Entities in Final Corporate Borrowers:")
            logger.info("Number of commercial entities:", len(commercial_entities_in_corpo))
            logger.info("First few commercial entities:")
            logger.info(commercial_entities_in_corpo.head())
        else:
            logger.info("\nWARNING: 'BUSINESSNAME' column not found in corporate DataFrame")
            logger.info("Cannot identify commercial entities in corporate borrowers")
    
    # Step 5: Split consumer entities from corporate borrowers
    corpo, auto_consumer_reg, manual_consumer_reg = split_consumer_entities(corpo)
    indi2 = pd.concat([auto_consumer_reg, manual_consumer_reg], ignore_index=True)
    
    logger.info("\n=== SPLIT CONSUMER ENTITIES ===")
    logger.info("Corporate records after split:", len(corpo))
    logger.info("Consumer entities extracted:", len(indi2))
    
    # Step 6: Rename consumer entities before combining
    if not indi2.empty:
        # Rename indi2 columns to match individual borrower template
        logger.info("\nRenaming columns for extracted consumer entities...")
        indi2 = rename_columns(indi2, CommToConsu.copy())
        
        # Debug statement to show indi2 details before concatenation
        logger.info("Number of consumer entities:", len(indi2))
        logger.info("First few rows of indi2:")
        logger.info(indi2.head())
        
        # Ensure both dataframes have reset indexes
        indi = indi.reset_index(drop=True)
        indi2 = indi2.reset_index(drop=True)
        
        # Combine consumer entities with existing individual borrowers
        try:
            # If indi is empty, just use indi2
            if indi.empty:
                indi = indi2
                logger.info(f"Using extracted consumer entities as individual data: {len(indi)} rows")
            else:
                # Ensure indi has the same column ordering as CommToConsu
                indi = rename_columns(indi, CommToConsu.copy())
                
                # Now both dataframes have exactly the same columns in the same order,
                # we can safely concatenate them
                indi = pd.concat([indi, indi2], ignore_index=True)
                logger.info(f"Total individual records after concatenation: {len(indi)}")
                    
            # Debug statement to confirm addition
            logger.info("\nAfter Adding Consumer Entities:")
            logger.info("Total individual borrowers:", len(indi))
            logger.info("Columns in final indi:", list(indi.columns[:10]) + ["..."])  # Show first 10 columns
            logger.info("First few rows after addition:")
            logger.info(indi.head())
        except Exception as e:
            logger.info(f"Error during consumer concatenation: {str(e)}")
            logger.info(f"indi columns: {list(indi.columns)}")
            logger.info(f"indi2 columns: {list(indi2.columns)}")
            # If concatenation fails, at least ensure indi2 is preserved
            if indi.empty:
                indi = indi2.copy()
                logger.info("Using only extracted consumer entities as individual data")
    
    return indi, corpo
 
def rename_columns(df, column_mapping):
    """
    Rename columns based on a mapping dictionary and strictly enforce column order
    
    Args:
        df (pd.DataFrame): Input DataFrame
        column_mapping (dict): Mapping of column names
    
    Returns:
        pd.DataFrame: DataFrame with renamed columns and ordered according to mapping
    """
    try:
        # Create a fresh copy of the dataframe to avoid modifying the original
        df = df.copy()

        # Rename columns that match the mapping
        for column in list(df.columns):  # Use list() to create a copy of columns
            for mapped_column, alt_names in column_mapping.items():
                if column in alt_names or column.lower() in alt_names or column.upper() in alt_names:
                    df.rename(columns={column: mapped_column}, inplace=True)
                    logger.info(f"Renamed {column} to {mapped_column}")
                    break

        # Check for duplicate columns and make them unique
        if len(df.columns) != len(set(df.columns)):
            logger.info("WARNING: Duplicate column names detected, making them unique...")
            # Create a new columns list without duplicates
            seen = set()
            new_columns = []
            for col in df.columns:
                if col not in seen:
                    seen.add(col)
                    new_columns.append(col)
                else:
                    # For duplicates, add a suffix
                    i = 1
                    while f"{col}_{i}" in seen:
                        i += 1
                    seen.add(f"{col}_{i}")
                    new_columns.append(f"{col}_{i}")
            
            # Assign the new unique column names
            df.columns = new_columns
        
        # Print columns before final reordering
        logger.info("Columns before reordering:", list(df.columns))

        # Create ordered DataFrame using concat instead of adding columns one by one
        # This avoids DataFrame fragmentation warning
        column_dfs = []
        for col in column_mapping.keys():
            if col in df.columns:
                # For existing columns, use the values from the original DataFrame
                column_dfs.append(pd.DataFrame({col: df[col]}))
            else:
                # For missing columns, create a new DataFrame with None values
                column_dfs.append(pd.DataFrame({col: [None] * len(df)}))
        
        # Use concat to join all columns at once
        if column_dfs:
            ordered_df = pd.concat(column_dfs, axis=1)
        else:
            # If no columns were found, create an empty DataFrame with the right columns
            ordered_df = pd.DataFrame(columns=list(column_mapping.keys()))
        
        # Reset index to ensure clean index for concatenation
        ordered_df = ordered_df.reset_index(drop=True)
        
        # Print final columns
        logger.info("Final columns after strict reordering:", list(ordered_df.columns))
        logger.info(f"Final dataframe has {len(ordered_df.columns)} columns and {len(ordered_df)} rows")

        return ordered_df
    except Exception as e:
        logger.info(f"Error in rename_columns: {e}")
        traceback.print_exc()
        return df


def modify_middle_names(df):
    """Keep only the first name in the specified middle name columns, removing standalone 'and'."""
    middle_name_columns = [
        'MIDDLENAME',
        'SPOUSEMIDDLENAME',
        'GUARANTORMIDDLENAME',
        'PRINCIPALOFFICER1MIDDLENAME',
        'PRINCIPALOFFICER2MIDDLENAME'
    ]
    
    for col in middle_name_columns:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: (
                str(x).split()[0] if pd.notna(x) and str(x).strip() and str(x).split()[0].lower() != 'and' else ''
            ) if pd.notna(x) and str(x).strip() else '')
    
    return df
def trim_strings_to_59(df):
    """
    Trim all string values in the DataFrame to 59 characters maximum
    
    Args:
        df (pd.DataFrame): Input DataFrame
        
    Returns:
        pd.DataFrame: DataFrame with all string values trimmed to 59 characters
    """
    # Define the trimming function
    def trim_string(s):
        if isinstance(s, str) and len(s) > 59:
            return s[:58]  # Trim to 58 characters as requested
        return s
    
    # Apply the function to all elements in the DataFrame
    logger.info("\n=== TRIMMING STRING VALUES TO 59 CHARACTERS ===")
    df = df.map(trim_string)  # Updated from applymap (deprecated)
    logger.info("String trimming completed")
    
    return df

# def normalize_spaces(text):
#     """Replace multiple consecutive spaces with single spaces."""
#     if text is None:
#         return ""
#     return re.sub(r' +', ' ', str(text)).strip()

def convert_numpy(obj):
    if isinstance(obj, np.integer):
        return int(obj)
    elif isinstance(obj, np.floating):
        return float(obj)
    elif isinstance(obj, np.ndarray):
        return obj.tolist()
    return obj


def save_excel_as_text(df, filename):
    """
    Save DataFrame to Excel with all cells formatted as Text.
    This ensures the Excel ribbon shows "Text" format instead of "General".
    
    Args:
        df (pd.DataFrame): DataFrame to save
        filename (str): Full path to output Excel file
    """
    try:
        # Create Excel writer with xlsxwriter engine
        writer = pd.ExcelWriter(filename, engine='xlsxwriter')
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = workbook.add_worksheet('Sheet1')
        
        # Define text format (@ is Excel's text format code)
        text_format = workbook.add_format({'num_format': '@'})
        
        # Write headers with text format
        for col_idx, header in enumerate(df.columns):
            worksheet.write(0, col_idx, str(header), text_format)
        
        # Write data with text format - this ensures ribbon shows "Text"
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row):
                # Convert all values to string and write with text format
                cell_value = str(value) if pd.notna(value) else ''
                worksheet.write(row_idx + 1, col_idx, cell_value, text_format)
        
        # Apply text format to ALL columns (for any future data)
        for col_idx in range(len(df.columns)):
            worksheet.set_column(col_idx, col_idx, None, text_format)
        
        # Close and save the file
        writer.close()
        
        logger.info(f"✓ Saved Excel file with Text formatting: {filename}")
        
    except Exception as e:
        logger.info(f"Error saving Excel with text format: {str(e)}")
        # Fallback to regular save if text formatting fails
        logger.info(f"Falling back to standard Excel save...")
        df.to_excel(filename, index=False)


def _apply_rate_limit(view_func):
    """
    Apply rate limiting decorator if django-ratelimit is available.
    Rate limits: 3 uploads per minute, 10 uploads per hour for regular users.
    """
    if RATELIMIT_AVAILABLE:
        # Apply multiple rate limits
        view_func = ratelimit(key='user', rate='3/m', method='POST', block=False)(view_func)
        view_func = ratelimit(key='user', rate='10/h', method='POST', block=False)(view_func)
    return view_func


@login_required
@_apply_rate_limit
def upload_file(request):
    # Check if rate limited
    if getattr(request, 'limited', False):
        logger.warning(f"[RATE LIMIT] User {request.user.username} exceeded upload rate limit")
        messages.error(
            request,
            'Upload rate limit exceeded. You can upload up to 3 files per minute or 10 files per hour. '
            'Please wait a few minutes before trying again.'
        )
        return redirect('auto:upload')
    
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            # Get list of uploaded files (form now returns a list)
            uploaded_files = form.cleaned_data['file']
            if not isinstance(uploaded_files, list):
                uploaded_files = [uploaded_files]
            
            # For display/tracking, combine all filenames
            # original_filenames = [os.path.splitext(f.name)[0] for f in uploaded_files]
            combined_filename = ', '.join([f.name for f in uploaded_files])
                        
            # Truncate to 255 characters to fit in database field (max_length=255)
            if len(combined_filename) > 255:
                combined_filename = combined_filename[:252] + '...'
            total_size = sum(f.size for f in uploaded_files)
            
            # Capture user-selected reporting period
            selected_month = int(form.cleaned_data['month'])
            selected_year = int(form.cleaned_data['year'])
            
            # Get subscriber information based on user type
            from acctmgt.models import UserProfile
            from .models import UploadSession, Subscriber
            import time
            import calendar
            
            # Check if user is in multi_subscriber group
            is_multi_subscriber = request.user.groups.filter(name='multi_subscriber').exists()
            
            if is_multi_subscriber:
                # Multi-subscriber user: get subscriber from form dropdown
                selected_subscriber_id = form.cleaned_data.get('subscriber_id')
                if not selected_subscriber_id:
                    messages.error(request, 'Please select a subscriber from the dropdown.')
                    return redirect('auto:upload')
                
                try:
                    bound_subscriber = Subscriber.objects.get(subscriber_id=selected_subscriber_id)
                    subscriber_id = bound_subscriber.subscriber_id
                    subscriber_name = bound_subscriber.subscriber_name
                    logger.info(f"[UPLOAD] Multi-subscriber user {request.user.username} uploading for subscriber: {subscriber_name} (ID: {subscriber_id})")
                except Subscriber.DoesNotExist:
                    messages.error(request, 'Selected subscriber not found.')
                    return redirect('auto:upload')
            else:
                # Regular user: get subscriber from user's permanent binding
                user_profile = UserProfile.get_or_create_profile(request.user)
                
                # Ensure user is bound to a subscriber
                if not user_profile.is_bound:
                    messages.error(request, 'Your account is not bound to any organization. Please complete the binding process.')
                    return redirect('acctmgt:subscriber_selection')
                
                # Get bound subscriber information
                bound_subscriber = user_profile.get_bound_subscriber()
                if not bound_subscriber:
                    messages.error(request, 'Unable to retrieve your organization information. Please contact support.')
                    return redirect('acctmgt:subscriber_selection')
                
                subscriber_id = bound_subscriber.subscriber_id
                subscriber_name = bound_subscriber.subscriber_name
            
            # Save original files for archival purposes before any processing
            upload_timestamp = timezone.now()
            original_file_paths = []
            file_paths = []
            
            fs = FileSystemStorage()
            
            for uploaded_file in uploaded_files:
                # Save original file for archival
                original_save_success, original_file_path, original_error = save_original_file(
                    uploaded_file, 
                    subscriber_id, 
                    upload_timestamp
                )
                
                if not original_save_success:
                    messages.error(request, f'Failed to archive original file {uploaded_file.name}: {original_error}')
                    return redirect('auto:upload_file')
                
                original_file_paths.append(original_file_path)
                
                # Reset file pointer for processing
                uploaded_file.seek(0)
                
                # Save file for processing
                filename = fs.save(uploaded_file.name, uploaded_file)
                file_path = os.path.join(settings.MEDIA_ROOT, filename)
                file_paths.append(file_path)
            
            # Create UploadSession to track this upload (use first file as primary, note all in filename)
            upload_session = UploadSession.objects.create(
                user=request.user,
                subscriber_id=subscriber_id,
                filename=combined_filename,
                original_filename=combined_filename,
                file_size=total_size,
                status='uploading',
                original_file_path=','.join(original_file_paths),  # Store all paths
                reporting_month=selected_month,
                reporting_year=selected_year
            )
            
            # Show confirmation message with selected reporting period
            month_name = calendar.month_name[selected_month]
            file_count_msg = f"Processing {len(uploaded_files)} file(s)" if len(uploaded_files) > 1 else "Processing file"
            messages.info(request, f'{file_count_msg} for reporting period: {month_name} {selected_year}')
            
            # Queue async processing task
            try:
                from django_q.tasks import async_task
                from .tasks import process_uploaded_file
                task_id = async_task(
                    process_uploaded_file,
                    upload_session.id,
                    file_paths,  # Pass list of file paths
                    [f.name for f in uploaded_files],  # Pass list of original filenames
                    subscriber_id,
                    subscriber_name,
                    request.user.id,
                    selected_month,  # Pass reporting month to task
                    selected_year,   # Pass reporting year to task
                    task_name=f'process_file_{upload_session.id}'
                )
                
                upload_session.update_progress('upload_validation', 2, f'File(s) uploaded, queued for processing (Task ID: {task_id})...')
                
                logger.info(f"[UPLOAD] ✓ Queued async task {task_id} for UploadSession {upload_session.id}")
                logger.info(f"[UPLOAD] ℹ Files: {[f.name for f in uploaded_files]}")
                logger.info(f"[UPLOAD] ℹ Reporting Period: {month_name} {selected_year}")
                logger.info(f"[UPLOAD] ⚠ IMPORTANT: Make sure Q cluster worker is running!")
                logger.info(f"[UPLOAD] → Run in separate terminal: python manage.py qcluster")
                
            except Exception as e:
                logger.info(f"[UPLOAD] ✗ Failed to queue async task: {e}")
                upload_session.update_progress('upload_validation', 2, f'File uploaded (sync fallback - worker not available)')
            
            # Immediately redirect to progress tracking page
            return redirect('auto:progress_tracking', session_id=upload_session.id)
            
    else:
        form = ExcelUploadForm()
    
    # Add subscriber context for all users
    context = {'form': form}
    if request.user.is_authenticated:
        from acctmgt.models import UserProfile
        from .models import UploadSession, Subscriber
        
        # Check if user is in multi_subscriber group
        is_multi_subscriber = request.user.groups.filter(name='multi_subscriber').exists()
        context['is_multi_subscriber'] = is_multi_subscriber
        
        if is_multi_subscriber:
            # Provide list of all subscribers for dropdown
            context['subscribers'] = Subscriber.objects.all().order_by('subscriber_name')
        else:
            # Regular user: show bound subscriber info
            user_profile = UserProfile.get_or_create_profile(request.user)
            if user_profile.is_bound:
                bound_subscriber = user_profile.get_bound_subscriber()
                if bound_subscriber:
                    context['bound_subscriber'] = bound_subscriber
                    context['subscriber_name'] = bound_subscriber.subscriber_name
                    context['subscriber_id'] = bound_subscriber.subscriber_id
        
        # Always check for last completed upload session
        last_upload = UploadSession.objects.filter(
            user=request.user,
            status='completed'
        ).order_by('-completed_at').first()
        
        # Set context flags for template
        context['has_results'] = bool(last_upload and last_upload.individual_file_path)
        context['current_view'] = request.GET.get('view', 'upload')
        
        # If viewing results, load the data from the last completed upload session
        if context['current_view'] == 'results' and last_upload:
            # Debug logging to track download links issue
            logger.info(f"[RESULTS VIEW] Loading results for multi_subscriber={is_multi_subscriber}")
            logger.info(f"[RESULTS VIEW] upload_session.id={last_upload.id}, individual_file_path={last_upload.individual_file_path}, corporate_file_path={last_upload.corporate_file_path}")
            
            # Force refresh from database to ensure we have the latest file paths
            # (fixes issue where async task updates DB but cached object has stale data)
            last_upload.refresh_from_db()
            logger.info(f"[RESULTS VIEW] After refresh: individual_file_path={last_upload.individual_file_path}, corporate_file_path={last_upload.corporate_file_path}")
            
            # Load processing data from the upload session
            if last_upload.processing_data:
                try:
                    processing_data = json.loads(last_upload.processing_data)
                    context['processing_stats'] = processing_data.get('processing_stats', [])
                    context['success_message'] = True

                    context['upload_session'] = last_upload
                    
                    # Calculate totals from the stored DataFrames
                    split_indi_json = processing_data.get('split_indi')
                    split_corpo_json = processing_data.get('split_corpo')
                    
                    if split_indi_json:
                        split_indi = pd.read_json(split_indi_json, orient='split')
                        context['total_individual'] = len(split_indi)
                    else:
                        context['total_individual'] = last_upload.individual_records
                    
                    if split_corpo_json:
                        split_corpo = pd.read_json(split_corpo_json, orient='split')
                        context['total_corporate'] = len(split_corpo)
                    else:
                        context['total_corporate'] = last_upload.corporate_records
                    
                    # Credit unmerged data
                    context['total_credit_unmerged_count'] = last_upload.unmatched_credit_records or 0
                    
                    # Excluded records data
                    context['excluded_individual'] = last_upload.excluded_individual_records or 0
                    context['excluded_corporate'] = last_upload.excluded_corporate_records or 0
                    
                except Exception as e:
                    logger.info(f"[UPLOAD] Error loading results data: {e}")
                    # Fallback to basic metrics from UploadSession
                    context['total_individual'] = last_upload.individual_records
                    context['total_corporate'] = last_upload.corporate_records
                    context['total_credit_unmerged_count'] = last_upload.unmatched_credit_records or 0
                    context['excluded_individual'] = last_upload.excluded_individual_records or 0
                    context['excluded_corporate'] = last_upload.excluded_corporate_records or 0
                    context['success_message'] = True
                    context['upload_session'] = last_upload
    
    return render(request, 'upload.html', context)

def clean_for_output(df):
    # Canonical null values (all lowercase) - single source of truth
    NULL_VALUES = {
        'n/a', 'n.a', 'n.a.', 'none', 'nan', 'null', '#n/a', 
        'nil', 'nill', 'na', 'unknown', 'blank', 'missing',
        'not available', 'not applicable', '-', '--', '---', '?'
    }
    
    # Convert all columns to string
    for col in df.columns:
        df[col] = df[col].astype(str)
    
    # Case-insensitive null value cleaning
    df = df.apply(lambda col: col.apply(
        lambda x: '' if str(x).strip().lower() in NULL_VALUES else x
    ))
    return df

def enforce_string_columns(df):
    for col in df.columns:
        df[col] = df[col].astype(str)
    return df

def reorder_consumer_columns(columns):
    """
    Reorder columns to place ORIGINAL_BUSINESSNAME at the beginning for better visibility,
    DATEOFBIRTH after SURNAME, and move FIRSTNAME and MIDDLENAME to the end.
    
    Args:
        columns (list): List of column names
        
    Returns:
        list: Reordered list of column names
    """
    # Convert to list if it's a pandas Index
    columns = list(columns)
    
    # Define the columns we want to move to the end
    columns_to_end = ['FIRSTNAME', 'MIDDLENAME']
    
    # Remove the columns we want to move to the end
    remaining_columns = [col for col in columns if col not in columns_to_end]
    
    # Handle ORIGINAL_BUSINESSNAME placement at the beginning (case-insensitive)
    original_businessname_col = None
    for col in remaining_columns:
        if col.upper() == 'ORIGINAL_BUSINESSNAME':
            original_businessname_col = col
            break
    
    if original_businessname_col:
        # Remove ORIGINAL_BUSINESSNAME from its current position
        remaining_columns.remove(original_businessname_col)
        # Insert at the beginning
        remaining_columns.insert(3, original_businessname_col)
    
    # Handle DATEOFBIRTH placement after SURNAME
    if 'DATEOFBIRTH' in remaining_columns and 'SURNAME' in remaining_columns:
        # Remove DATEOFBIRTH from its current position
        remaining_columns.remove('DATEOFBIRTH')
        
        # Find the index of SURNAME and insert DATEOFBIRTH after it
        surname_index = remaining_columns.index('SURNAME')
        remaining_columns.insert(surname_index + 1, 'DATEOFBIRTH')
    
    # Add the columns that should be at the end
    for col in columns_to_end:
        if col in columns:
            remaining_columns.append(col)
    
    return remaining_columns


def reorder_commercial_columns(columns):
    """
    Reorder columns to place ORIGINAL_BUSINESSNAME after BUSINESSREGISTRATIONNUMBER,
    and move FIRSTNAME, MIDDLENAME, and DEPENDANTS to the end.
    
    Args:
        columns (list): List of column names
        
    Returns:
        list: Reordered list of column names
    """
    # Convert to list if it's a pandas Index
    columns = list(columns)
    
    # Define the columns we want to move to the end
    columns_to_end = ['FIRSTNAME', 'MIDDLENAME', 'DEPENDANTS']
    
    # Remove the columns we want to move to the end
    remaining_columns = [col for col in columns if col not in columns_to_end]
    
    # Handle ORIGINAL_BUSINESSNAME placement after BUSINESSREGISTRATIONNUMBER (case-insensitive)
    original_businessname_col = None
    businessreg_col = None
    
    for col in remaining_columns:
        if col.upper() == 'ORIGINAL_BUSINESSNAME':
            original_businessname_col = col
        elif col.upper() == 'BUSINESSREGISTRATIONNUMBER':
            businessreg_col = col
    
    if original_businessname_col and businessreg_col:
        # Remove ORIGINAL_BUSINESSNAME from its current position
        remaining_columns.remove(original_businessname_col)
        
        # Find the index of BUSINESSREGISTRATIONNUMBER and insert ORIGINAL_BUSINESSNAME after it
        brn_index = remaining_columns.index(businessreg_col)
        remaining_columns.insert(brn_index + 1, original_businessname_col)
    
    # Add the columns that should be at the end
    for col in columns_to_end:
        if col in columns:
            remaining_columns.append(col)
    
    return remaining_columns




def transform_to_commercial(df, columns_to_clear):
    """
    Transform individual records to commercial format and clear guarantor data.
    """
    if df.empty:
        return df
    
    df_copy = df.copy()
    
    # Use ORIGINAL_BUSINESSNAME if available, otherwise reconstruct from components
    if 'BUSINESSNAME' not in df_copy.columns:
        if 'ORIGINAL_BUSINESSNAME' in df_copy.columns:
            df_copy['BUSINESSNAME'] = df_copy['ORIGINAL_BUSINESSNAME']
        else:
            df_copy['BUSINESSNAME'] = (
                df_copy['SURNAME'].fillna('') + ' '
                + df_copy['FIRSTNAME'].fillna('') + ' '
                + df_copy['MIDDLENAME'].fillna('')
            ).str.strip()
            
    # --- Clear data from the specified guarantor columns ---
    for col in columns_to_clear:
        if col in df_copy.columns:
            df_copy[col] = ''
    # ---

    # Drop individual name, dependant, and temporary columns for corporate format
    columns_to_remove = ['SURNAME', 'FIRSTNAME', 'MIDDLENAME', 'DEPENDANTS', 'ORIGINAL_BUSINESSNAME']
    df_copy = df_copy.drop(columns=[col for col in columns_to_remove if col in df_copy.columns], errors='ignore')
    
    df_copy = rename_columns(df_copy, ConsuToComm.copy())
    df_copy = enforce_string_columns(df_copy)
    
    return df_copy


def transform_to_consumer(df, columns_to_clear):
    """
    Transform commercial records to consumer format and clear Principal Officer data.
    """
    if df.empty:
        return df
    
    df_copy = df.copy()

    # --- Clear data from the specified Principal Officer columns ---
    for col in columns_to_clear:
        if col in df_copy.columns:
            df_copy[col] = ''
    # ---

    # Apply column mapping from commercial to consumer
    df_copy = rename_columns(df_copy, CommToConsu.copy())
    df_copy = enforce_string_columns(df_copy)
    
    return df_copy
# 



@login_required
@csrf_exempt
def verify_split_decision(request):
    from .models import UploadSession as UploadSessionModel
    
    # Handle GET request with session_id (load data from UploadSession into request.session)
    if request.method == 'GET' and 'session_id' in request.GET:
        try:
            session_id = int(request.GET['session_id'])
            upload_session = UploadSessionModel.objects.get(id=session_id, user=request.user)
            
            # Load processing data from UploadSession
            if upload_session.processing_data:
                processing_data = json.loads(upload_session.processing_data)
                
                # Check if using new Parquet-based storage
                using_parquet = processing_data.get('using_parquet', False)
                
                if using_parquet:
                    # New flow: Only store verification candidates and metadata in session
                    # Full DataFrames will be loaded from Parquet when needed
                    logger.info(f"[[VERIFICATION] Loading session data (Parquet mode) for upload session {session_id}")
                else:
                    # Legacy flow: Load from JSON (for backward compatibility with old uploads)
                    logger.info(f"[[VERIFICATION] Loading session data (JSON mode - legacy) for upload session {session_id}")
                    request.session['split_indi'] = processing_data.get('split_indi')
                    request.session['split_corpo'] = processing_data.get('split_corpo')
                    request.session['indi'] = processing_data.get('split_indi')
                    request.session['corpo'] = processing_data.get('split_corpo')
                
                # Populate request.session with verification candidates and metadata
                request.session['split_candidates_commercial'] = pd.DataFrame(
                    processing_data.get('commercial_candidates', [])
                ).to_json(orient='split')
                request.session['split_candidates_consumer'] = pd.DataFrame(
                    processing_data.get('consumer_candidates', [])
                ).to_json(orient='split')
                request.session['columns_commercial'] = processing_data.get('columns_commercial', [])
                request.session['columns_consumer'] = processing_data.get('columns_consumer', [])
                request.session['processing_stats'] = processing_data.get('processing_stats', [])
                request.session['original_filename'] = processing_data.get('original_filename', '')
                request.session['subscriber_id'] = processing_data.get('subscriber_id')
                request.session['subscriber_name'] = processing_data.get('subscriber_name')
                request.session['individual_credit_unmerged'] = processing_data.get('individual_credit_unmerged', {})
                request.session['corporate_credit_unmerged'] = processing_data.get('corporate_credit_unmerged', {})
                request.session['upload_session_id'] = upload_session.id
                request.session['commercial_candidates'] = processing_data.get('commercial_candidates', [])
                request.session['consumer_candidates'] = processing_data.get('consumer_candidates', [])
                request.session['using_parquet'] = using_parquet  # Flag for POST handler
                
                # Save session data
                request.session.modified = True
                
                # Redirect to verification page WITHOUT session_id parameter
                # This ensures the session data is properly loaded on the next request
                return redirect('auto:verify_split_decision')
            else:
                messages.error(request, 'Processing data not found. Please restart the upload.')
                return redirect('auto:upload')
                
        except (UploadSessionModel.DoesNotExist, ValueError, json.JSONDecodeError):
            messages.error(request, 'Invalid session or data corrupted.')
            return redirect('auto:upload')
    
    # Handle GET request to display verification form
    if request.method == 'GET':
        commercial_candidates = request.session.get('commercial_candidates', [])
        consumer_candidates = request.session.get('consumer_candidates', [])
        columns_commercial = request.session.get('columns_commercial', [])
        columns_consumer = request.session.get('columns_consumer', [])
        
        return render(request, 'verify_split.html', {
            'commercial_candidates': commercial_candidates,
            'consumer_candidates': consumer_candidates,
            'columns_commercial': columns_commercial,
            'columns_consumer': columns_consumer,
            'upload_session_id': request.session.get('upload_session_id'),
        })
    
    # Handle POST request (user's verification decisions)
    if request.method == 'POST':
        try:
            # Validate required session keys exist
            required_session_keys = ['upload_session_id']
            missing_keys = [key for key in required_session_keys if key not in request.session]
            
            if missing_keys:
                messages.error(request, f'Session data missing: {missing_keys}. Please restart the process.')
                return redirect('auto:upload')
            
            # Get user checkbox moves from POST (lists of booleans)
            commercial_moves = json.loads(request.POST.get('commercial_moves', '[]'))
            consumer_moves = json.loads(request.POST.get('consumer_moves', '[]'))
            
            # Debug logging
            logger.info(f"[VERIFICATION] Commercial moves count: {len(commercial_moves)}")
            logger.info(f"[VERIFICATION] Consumer moves count: {len(consumer_moves)}")
            
            # Load upload session
            upload_session_id = request.session.get('upload_session_id')
            if not upload_session_id:
                messages.error(request, 'Session expired. Please upload the file again.')
                return redirect('auto:upload')
            
            # Get the upload session
            from .models import UploadSession as UploadSessionModel
            upload_session = UploadSessionModel.objects.get(id=upload_session_id, user=request.user)
            
            # Store verification decisions in processing_data for async processing
            processing_data = json.loads(upload_session.processing_data or '{}')
            processing_data['commercial_moves'] = commercial_moves
            processing_data['consumer_moves'] = consumer_moves
            processing_data['verification_submitted'] = True
            upload_session.processing_data = json.dumps(processing_data)
            
            # CRITICAL: Save processing_data BEFORE mark_verification_completed()
            # mark_verification_completed() uses save(update_fields=[...]) which doesn't include processing_data
            # Without this, the async task loads OLD data without verification decisions!
            upload_session.save(update_fields=['processing_data'])
            
            # Mark verification as completed (status transition)
            upload_session.mark_verification_completed()
            
            logger.info(f"[VERIFICATION] Stored verification decisions for session {upload_session_id}")
            logger.info(f"[VERIFICATION] Queuing async post-verification processing")
            
            # Queue async task for post-verification processing (all heavy work happens here)
            from django_q.tasks import async_task
            
            processing_stats = request.session.get('processing_stats', [])
            original_filename = request.session.get('original_filename', 'output')
            subscriber_id = request.session.get('subscriber_id')
            subscriber_name = request.session.get('subscriber_name')
            
            task_id = async_task(
                'auto.tasks.process_post_verification',
                upload_session_id,
                True,  # use_parquet=True
                processing_stats,
                original_filename,
                subscriber_id,
                subscriber_name,
                0,  # total_individual_records - calculated in task
                0   # total_corporate_records - calculated in task
            )
            
            logger.info(f"[VERIFICATION] Post-verification task queued with ID: {task_id}")
            
            # Clean up session verification data
            request.session.pop('commercial_candidates', None)
            request.session.pop('consumer_candidates', None)
            request.session.pop('columns_commercial', None)
            request.session.pop('columns_consumer', None)
            request.session.pop('consumer_data', None)
            request.session.pop('commercial_data', None)
            request.session.pop('processing_stats', None)
            request.session.pop('individual_credit_merged', None)
            request.session.pop('corporate_credit_merged', None)
            request.session.pop('individual_credit_unmerged', None)
            request.session.pop('corporate_credit_unmerged', None)
            request.session.pop('split_candidates_commercial', None)
            request.session.pop('split_candidates_consumer', None)
            
            # Redirect to progress tracking page immediately
            return redirect('auto:progress_tracking', session_id=upload_session_id)
        except Exception as e:
            # Log the full error for debugging
            logger.info("An error occurred in verify_split_decision:")
            logger.info(traceback.format_exc())
            
            # Show error message and redirect to upload page
            messages.error(request, 'An error occurred during verification. Please try uploading the file again.')
            return redirect('auto:upload')
    else:
        return render(request, 'upload.html', {'form': ExcelUploadForm(), 'error_message': 'Invalid request.'})

def clean_and_deduplicate_columns(df):
    """Clean column names and assign suffixes to duplicates after cleaning."""
    cleaned_cols = [remove_special_characters(str(col)).upper().strip() for col in df.columns]
    counts = {}
    new_cols = []
    for col in cleaned_cols:
        if col in counts:
            counts[col] += 1
            new_cols.append(f"{col}{counts[col]}")
        else:
            counts[col] = 0
            new_cols.append(col)
    df.columns = new_cols
    return df


@login_required
def dashboard(request):
    """
    Main dashboard view displaying upload statistics and user engagement metrics
    """
    from .models import UploadSession, SubscriberUtils
    from acctmgt.models import UserProfile
    from django.utils import timezone
    from datetime import timedelta
    
    # Get user profile and subscriber information
    user_profile = UserProfile.get_or_create_profile(request.user)
    
    # Check if user is multi-subscriber (they don't need binding)
    is_multi_subscriber = request.user.groups.filter(name='multi_subscriber').exists()
    
    if is_multi_subscriber:
        # Multi-subscriber users can view dashboard without specific binding
        bound_subscriber = None
    else:
        # Regular users must be bound to a subscriber
        if not user_profile.is_bound:
            messages.error(request, 'Your account is not bound to any organization. Please complete the binding process.')
            return redirect('acctmgt:subscriber_selection')
        
        bound_subscriber = user_profile.get_bound_subscriber()
        if not bound_subscriber:
            messages.error(request, 'Unable to retrieve your organization information. Please contact support.')
            return redirect('acctmgt:subscriber_selection')
    
    # Calculate date ranges
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Get upload statistics
    today_stats = UploadSession.get_user_stats(request.user, days=1)
    weekly_stats = UploadSession.get_user_stats(request.user, days=7)
    monthly_stats = UploadSession.get_user_stats(request.user, days=30)
    overall_stats = UploadSession.get_user_stats(request.user)
    
    # Get recent uploads
    recent_uploads = UploadSession.get_recent_uploads(request.user, limit=10)
    
    # Prepare context data
    context = {
        'user_profile': user_profile,
        'subscriber': bound_subscriber,
        'today_stats': today_stats,
        'weekly_stats': weekly_stats,
        'monthly_stats': monthly_stats,
        'overall_stats': overall_stats,
        'recent_uploads': recent_uploads,
        'current_date': today,
    }
    
    return render(request, 'dashboard.html', context)


@login_required
@csrf_exempt
def dashboard_api(request):
    """
    API endpoint for real-time dashboard data updates
    """
    from .models import UploadSession
    from django.utils import timezone
    
    if request.method == 'GET':
        # Get time range from query parameters
        time_range = request.GET.get('range', 'today')
        
        if time_range == 'today':
            stats = UploadSession.get_user_stats(request.user, days=1)
        elif time_range == 'week':
            stats = UploadSession.get_user_stats(request.user, days=7)
        elif time_range == 'month':
            stats = UploadSession.get_user_stats(request.user, days=30)
        else:
            stats = UploadSession.get_user_stats(request.user)
        
        # Get recent uploads
        recent_uploads = UploadSession.get_recent_uploads(request.user, limit=10)
        
        # Format recent uploads for JSON response
        uploads_data = []
        for upload in recent_uploads:
            uploads_data.append({
                'id': upload.id,
                'filename': upload.original_filename,
                'status': upload.status,
                'uploaded_at': upload.uploaded_at.isoformat(),
                'total_records': upload.total_records,
                'individual_records': upload.individual_records,
                'corporate_records': upload.corporate_records,
                'processing_time': upload.processing_time,
            })
        
        return JsonResponse({
            'stats': stats,
            'recent_uploads': uploads_data,
            'timestamp': timezone.now().isoformat()
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
@csrf_exempt
def delete_upload(request, upload_id):
    """
    Delete an upload session and its associated files.
    Only allows deletion by the user who created the upload.
    """
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get the upload session, ensuring it belongs to the current user
        upload_session = UploadSession.objects.get(
            id=upload_id,
            user=request.user
        )
        
        # Store file paths before deletion for cleanup
        file_paths_to_delete = []
        
        # Collect all file paths that need to be deleted
        if upload_session.individual_file_path:
            file_paths_to_delete.append(upload_session.individual_file_path)
        if upload_session.corporate_file_path:
            file_paths_to_delete.append(upload_session.corporate_file_path)
        if upload_session.individual_txt_path:
            file_paths_to_delete.append(upload_session.individual_txt_path)
        if upload_session.corporate_txt_path:
            file_paths_to_delete.append(upload_session.corporate_txt_path)
        
        # Also try to delete the original uploaded file
        if upload_session.filename:
            original_file_path = os.path.join(settings.MEDIA_ROOT, 'excel', upload_session.filename)
            if os.path.exists(original_file_path):
                file_paths_to_delete.append(original_file_path)
        
        # Delete the database record first
        upload_session.delete()
        
        # Clean up associated files
        deleted_files = []
        failed_deletions = []
        
        # Delete the original archived file if it exists
        if upload_session.original_file_path:
            try:
                delete_success, delete_error = delete_original_file(upload_session.original_file_path)
                if not delete_success:
                    failed_deletions.append(f"Original file: {delete_error}")
            except Exception as e:
                failed_deletions.append(f"Original file: {str(e)}")
        
        for file_path in file_paths_to_delete:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    deleted_files.append(file_path)
            except OSError as e:
                failed_deletions.append(f"{file_path}: {str(e)}")
        
        response_data = {
            'success': True,
            'message': 'Upload deleted successfully',
            'deleted_files_count': len(deleted_files)
        }
        
        # Include warning if some files couldn't be deleted
        if failed_deletions:
            response_data['warning'] = f"Some files could not be deleted: {', '.join(failed_deletions)}"
        
        return JsonResponse(response_data)
        
    except UploadSession.DoesNotExist:
        return JsonResponse({
            'error': 'Upload not found or you do not have permission to delete it'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'error': f'An error occurred while deleting the upload: {str(e)}'
        }, status=500)


@login_required
@csrf_exempt
def cancel_upload(request, upload_id):
    """
    Cancel an upload session that is stuck or in progress.
    Only allows cancellation of non-completed uploads by the user who created them.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get the upload session, ensuring it belongs to the current user
        upload_session = UploadSession.objects.get(
            id=upload_id,
            user=request.user
        )
        
        # Only allow cancellation of non-completed uploads
        if upload_session.status in ['completed', 'cancelled']:
            return JsonResponse({
                'error': f'Cannot cancel an upload that is already {upload_session.status}'
            }, status=400)
        
        # Mark as cancelled
        upload_session.mark_cancelled(reason='Cancelled by user from dashboard')
        
        return JsonResponse({
            'success': True,
            'message': 'Upload cancelled successfully'
        })
        
    except UploadSession.DoesNotExist:
        return JsonResponse({
            'error': 'Upload not found or you do not have permission to cancel it'
        }, status=404)
    
    except Exception as e:
        return JsonResponse({
            'error': f'An error occurred while cancelling the upload: {str(e)}'
        }, status=500)


def home(request):
    """
    Home view that redirects users based on authentication status.
    Authenticated users go to dashboard, unauthenticated users go to login.
    """
    if request.user.is_authenticated:
        return redirect('auto:dashboard')
    else:
        return redirect('acctmgt:login')




def build_excel_report(upload_session, user=None):
    """
    Build Excel report workbook as BytesIO buffer (no request dependency).
    
    Args:
        upload_session: UploadSession instance
        user: User instance (optional, for report metadata)
    
    Returns:
        tuple: (BytesIO buffer, filename string)
    """
    from .tasks import load_from_parquet
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from io import BytesIO
    
    # Get subscriber info using the model's method
    subscriber = upload_session.get_subscriber()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # ============================================================
    # SHEET 1: PROCESSING SUMMARY
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = "Processing Summary"
    
    # Styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=10, color="1F4E78")
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # Color fills for excluded records highlighting
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red for missing cells
    reason_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")  # Light yellow for exclusion reason
    reason_header_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Orange for reason header
    

    row = 1
    
    # Title
    ws_summary.merge_cells(f'A{row}:B{row}')
    title_cell = ws_summary[f'A{row}']
    title_cell.value = "FCB AUTO PROCESSING REPORT"
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    title_cell.alignment = Alignment(horizontal='center')
    row += 2
    
    # Report Information Section
    ws_summary[f'A{row}'] = "REPORT INFORMATION"
    ws_summary[f'A{row}'].font = section_font
    row += 1
    
    # Build user info for report
    if user:
        generated_by = f"{user.first_name} {user.last_name}" if user.first_name else user.username
        user_email = user.email or 'Not provided'
    else:
        generated_by = 'System (Auto-generated)'
        user_email = 'N/A'
    
    report_data = [
        ['Generated On:', upload_session.completed_at.strftime('%B %d, %Y at %I:%M %p') if upload_session.completed_at else 'N/A'],
        ['Generated By:', generated_by],
        ['User Email:', user_email],
        ['Report ID:', f"RPT-{upload_session.id:06d}"]
    ]
        
    for label, value in report_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True, size=10)
        row += 1
    
    row += 1
    
    # Subscriber Information Section
    ws_summary[f'A{row}'] = "SUBSCRIBER INFORMATION"
    ws_summary[f'A{row}'].font = section_font
    row += 1
    
    subscriber_data = [
        ['Subscriber Name:', subscriber.subscriber_name if subscriber else 'Unknown'],
        ['Subscriber ID:', str(upload_session.subscriber_id)],
        ['Processing Date:', upload_session.uploaded_at.strftime('%B %d, %Y') if upload_session.uploaded_at else 'N/A'],
        ['Processing Time:', upload_session.completed_at.strftime('%I:%M %p') if upload_session.completed_at else 'N/A']
    ]
    
    for label, value in subscriber_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True, size=10)
        row += 1
    
    row += 1
    
    # File Processing Details Section
    ws_summary[f'A{row}'] = "FILE PROCESSING DETAILS"
    ws_summary[f'A{row}'].font = section_font
    row += 1
    
    processing_data = [
        ['Original Filename:', upload_session.original_filename or 'N/A'],
        ['Processing Duration:', f"{upload_session.processing_time:.2f} seconds" if upload_session.processing_time else 'Not recorded'],
        ['Processing Status:', upload_session.get_status_display()],
        ['Processing Method:', 'Automated Excel Processing with Data Validation']
    ]
    
    for label, value in processing_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True, size=10)
        row += 1
    
    row += 1
    
    # Processing Results Section (replacing Credit Records Merge Analysis)
    ws_summary[f'A{row}'] = "PROCESSING RESULTS"
    ws_summary[f'A{row}'].font = section_font
    row += 1
    
    # Get credit analysis data
    total_credit_unmerged = upload_session.unmatched_credit_records or 0
    individual_matched = upload_session.individual_records or 0
    corporate_matched = upload_session.corporate_records or 0
    total_credits_processed = individual_matched + corporate_matched + total_credit_unmerged
    total_matched = individual_matched + corporate_matched
    
    # Get excluded record counts
    excluded_individual = upload_session.excluded_individual_records or 0
    excluded_corporate = upload_session.excluded_corporate_records or 0
    total_excluded = excluded_individual + excluded_corporate
    
    processing_results_data = [
        ['Individual Borrowers:', f"{individual_matched:,} records"],
        ['Corporate Borrowers:', f"{corporate_matched:,} records"],
        ['Total Credits Processed:', f"{total_credits_processed:,} records"],
        ['Matched Credits:', f"{total_matched:,} records"],
        ['Unmatched Credits:', f"{total_credit_unmerged:,} records"],
        ['', ''],  # Blank row for spacing
        ['EXCLUDED RECORDS (Data Quality):', ''],
        ['Individual Excluded:', f"{excluded_individual:,} records"],
        ['Corporate Excluded:', f"{excluded_corporate:,} records"],
        ['Total Excluded:', f"{total_excluded:,} records"],
    ]
    
    for label, value in processing_results_data:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        if label and not label.startswith('EXCLUDED'):
            ws_summary[f'A{row}'].font = Font(bold=True, size=10)
        elif label.startswith('EXCLUDED'):
            ws_summary[f'A{row}'].font = Font(bold=True, size=10, color="C65911")
        row += 1
    
    row += 1
    
    # Footer
    ws_summary.merge_cells(f'A{row}:D{row}')
    footer_cell = ws_summary[f'A{row}']
    footer_cell.value = f"This report was automatically generated by FCB Auto Processing System on {upload_session.completed_at.strftime('%B %d, %Y') if upload_session.completed_at else 'N/A'}."
    footer_cell.font = Font(size=9, italic=True, color="666666")
    footer_cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths for summary sheet
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 20
    
    # ============================================================
    # SHEET 2: UNMATCHED CREDITS
    # ============================================================
    ws_unmatched = wb.create_sheet(title="Unmatched Credits")
    
    # Load unmatched credits from Parquet
    try:
        unmatched_credits = load_from_parquet(upload_session.id, 'unmatched_credits')
        logger.info(f"[EXCEL REPORT] Loaded {len(unmatched_credits)} unmatched credits from Parquet")
    except FileNotFoundError as e:
        # If no Parquet file, create empty DataFrame
        logger.info(f"[EXCEL REPORT] Parquet file not found: {e}")
        unmatched_credits = pd.DataFrame(columns=['CUSTOMERID', 'ACCOUNTNUMBER', 'ACCOUNTSTATUS', 'CREDITLIMIT', 'AVAILEDLIMIT'])
    except Exception as e:
        # Catch any other errors
        logger.info(f"[EXCEL REPORT] Error loading unmatched credits: {e}")
        unmatched_credits = pd.DataFrame(columns=['CUSTOMERID', 'ACCOUNTNUMBER', 'ACCOUNTSTATUS', 'CREDITLIMIT', 'AVAILEDLIMIT'])
    
    # Title
    ws_unmatched.merge_cells('A1:E1')
    title_cell = ws_unmatched['A1']
    title_cell.value = "Unmatched Credit Records"
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    title_cell.alignment = Alignment(horizontal='center')
    
    # Description
    ws_unmatched.merge_cells('A2:E2')
    desc_cell = ws_unmatched['A2']
    desc_cell.value = f"These credit records could not be matched to any borrower (Individual or Corporate). Total Records: {len(unmatched_credits):,}"
    desc_cell.font = Font(size=10, italic=True)
    desc_cell.alignment = Alignment(horizontal='center')
    
    # Add unmatched credits data
    if not unmatched_credits.empty:
        # Write headers
        for col_idx, col_name in enumerate(unmatched_credits.columns, start=1):
            cell = ws_unmatched.cell(row=4, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Write data rows
        for row_idx, row_data in enumerate(dataframe_to_rows(unmatched_credits, index=False, header=False), start=5):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_unmatched.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # Format numeric columns
                if col_idx in [4, 5]:  # CREDITLIMIT and AVAILEDLIMIT
                    try:
                        cell.number_format = '#,##0.00'
                    except Exception:
                        pass
    else:
        # No unmatched credits
        ws_unmatched.merge_cells('A4:E4')
        no_data_cell = ws_unmatched['A4']
        no_data_cell.value = "✓ All credit records were successfully matched to borrowers!"
        no_data_cell.font = Font(size=12, color="008000", bold=True)
        no_data_cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths for unmatched credits sheet
    ws_unmatched.column_dimensions['A'].width = 20  # CUSTOMERID
    ws_unmatched.column_dimensions['B'].width = 20  # ACCOUNTNUMBER
    ws_unmatched.column_dimensions['C'].width = 20  # ACCOUNTSTATUS
    ws_unmatched.column_dimensions['D'].width = 18  # CREDITLIMIT
    ws_unmatched.column_dimensions['E'].width = 18  # AVAILEDLIMIT
    
    # ============================================================
    # SHEET 3: EXCLUDED RECORDS
    # ============================================================
    ws_excluded = wb.create_sheet(title="Excluded Records")
    
    # Load excluded records from Parquet
    try:
        excluded_individual_df = load_from_parquet(upload_session.id, 'excluded_individual')
        logger.info(f"[EXCEL REPORT] Loaded {len(excluded_individual_df)} excluded individual records from Parquet")
    except FileNotFoundError:
        excluded_individual_df = pd.DataFrame()
    except Exception as e:
        logger.info(f"[EXCEL REPORT] Error loading excluded individual records: {e}")
        excluded_individual_df = pd.DataFrame()
    
    try:
        excluded_corporate_df = load_from_parquet(upload_session.id, 'excluded_corporate')
        logger.info(f"[EXCEL REPORT] Loaded {len(excluded_corporate_df)} excluded corporate records from Parquet")
    except FileNotFoundError:
        excluded_corporate_df = pd.DataFrame()
    except Exception as e:
        logger.info(f"[EXCEL REPORT] Error loading excluded corporate records: {e}")
        excluded_corporate_df = pd.DataFrame()
    
    current_row = 1
    
    # ============================================================
    # SECTION 1: EXCLUDED INDIVIDUAL BORROWERS
    # ============================================================
    # Section title
    ws_excluded.merge_cells(f'A{current_row}:F{current_row}')
    title_cell = ws_excluded[f'A{current_row}']
    title_cell.value = "EXCLUDED INDIVIDUAL BORROWERS"
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Description
    ws_excluded.merge_cells(f'A{current_row}:F{current_row}')
    desc_cell = ws_excluded[f'A{current_row}']
    desc_cell.value = f"Records excluded due to missing validation columns. Total: {len(excluded_individual_df):,} records"
    desc_cell.font = Font(size=10, italic=True)
    current_row += 2
    
    # Individual data
    if not excluded_individual_df.empty:
        # Get column positions for highlighting
        columns_list = list(excluded_individual_df.columns)
        exclusion_reason_col_idx = columns_list.index('EXCLUSION_REASON') + 1 if 'EXCLUSION_REASON' in columns_list else None
        
        # Map validated fields to their column indices (for individual borrowers)
        individual_validated_columns = {
            'AVAILEDLIMIT': columns_list.index('AVAILEDLIMIT') + 1 if 'AVAILEDLIMIT' in columns_list else None,
            'DATEOFBIRTH': columns_list.index('DATEOFBIRTH') + 1 if 'DATEOFBIRTH' in columns_list else None,
            'BVNNUMBER': columns_list.index('BVNNUMBER') + 1 if 'BVNNUMBER' in columns_list else None,
            'NATIONALIDENTITYNUMBER': columns_list.index('NATIONALIDENTITYNUMBER') + 1 if 'NATIONALIDENTITYNUMBER' in columns_list else None,
            'LOANEFFECTIVEDATE': columns_list.index('LOANEFFECTIVEDATE') + 1 if 'LOANEFFECTIVEDATE' in columns_list else None,
        }
        
        # Write headers
        for col_idx, col_name in enumerate(excluded_individual_df.columns, start=1):
            cell = ws_excluded.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            # Use orange header for EXCLUSION_REASON column
            if col_name == 'EXCLUSION_REASON':
                cell.fill = reason_header_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        current_row += 1
        
        # Write data rows with color highlighting for missing cells
        for row_idx, df_row in excluded_individual_df.iterrows():
            exclusion_reason = str(df_row.get('EXCLUSION_REASON', ''))
            
            for col_idx, col_name in enumerate(columns_list, start=1):
                value = df_row[col_name]
                cell = ws_excluded.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                
                # Apply highlighting based on column and exclusion reason
                if col_name == 'EXCLUSION_REASON':
                    cell.fill = reason_fill
                    cell.font = Font(size=10, bold=True)
                else:
                    # Check if this column is mentioned in the exclusion reason as missing
                    value_is_empty = pd.isna(value) or str(value).strip() == ''
                    
                    # Highlight if cell is empty AND column is related to a missing field
                    should_highlight = False
                    if value_is_empty:
                        if col_name == 'AVAILEDLIMIT' and 'AVAILEDLIMIT' in exclusion_reason:
                            should_highlight = True
                        elif col_name == 'DATEOFBIRTH' and 'DATEOFBIRTH' in exclusion_reason:
                            should_highlight = True
                        elif col_name == 'BVNNUMBER' and 'BVN/NIN' in exclusion_reason:
                            should_highlight = True
                        elif col_name == 'NATIONALIDENTITYNUMBER' and 'BVN/NIN' in exclusion_reason:
                            should_highlight = True
                        elif col_name == 'LOANEFFECTIVEDATE' and 'LOANEFFECTIVEDATE' in exclusion_reason:
                            should_highlight = True
                    
                    if should_highlight:
                        cell.fill = missing_fill
            
            current_row += 1
    else:
        ws_excluded.merge_cells(f'A{current_row}:F{current_row}')
        no_data_cell = ws_excluded[f'A{current_row}']
        no_data_cell.value = "✓ No individual records were excluded"
        no_data_cell.font = Font(size=10, color="008000", bold=True)
        current_row += 1
    
    # Add spacing between sections
    current_row += 3
    
    # ============================================================
    # SECTION 2: EXCLUDED CORPORATE BORROWERS
    # ============================================================
    # Section title
    ws_excluded.merge_cells(f'A{current_row}:F{current_row}')
    title_cell = ws_excluded[f'A{current_row}']
    title_cell.value = "EXCLUDED CORPORATE BORROWERS"
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='left')
    current_row += 1
    
    # Description
    ws_excluded.merge_cells(f'A{current_row}:F{current_row}')
    desc_cell = ws_excluded[f'A{current_row}']
    desc_cell.value = f"Records excluded due to missing validation columns. Total: {len(excluded_corporate_df):,} records"
    desc_cell.font = Font(size=10, italic=True)
    current_row += 2
    
    # Corporate data
    if not excluded_corporate_df.empty:
        # Get column positions for highlighting
        columns_list_corp = list(excluded_corporate_df.columns)
        
        # Write headers
        for col_idx, col_name in enumerate(excluded_corporate_df.columns, start=1):
            cell = ws_excluded.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = header_font
            # Use orange header for EXCLUSION_REASON column
            if col_name == 'EXCLUSION_REASON':
                cell.fill = reason_header_fill
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        current_row += 1
        
        # Write data rows with color highlighting for missing cells
        for row_idx, df_row in excluded_corporate_df.iterrows():
            exclusion_reason = str(df_row.get('EXCLUSION_REASON', ''))
            
            for col_idx, col_name in enumerate(columns_list_corp, start=1):
                value = df_row[col_name]
                cell = ws_excluded.cell(row=current_row, column=col_idx, value=value)
                cell.border = border
                
                # Apply highlighting based on column and exclusion reason
                if col_name == 'EXCLUSION_REASON':
                    cell.fill = reason_fill
                    cell.font = Font(size=10, bold=True)
                else:
                    # Check if this column is mentioned in the exclusion reason as missing
                    value_is_empty = pd.isna(value) or str(value).strip() == ''
                    
                    # Highlight if cell is empty AND column is related to a missing field
                    should_highlight = False
                    if value_is_empty:
                        if col_name == 'AVAILEDLIMIT' and 'AVAILEDLIMIT' in exclusion_reason:
                            should_highlight = True
                        elif col_name == 'BUSINESSNAME' and 'BUSINESSNAME' in exclusion_reason:
                            should_highlight = True
                        elif col_name == 'LOANEFFECTIVEDATE' and 'LOANEFFECTIVEDATE' in exclusion_reason:
                            should_highlight = True
                    
                    if should_highlight:
                        cell.fill = missing_fill
            
            current_row += 1
    else:
        ws_excluded.merge_cells(f'A{current_row}:F{current_row}')
        no_data_cell = ws_excluded[f'A{current_row}']
        no_data_cell.value = "✓ No corporate records were excluded"
        no_data_cell.font = Font(size=10, color="008000", bold=True)
        current_row += 1
    
    # ============================================================
    # SAVE AND RETURN EXCEL FILE
    # ============================================================
    
    # Create filename
    filename = f"FCB_Report_{upload_session.id}_{upload_session.uploaded_at.strftime('%Y%m%d') if upload_session.uploaded_at else 'unknown'}.xlsx"
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer, filename


@login_required
def generate_excel_report(request, session_id):
    """
    View wrapper for build_excel_report — returns HTTP download response.
    """
    try:
        upload_session = UploadSession.objects.filter(
            id=session_id,
            user=request.user
        ).first()
        
        if not upload_session:
            return HttpResponse("Upload session not found or access denied.", status=404)
        
        buffer, filename = build_excel_report(upload_session, user=request.user)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.info(f"[EXCEL REPORT ERROR] {error_details}")
        return HttpResponse(f"Error generating Excel report: {str(e)}", status=500)


@login_required
@csrf_exempt
def engagement_metrics_api(request):
    """
    API endpoint for user engagement metrics
    """

    
    if request.method == 'GET':
        # Get upload frequency data for the last 30 days
        thirty_days_ago = timezone.now() - timedelta(days=30)
        
        # Get basic upload statistics instead of complex date queries
        # This avoids SQL Server compatibility issues with date/strftime functions
        total_uploads = UploadSession.objects.filter(
            user=request.user,
            uploaded_at__gte=thirty_days_ago
        ).count()
        
        completed_uploads = UploadSession.objects.filter(
            user=request.user,
            uploaded_at__gte=thirty_days_ago,
            status='completed'
        ).count()
        
        # Get actual upload data for daily grouping
        uploads_queryset = UploadSession.objects.filter(
            user=request.user,
            uploaded_at__gte=thirty_days_ago
        ).values('uploaded_at')
        
        # Group uploads by date using Python to avoid SQL compatibility issues
        from collections import defaultdict
        daily_counts = defaultdict(int)
        
        # Process each upload and group by date
        for upload in uploads_queryset:
            upload_date = upload['uploaded_at'].date()
            daily_counts[upload_date] += 1
        
        # Create daily uploads data for the last 7 days (for chart)
        daily_uploads = []
        today = timezone.now().date()
        
        for i in range(6, -1, -1):  # Last 7 days
            date = today - timedelta(days=i)
            count = daily_counts.get(date, 0)
            daily_uploads.append({
                'date': date.isoformat(),
                'count': count,
                'label': date.strftime('%a, %b %d')
            })
        
        # For now, keep these empty as they're not used in the chart
        weekly_pattern = []
        hourly_pattern = []
        
        # Average processing time
        avg_processing_time = UploadSession.objects.filter(
            user=request.user,
            status='completed',
            processing_time__isnull=False
        ).aggregate(avg_time=Avg('processing_time'))['avg_time'] or 0
        
        return JsonResponse({
            'total_uploads': total_uploads,
            'completed_uploads': completed_uploads,
            'daily_uploads': daily_uploads,
            'weekly_pattern': weekly_pattern,
            'hourly_pattern': hourly_pattern,
            'avg_processing_time': round(avg_processing_time, 2),
            'timestamp': timezone.now().isoformat()
        })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def progress_tracking(request, session_id):
    """
    View to display real-time progress tracking page for file processing
    """
    try:
        # Get upload session - ensure user has access
        upload_session = UploadSession.objects.get(
            id=session_id,
            user=request.user
        )
        
        # Get subscriber name
        subscriber_name = None
        if upload_session.subscriber_id:
            try:
                from .models import Subscriber
                subscriber = Subscriber.objects.get(subscriber_id=upload_session.subscriber_id)
                subscriber_name = subscriber.subscriber_name
            except Exception:
                subscriber_name = f"Subscriber {upload_session.subscriber_id}"
        
        context = {
            'upload_session': upload_session,
            'subscriber_name': subscriber_name,
        }
        
        return render(request, 'progress_tracking.html', context)
        
    except UploadSession.DoesNotExist:
        messages.error(request, 'Upload session not found or access denied.')
        return redirect('auto:upload')


@login_required
def progress_api(request, session_id):
    """
    API endpoint to return progress data as JSON for polling
    """
    try:
        # Get upload session - ensure user has access
        upload_session = UploadSession.objects.get(
            id=session_id,
            user=request.user
        )
        
        # Prepare response data
        data = {
            'id': upload_session.id,
            'status': upload_session.status,
            'processing_stage': upload_session.processing_stage,
            'progress': upload_session.progress_percentage,
            'current_message': upload_session.current_message,
            'error_message': upload_session.error_message,
            
            # Processing metrics
            'individual_records': upload_session.individual_records,
            'corporate_records': upload_session.corporate_records,
            'total_records': upload_session.total_records,
            'individual_credit_matched': upload_session.individual_credit_matched,
            'corporate_credit_matched': upload_session.corporate_credit_matched,
            'unmatched_credit_records': upload_session.unmatched_credit_records,
            
            # Verification data
            'has_verification_candidates': upload_session.has_verification_candidates,
            'commercial_candidates_count': upload_session.commercial_candidates_count,
            'consumer_candidates_count': upload_session.consumer_candidates_count,
            'verification_skipped': upload_session.verification_skipped,
            
            # File paths
            'individual_file_path': upload_session.individual_file_path,
            'corporate_file_path': upload_session.corporate_file_path,
            'individual_txt_path': upload_session.individual_txt_path,
            'corporate_txt_path': upload_session.corporate_txt_path,
            
            # Activity log
            'activity_log': upload_session.activity_log if isinstance(upload_session.activity_log, list) else [],
            
            # Timestamps
            'uploaded_at': upload_session.uploaded_at.isoformat() if upload_session.uploaded_at else None,
            'processing_started_at': upload_session.processing_started_at.isoformat() if upload_session.processing_started_at else None,
            'completed_at': upload_session.completed_at.isoformat() if upload_session.completed_at else None,
            'verification_completed_at': upload_session.verification_completed_at.isoformat() if upload_session.verification_completed_at else None,
        }
        
        # Add friendly error information if there's an error
        if upload_session.error_message:
            try:
                from .error_messages import get_friendly_error
                friendly_error = get_friendly_error(upload_session.error_message)
                data['friendly_error'] = friendly_error
            except ImportError:
                pass
        
        return JsonResponse(data)
        
    except UploadSession.DoesNotExist:
        return JsonResponse({'error': 'Upload session not found or access denied'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@login_required
def submit_feedback(request):
    """
    Handle feedback submission from the in-app feedback modal.
    Accepts POST with JSON body containing rating, category, message.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Parse JSON body
        data = json.loads(request.body)
        
        rating = data.get('rating')
        category = data.get('category', 'general')
        message = data.get('message', '').strip()
        page_url = data.get('page_url', '')
        
        # Validate required fields
        if not rating or not isinstance(rating, int) or rating < 1 or rating > 5:
            return JsonResponse({'error': 'Please provide a valid rating (1-5)'}, status=400)
        
        if not message:
            return JsonResponse({'error': 'Please provide feedback message'}, status=400)
        
        # Validate category
        valid_categories = ['bug', 'feature', 'general']
        if category not in valid_categories:
            category = 'general'
        
        # Import and create feedback
        from .models import Feedback
        
        feedback = Feedback.objects.create(
            user=request.user if request.user.is_authenticated else None,
            rating=rating,
            category=category,
            message=message,
            page_url=page_url
        )
        
        logger.info(f"[FEEDBACK] New feedback submitted: {feedback.category} - {feedback.rating}★ from {request.user.username}")
        
        # Send email notification
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            # Get subscriber name if user is bound to one
            subscriber_name = 'Not bound'
            try:
                from acctmgt.models import UserProfile
                profile = UserProfile.objects.filter(user=request.user, is_bound=True).first()
                if profile:
                    subscriber = profile.get_bound_subscriber()
                    if subscriber:
                        subscriber_name = subscriber.subscriber_name
            except Exception:
                pass
            
            star_display = '★' * rating + '☆' * (5 - rating)
            email_subject = f'[Feedback] {feedback.get_category_display()} - {star_display}'
            email_body = f"""New feedback received from the Data Processing Suite:

User: {request.user.username if request.user.is_authenticated else 'Anonymous'}
Subscriber: {subscriber_name}
Rating: {star_display} ({rating}/5)
Category: {feedback.get_category_display()}
Page: {page_url}
Time: {feedback.created_at.strftime('%Y-%m-%d %H:%M:%S') if feedback.created_at else 'N/A'}

Message:
{message}

---
View all feedback at: /admin/auto/feedback/
"""
            send_mail(
                subject=email_subject,
                message=email_body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[settings.FEEDBACK_EMAIL_RECIPIENT],
                fail_silently=True  # Don't break submission if email fails
            )
            logger.info(f"[FEEDBACK] Email notification sent to {settings.FEEDBACK_EMAIL_RECIPIENT}")
        except Exception as email_error:
            logger.warning(f"[FEEDBACK] Email notification failed: {str(email_error)}")
        
        return JsonResponse({
            'success': True,
            'message': 'Thank you for your feedback!'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"[FEEDBACK] Error saving feedback: {str(e)}")
        return JsonResponse({'error': 'Failed to save feedback. Please try again.'}, status=500)
