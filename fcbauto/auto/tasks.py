import os
import logging
import glob
import shutil
import json
import time
import gc

logger = logging.getLogger(__name__)
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from datetime import datetime as dt
from .models import UploadSession
from .exceptions import (
    DataValidationError, FileProcessingError, MergeError,
    OutputGenerationError, VerificationError
)
from .map import (
    consu_mapping, comm_mapping, credit_mapping, guar_mapping, prin_mapping,
    consumer_merged_mapping, commercial_merged_mapping
)
from .views import (
    # Import all the processing functions
    ensure_all_sheets_exist, clean_sheet_name, convert_numpy,
    remove_special_characters, make_column_names_unique,
    preprocess_tenor_from_headers, preprocess_arrears_from_headers, rename_columns_with_fuzzy_rapidfuzz,
    process_dates, process_names, replace_ampersands, process_special_characters,
    process_nationality, process_gender, process_states, process_marital_status,
    process_borrower_type, process_employment_status, process_phone_columns,
    process_title, process_account_status, process_loan_type, process_currency,
    process_repayment, process_classification, process_collateral_type,
    process_loan_tenor, clear_previous_info_columns, process_numeric_columns,
    fill_data_column, fill_depend_column, process_identity_numbers,
    process_passport_number, process_business_id, process_bvn_number,
    process_DriversLicense, process_otherid, process_tax_numbers,
    process_collateral_details, positioninBusiness, trim_strings_to_59,
    remove_duplicates, merge_individual_borrowers, merge_corporate_borrowers,
    split_commercial_entities, split_consumer_entities, reorder_commercial_columns,
    reorder_consumer_columns, extract_date_from_filename, clean_business_name
)
from .header_detection import find_header_row


# Constants for chunked processing
LARGE_SHEET_THRESHOLD = 50000  # Process sheets with >50k rows in chunks
CHUNK_SIZE = 50000  # Rows per chunk for large sheets


# ============================================================================
# COLUMN VALIDATION FUNCTIONS (Data Quality Checks)
# ============================================================================

def is_column_populated(df, column_name):
    """
    Check if a column has at least one non-empty value in the entire dataset.
    Uses optimized pandas operations with early termination for performance.
    
    Args:
        df (pd.DataFrame): DataFrame to check
        column_name (str): Name of column to validate
    
    Returns:
        bool: True if column has at least one non-empty value, False otherwise
    """
    # Check if column exists
    if column_name not in df.columns:
        return False
    
    # Quick null check first (early exit if all null)
    if not df[column_name].notna().any():
        return False
    
    # Drop null values for efficiency
    non_null_values = df[column_name].dropna()
    if len(non_null_values) == 0:
        return False
    
    # Check for non-empty strings (early exit on first match)
    return (non_null_values.astype(str).str.strip() != '').any()


def validate_required_columns(processed_sheets):
    """
    Validate required columns after fuzzy mapping/column renaming.
    Checks business rules for each sheet type.
    
    Validation Rules:
    - Credit Information: CUSTOMERID + ACCOUNTNUMBER both required
    - Consumer/Commercial Merged: ACCOUNTNUMBER required
    - Individual Borrower: CUSTOMERID required + (SURNAME or FIRSTNAME) required
    - Corporate Borrower: CUSTOMERID required
    - Principal Officers: CUSTOMERID required
    - Guarantors Information: Skip validation
    - Auto-generated (empty): Skip validation
    
    Args:
        processed_sheets (dict): Dictionary of sheet_name -> DataFrame
    
    Returns:
        tuple: (is_valid: bool, error_messages: list of str)
    """
    error_messages = []
    
    for sheet_key, df in processed_sheets.items():
        # Skip auto-generated empty sheets (headers only, no data)
        if df.empty:
            logger.info(f"[[VALIDATION] Skipping auto-generated empty sheet: {sheet_key}")
            continue
        
        # Skip guarantors information
        if sheet_key == 'guarantorsinformation':
            logger.info(f"[[VALIDATION] Skipping guarantors sheet: {sheet_key}")
            continue
        
        sheet_errors = []
        
        # Rule 1: Consumer/Commercial Merged - ACCOUNTNUMBER required
        if sheet_key in ['consumermerged', 'commercialmerged']:
            if not is_column_populated(df, 'ACCOUNTNUMBER'):
                sheet_errors.append("ACCOUNTNUMBER column is empty")
        
        # Rule 2: Credit Information - At least one of CUSTOMERID or ACCOUNTNUMBER required
        # (Merge logic can use ACCOUNTNUMBER as fallback when CUSTOMERID is missing)
        # Also require AVAILEDLIMIT to be populated
        elif sheet_key == 'creditinformation':
            has_customerid = is_column_populated(df, 'CUSTOMERID')
            has_accountnumber = is_column_populated(df, 'ACCOUNTNUMBER')
            has_availedlimit = is_column_populated(df, 'AVAILEDLIMIT')
            
            if not has_customerid and not has_accountnumber:
                sheet_errors.append("Both CUSTOMERID and ACCOUNTNUMBER columns are empty (at least one required for merging)")
            
            if not has_availedlimit:
                sheet_errors.append("AVAILEDLIMIT column is empty (required for credit records)")
        
        # Rule 3: Individual Borrower - CUSTOMERID + (SURNAME or FIRSTNAME) + BVNNUMBER required
        elif sheet_key == 'individualborrowertemplate':
            if not is_column_populated(df, 'CUSTOMERID'):
                sheet_errors.append("CUSTOMERID column is empty")
            
            has_surname = is_column_populated(df, 'SURNAME')
            has_firstname = is_column_populated(df, 'FIRSTNAME')
            
            if not has_surname and not has_firstname:
                sheet_errors.append("SURNAME and FIRSTNAME columns are both empty (at least one required)")
            
            # Check BVN: If empty in individual borrower, check if it's available in credit info
            if not is_column_populated(df, 'BVNNUMBER'):
                # Check if credit information has BVNNUMBER as fallback
                credit_df = processed_sheets.get('creditinformation', pd.DataFrame())
                if credit_df.empty or not is_column_populated(credit_df, 'BVNNUMBER'):
                    sheet_errors.append("BVNNUMBER column is empty")
        
        # Rule 4: Corporate Borrower - CUSTOMERID required
        elif sheet_key == 'corporateborrowertemplate':
            if not is_column_populated(df, 'CUSTOMERID'):
                sheet_errors.append("CUSTOMERID column is empty")
        
        # Rule 5: Principal Officers - CUSTOMERID required
        elif sheet_key == 'principalofficerstemplate':
            if not is_column_populated(df, 'CUSTOMERID'):
                sheet_errors.append("CUSTOMERID column is empty")
        
        # Add sheet errors to master list
        if sheet_errors:
            # Format sheet name for display
            sheet_display_name = sheet_key.replace('template', ' Template').replace('information', ' Information').title()
            # Format: "Sheet Name: error1. error2."
            formatted_errors = ". ".join(sheet_errors) + "."
            error_messages.append(f"{sheet_display_name}: {formatted_errors}")
    
    # Return validation result
    if error_messages:
        full_error_message = (
            "❌ Validation Failed - Required Columns Are Empty.\n"
            + "\n".join(error_messages)
            + "\nPlease correct these issues and re-upload your file."
        )
        return False, full_error_message
    
    return True, "✓ All required columns validated successfully"


# ============================================================================
# DATA QUALITY VALIDATION FUNCTIONS (Pre-Output Validation)
# ============================================================================

def validate_individual_record(row):
    """
    Validate an individual borrower record for data quality.
    
    Rule: All of the following must be present:
          - AVAILEDLIMIT
          - DATEOFBIRTH
          - LOANEFFECTIVEDATE
          (BVN/NIN are optional)
    
    Args:
        row: DataFrame row (Series)
    
    Returns:
        tuple: (is_valid: bool, reason: str or None)
    """
    availedlimit = row.get('AVAILEDLIMIT', '')
    dob = row.get('DATEOFBIRTH', '')
    loan_effective_date = row.get('LOANEFFECTIVEDATE', '')
    
    # Check if each field is empty/missing
    availedlimit_empty = pd.isna(availedlimit) or str(availedlimit).strip() == ''
    dob_empty = pd.isna(dob) or str(dob).strip() == ''
    loan_effective_date_empty = pd.isna(loan_effective_date) or str(loan_effective_date).strip() == ''
    
    # Build exclusion reason based on what's missing
    missing_fields = []
    
    if availedlimit_empty:
        missing_fields.append('AVAILEDLIMIT')
    
    # DATEOFBIRTH is mandatory
    if dob_empty:
        missing_fields.append('DATEOFBIRTH')
    
    # LOANEFFECTIVEDATE is mandatory
    if loan_effective_date_empty:
        missing_fields.append('LOANEFFECTIVEDATE')
    
    # Record is invalid if any required field is missing
    if missing_fields:
        return False, f"Missing {', '.join(missing_fields)}"
    
    return True, None




def validate_corporate_record(row):
    """
    Validate a corporate borrower record for data quality.
    
    Rule: All of the following must be present:
          - AVAILEDLIMIT
          - BUSINESSNAME
          - LOANEFFECTIVEDATE
    
    Args:
        row: DataFrame row (Series)
    
    Returns:
        tuple: (is_valid: bool, reason: str or None)
    """
    availedlimit = row.get('AVAILEDLIMIT', '')
    businessname = row.get('BUSINESSNAME', '')
    loan_effective_date = row.get('LOANEFFECTIVEDATE', '')
    
    # Check if each field is empty/missing
    availedlimit_empty = pd.isna(availedlimit) or str(availedlimit).strip() == ''
    businessname_empty = pd.isna(businessname) or str(businessname).strip() == ''
    loan_effective_date_empty = pd.isna(loan_effective_date) or str(loan_effective_date).strip() == ''
    
    # Build exclusion reason based on what's missing
    missing_fields = []
    
    if availedlimit_empty:
        missing_fields.append('AVAILEDLIMIT')
    
    if businessname_empty:
        missing_fields.append('BUSINESSNAME')
    
    if loan_effective_date_empty:
        missing_fields.append('LOANEFFECTIVEDATE')
    
    # Record is invalid if any required field is missing
    if missing_fields:
        return False, f"Missing {', '.join(missing_fields)}"
    
    return True, None




def apply_data_quality_validation(indi_df, corpo_df):
    """
    Apply data quality validation to individual and corporate DataFrames.
    
    Separates valid records from excluded records based on validation rules.
    This should be called AFTER remove_duplicates() to ensure no duplicate
    records in the excluded set.
    
    Args:
        indi_df: DataFrame of individual borrower records
        corpo_df: DataFrame of corporate borrower records
    
    Returns:
        tuple: (
            valid_indi: DataFrame of valid individual records,
            valid_corpo: DataFrame of valid corporate records,
            excluded_indi: DataFrame of excluded individual records (with EXCLUSION_REASON),
            excluded_corpo: DataFrame of excluded corporate records (with EXCLUSION_REASON)
        )
    """
    logger.info("[DATA QUALITY] Starting data quality validation...")
    
    # Validate Individual Records
    if not indi_df.empty:
        indi_validation = indi_df.apply(validate_individual_record, axis=1)
        indi_is_valid = indi_validation.apply(lambda x: x[0])
        indi_reasons = indi_validation.apply(lambda x: x[1])
        
        valid_indi = indi_df[indi_is_valid].copy().reset_index(drop=True)
        excluded_indi = indi_df[~indi_is_valid].copy().reset_index(drop=True)
        if not excluded_indi.empty:
            excluded_indi['EXCLUSION_REASON'] = indi_reasons[~indi_is_valid].values
        logger.info(f"[[DATA QUALITY] Individual: {len(valid_indi)} valid, {len(excluded_indi)} excluded")
    else:
        valid_indi = indi_df.copy()
        excluded_indi = pd.DataFrame()
        logger.info("[DATA QUALITY] Individual: No records to validate")
    
    # Validate Corporate Records
    if not corpo_df.empty:
        corpo_validation = corpo_df.apply(validate_corporate_record, axis=1)
        corpo_is_valid = corpo_validation.apply(lambda x: x[0])
        corpo_reasons = corpo_validation.apply(lambda x: x[1])
        
        valid_corpo = corpo_df[corpo_is_valid].copy().reset_index(drop=True)
        excluded_corpo = corpo_df[~corpo_is_valid].copy().reset_index(drop=True)
        if not excluded_corpo.empty:
            excluded_corpo['EXCLUSION_REASON'] = corpo_reasons[~corpo_is_valid].values
        logger.info(f"[[DATA QUALITY] Corporate: {len(valid_corpo)} valid, {len(excluded_corpo)} excluded")
    else:
        valid_corpo = corpo_df.copy()
        excluded_corpo = pd.DataFrame()
        logger.info("[DATA QUALITY] Corporate: No records to validate")
    
    return valid_indi, valid_corpo, excluded_indi, excluded_corpo


# ============================================================================
# PARQUET TEMP STORAGE FUNCTIONS (Memory Management for Large Files)
# ============================================================================

def get_temp_parquet_dir(session_id):
    """
    Get the temporary Parquet directory path for a session
    
    Args:
        session_id: UploadSession ID
    
    Returns:
        Path to temp directory
    """
    temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp', str(session_id))
    os.makedirs(temp_dir, exist_ok=True)
    return temp_dir


def save_to_parquet(df, session_id, sheet_name):
    """
    Save DataFrame to temporary Parquet file
    
    Args:
        df: DataFrame to save (can be empty but must have proper columns)
        session_id: UploadSession ID
        sheet_name: Name identifier for the sheet (e.g., 'consumer', 'credit', 'corporate')
    
    Returns:
        Path to saved Parquet file
    """
    temp_dir = get_temp_parquet_dir(session_id)
    parquet_path = os.path.join(temp_dir, f'{sheet_name}_processed.parquet')
    
    # Log saving status - important for empty auto-generated sheets
    row_status = f"({len(df)} rows)" if not df.empty else "(empty but with headers)"
    logger.info(f"[[PARQUET] Saving {sheet_name} to {parquet_path} {row_status}")
    
    df.to_parquet(parquet_path, engine='pyarrow', compression='snappy', index=False)
    logger.info(f"[[PARQUET] Successfully saved {sheet_name}")
    
    return parquet_path


def load_from_parquet(session_id, sheet_name, columns=None):
    """
    Load DataFrame from temporary Parquet file
    
    Args:
        session_id: UploadSession ID
        sheet_name: Name identifier for the sheet
        columns: Optional list of specific columns to load (memory optimization)
    
    Returns:
        DataFrame loaded from Parquet
    """
    temp_dir = get_temp_parquet_dir(session_id)
    parquet_path = os.path.join(temp_dir, f'{sheet_name}_processed.parquet')
    
    if not os.path.exists(parquet_path):
        raise FileNotFoundError(f"Parquet file not found: {parquet_path}")
    
    logger.info(f"[[PARQUET] Loading {sheet_name} from {parquet_path}" + 
          (f" (columns: {columns})" if columns else ""))
    
    df = pd.read_parquet(parquet_path, engine='pyarrow', columns=columns)
    logger.info(f"[[PARQUET] Successfully loaded {sheet_name} ({len(df)} rows)")
    
    return df


def cleanup_temp_files(session_id, keep_unmatched_credits=True):
    """
    Delete temporary Parquet files for a session, optionally keeping 
    unmatched_credits and excluded records for Excel report generation
    
    Args:
        session_id: UploadSession ID
        keep_unmatched_credits: If True, preserve unmatched_credits.parquet 
                                and excluded_*.parquet for Excel report generation
    """

    
    temp_dir = get_temp_parquet_dir(session_id)
    
    if os.path.exists(temp_dir):
        if keep_unmatched_credits:
            # Files to preserve for Excel report generation
            files_to_preserve = [
                os.path.normpath(os.path.join(temp_dir, 'unmatched_credits_processed.parquet')),
                os.path.normpath(os.path.join(temp_dir, 'excluded_individual_processed.parquet')),
                os.path.normpath(os.path.join(temp_dir, 'excluded_corporate_processed.parquet')),
            ]
            
            # Delete individual files except preserved ones
            files_to_delete = glob.glob(os.path.join(temp_dir, '*_processed.parquet'))
            
            logger.info(f"[[PARQUET] Cleanup with keep_unmatched_credits=True")
            logger.info(f"[[PARQUET] Files to preserve: {[os.path.basename(f) for f in files_to_preserve]}")
            logger.info(f"[[PARQUET] Files found: {len(files_to_delete)}")
            
            deleted_count = 0
            preserved_count = 0
            for file_path in files_to_delete:
                normalized_path = os.path.normpath(file_path)
                if normalized_path in files_to_preserve:
                    logger.info(f"[[PARQUET] ✓ Preserving: {os.path.basename(file_path)}")
                    preserved_count += 1
                else:
                    try:
                        os.remove(file_path)
                        deleted_count += 1
                        logger.info(f"[[PARQUET] ✗ Deleted: {os.path.basename(file_path)}")
                    except Exception as e:
                        logger.info(f"[[PARQUET] Error deleting {file_path}: {e}")
            
            logger.info(f"[[PARQUET] Cleanup complete: deleted {deleted_count} files, preserved {preserved_count} files")
        else:
            # Delete entire directory
            logger.info(f"[[PARQUET] Cleaning up entire temp directory: {temp_dir}")
            shutil.rmtree(temp_dir)
            logger.info(f"[[PARQUET] Cleanup complete")
    else:
        logger.info(f"[[PARQUET] No temp directory to cleanup for session {session_id}")


# ============================================================================
# SHEET PROCESSING FUNCTIONS
# ============================================================================

def preprocess_sheet_columns(sheet_data, cleaned_name):
    """
    Preprocess a sheet by cleaning and normalizing column names with fuzzy mapping.
    This is step 1 of processing - it makes columns consistent so sheets from
    different files can be safely concatenated before full processing.
    
    Args:
        sheet_data: DataFrame containing sheet data
        cleaned_name: Cleaned sheet name for mapping detection
    
    Returns:
        DataFrame with normalized column names (ready for concatenation)
    """
    cleaned_df = sheet_data.copy()
    
    # Comprehensive null value cleaning (case-insensitive)
    # Base null values - will be matched case-insensitively
    null_values_base = [
        'n/a', 'n.a', 'none', 'nan', 'null', '#n/a', 'nil', 'nill', 'na',
        'unknown', 'blank', 'n.a.', 'not available', 'not applicable', 
        'missing', '-', '--', '---'
    ]
    
    # Apply case-insensitive null value cleaning to all columns
    for col in cleaned_df.columns:
        # Convert column to string first
        cleaned_df[col] = cleaned_df[col].astype(str)
        # Case-insensitive replacement: check if stripped lowercase value is in null list
        cleaned_df[col] = cleaned_df[col].apply(
            lambda x: '' if str(x).strip().lower() in null_values_base else x
        )
    
    # Clean Excel XML escape sequences (_x000D_ = carriage return, _x000A_ = line feed)
    # These appear when Excel stores line breaks in XML format
    import re
    xml_escape_pattern = re.compile(r'_x[0-9A-Fa-f]{4}_')
    for col in cleaned_df.columns:
        cleaned_df[col] = cleaned_df[col].apply(
            lambda x: xml_escape_pattern.sub(' ', str(x)).strip() if isinstance(x, str) else x
        )
    
    # Clean column names
    cleaned_df.columns = [str(col).upper().strip() for col in cleaned_df.columns]
    cleaned_df = preprocess_tenor_from_headers(cleaned_df)
    cleaned_df = preprocess_arrears_from_headers(cleaned_df)  # Extract numbers, convert months to days
    cleaned_df.columns = [remove_special_characters(col) for col in cleaned_df.columns]
    cleaned_df = make_column_names_unique(cleaned_df)
    cleaned_df.columns = [remove_special_characters(col) for col in cleaned_df.columns]
    
    # Fuzzy column mapping based on sheet type
    if cleaned_name == 'individualborrowertemplate':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, consu_mapping)
    elif cleaned_name == 'corporateborrowertemplate':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, comm_mapping)
    elif cleaned_name == 'principalofficerstemplate':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, prin_mapping)
    elif cleaned_name == 'creditinformation':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, credit_mapping)
    elif cleaned_name == 'guarantorsinformation':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, guar_mapping)
    elif cleaned_name == 'consumermerged':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, consumer_merged_mapping)
    elif cleaned_name == 'commercialmerged':
        cleaned_df = rename_columns_with_fuzzy_rapidfuzz(cleaned_df, commercial_merged_mapping)
    
    return cleaned_df

def process_single_sheet(sheet_data, cleaned_name, cutoff_date=None, skip_preprocessing=False):
    """
    Apply all data transformation functions to a single sheet
    
    Args:
        sheet_data: DataFrame containing sheet data
        cleaned_name: Cleaned sheet name for mapping detection
        cutoff_date: Optional datetime object for LOANEFFECTIVEDATE/ACCOUNTSTATUSDATE validation
        skip_preprocessing: If True, skip column preprocessing (already done for multi-file merge)
    
    Returns:
        Processed DataFrame
    """
    # Step 1: Preprocessing (column normalization, fuzzy mapping)
    # Skip if already done (e.g., for merged multi-file data)
    if skip_preprocessing:
        cleaned_df = sheet_data.copy()
    else:
        cleaned_df = preprocess_sheet_columns(sheet_data, cleaned_name)
    
    # Step 2: Apply all data cleaning transformations
    cleaned_df = process_dates(cleaned_df, cutoff_date=cutoff_date)
    cleaned_df = process_names(cleaned_df)
    cleaned_df = clean_business_name(cleaned_df)  # Clean repetitive numbers like 0, 000, 11111
    cleaned_df = replace_ampersands(cleaned_df)
    cleaned_df = process_special_characters(cleaned_df)
    cleaned_df = process_nationality(cleaned_df)
    cleaned_df = process_gender(cleaned_df)
    cleaned_df = process_states(cleaned_df)
    cleaned_df = process_marital_status(cleaned_df)
    cleaned_df = process_borrower_type(cleaned_df)
    cleaned_df = process_employment_status(cleaned_df)
    cleaned_df = process_phone_columns(cleaned_df)
    cleaned_df = process_title(cleaned_df)
    cleaned_df = process_account_status(cleaned_df)
    cleaned_df = process_loan_type(cleaned_df)
    cleaned_df = process_currency(cleaned_df)
    cleaned_df = process_repayment(cleaned_df)
    cleaned_df = process_classification(cleaned_df)
    cleaned_df = process_collateral_type(cleaned_df)
    cleaned_df = process_loan_tenor(cleaned_df)
    cleaned_df = clear_previous_info_columns(cleaned_df)
    cleaned_df = process_numeric_columns(cleaned_df)
    cleaned_df = fill_data_column(cleaned_df)
    cleaned_df = fill_depend_column(cleaned_df)
    cleaned_df = process_identity_numbers(cleaned_df)
    cleaned_df = process_passport_number(cleaned_df)
    cleaned_df = process_business_id(cleaned_df)
    cleaned_df = process_bvn_number(cleaned_df)
    cleaned_df = process_DriversLicense(cleaned_df)
    cleaned_df = process_otherid(cleaned_df)
    cleaned_df = process_tax_numbers(cleaned_df)
    cleaned_df = process_collateral_details(cleaned_df)
    cleaned_df = positioninBusiness(cleaned_df)
    cleaned_df = trim_strings_to_59(cleaned_df)
    cleaned_df = remove_duplicates(cleaned_df)
    
    return cleaned_df


def process_large_sheet_chunked(file_path, sheet_name, cleaned_name, chunk_size=CHUNK_SIZE, cutoff_date=None):
    """
    Process large sheets (>50k rows) in chunks to avoid memory issues and timeouts.
    Uses openpyxl for streaming read to minimize memory footprint.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet to process
        cleaned_name: Cleaned sheet name for mapping detection
        chunk_size: Number of rows per chunk (default 50,000)
        cutoff_date: Optional datetime for LOANEFFECTIVEDATE/ACCOUNTSTATUSDATE validation
    
    Returns:
        Fully processed DataFrame
    """
    logger.info(f"[[CHUNKED PROCESSING] Processing large sheet '{sheet_name}' in chunks of {chunk_size} rows")
    
    processed_chunks = []
    
    # Detect header row before opening workbook (handles title rows, notes before headers)
    header_row = find_header_row(file_path, sheet_name)
    # Convert to 1-indexed for openpyxl (openpyxl uses 1-based row indices)
    header_row_1indexed = header_row + 1
    data_start_row = header_row_1indexed + 1
    
    logger.info(f"[CHUNKED PROCESSING] Header detected at row {header_row} (1-indexed: {header_row_1indexed})")
    
    # Open workbook in read-only mode for memory efficiency
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    
    # Get header row from detected row
    header = [cell.value for cell in sheet[header_row_1indexed]]
    logger.info(f"[CHUNKED PROCESSING] Header extracted: {len(header)} columns")
    
    chunk_data = []
    chunk_number = 0
    
    # Iterate through rows starting from data row (skip header and any rows before it)
    for row_index, row in enumerate(sheet.iter_rows(min_row=data_start_row), start=1):
        row_values = [cell.value for cell in row]
        chunk_data.append(row_values)
        
        # When chunk is full, process it
        if row_index % chunk_size == 0:
            chunk_number += 1
            logger.info(f"[[CHUNKED PROCESSING] Processing chunk {chunk_number} ({len(chunk_data)} rows)")
            
            # Create DataFrame from chunk
            chunk_df = pd.DataFrame(chunk_data, columns=header)
            
            # Convert all columns to string type
            for col in chunk_df.columns:
                chunk_df[col] = chunk_df[col].astype(str)
                chunk_df[col] = chunk_df[col].replace({'nan': '', 'None': '', 'NaN': ''})
            
            # Process this chunk
            processed_chunk = process_single_sheet(chunk_df, cleaned_name, cutoff_date=cutoff_date)
            processed_chunks.append(processed_chunk)
            
            # Clear chunk data and force garbage collection
            chunk_data = []
            del chunk_df
            gc.collect()
    
    # Process the final chunk if any rows remain
    if chunk_data:
        chunk_number += 1
        logger.info(f"[[CHUNKED PROCESSING] Processing final chunk {chunk_number} ({len(chunk_data)} rows)")
        
        chunk_df = pd.DataFrame(chunk_data, columns=header)
        
        # Convert all columns to string type
        for col in chunk_df.columns:
            chunk_df[col] = chunk_df[col].astype(str)
            chunk_df[col] = chunk_df[col].replace({'nan': '', 'None': '', 'NaN': ''})
        
        processed_chunk = process_single_sheet(chunk_df, cleaned_name, cutoff_date=cutoff_date)
        processed_chunks.append(processed_chunk)
        
        del chunk_df
        gc.collect()
    
    # Close workbook
    workbook.close()
    
    # Combine all processed chunks
    if not processed_chunks:
        logger.info(f"[[CHUNKED PROCESSING] Warning: No data chunks processed for sheet '{sheet_name}'")
        return pd.DataFrame(columns=header)
    
    logger.info(f"[[CHUNKED PROCESSING] Combining {len(processed_chunks)} processed chunks")
    final_result = pd.concat(processed_chunks, ignore_index=True)
    
    # Clean up chunk list
    del processed_chunks
    gc.collect()
    
    logger.info(f"[[CHUNKED PROCESSING] Completed processing of '{sheet_name}': {len(final_result)} total rows")
    return final_result


def get_sheet_row_count(file_path, sheet_name):
    """
    Quickly get the row count of a sheet without loading all data.
    
    Args:
        file_path: Path to Excel file
        sheet_name: Name of sheet
    
    Returns:
        Number of rows in sheet (excluding header)
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook[sheet_name]
        row_count = sheet.max_row - 1  # Subtract 1 for header row
        workbook.close()
        return row_count
    except Exception as e:
        logger.warning(f"[[ROW COUNT] Error getting row count for '{sheet_name}': {e}")
        return 0


def process_uploaded_file(upload_session_id, file_paths, original_filenames, 
                          subscriber_id, subscriber_name, user_id,
                          reporting_month=None, reporting_year=None):
    """
    Asynchronous task to process uploaded Excel file(s)
    
    Args:
        upload_session_id: ID of the UploadSession record
        file_paths: Path or list of paths to uploaded file(s) - accepts both for backward compatibility
        original_filenames: Original name(s) of uploaded file(s) - accepts both string and list
        subscriber_id: Subscriber ID
        subscriber_name: Subscriber name
        user_id: User ID who uploaded the file
        reporting_month: User-selected reporting month (1-12) or None
        reporting_year: User-selected reporting year or None
    """
    # Normalize inputs to lists for consistent handling (backward compatible)
    if isinstance(file_paths, str):
        file_paths = [file_paths]
    if isinstance(original_filenames, str):
        original_filenames = [original_filenames]
    
    file_count = len(file_paths)
    logger.info(f"[ASYNC TASK] Starting file processing for UploadSession {upload_session_id}")
    logger.info(f"[ASYNC TASK] Processing {file_count} file(s): {original_filenames}")
    logger.info(f"[ASYNC TASK] User-selected reporting period: {reporting_month}/{reporting_year}")
    
    try:
        # Get upload session
        upload_session = UploadSession.objects.get(id=upload_session_id)
        
        # Archive original file early in async processing (non-blocking)
        print(f"[ASYNC TASK] Archiving original file...")
        from .file_archival_utils import save_original_file
        try:
            with open(file_path, 'rb') as f:
                from django.core.files.base import File
                original_save_success, original_file_path, original_error = save_original_file(
                    File(f),
                    subscriber_id,
                    upload_session.uploaded_at
                )
                if original_save_success:
                    upload_session.original_file_path = original_file_path
                    upload_session.save(update_fields=['original_file_path'])
                    print(f"[ASYNC TASK] Original file archived: {original_file_path}")
                else:
                    print(f"[ASYNC TASK] Warning: Could not archive original file: {original_error}")
        except Exception as e:
            print(f"[ASYNC TASK] Warning: File archival error: {e}")
        
        # Update progress: Starting
        upload_session.update_progress('upload_validation', 5, f'File(s) uploaded successfully, starting processing...')
        upload_session.mark_processing_started()
        
        # Combine sheets from all files using worksheet names for type detection
        # This allows users to split data across multiple files
        all_sheet_info = {}  # {file_path: {sheet_name: {row_count, use_chunking}}}
        combined_sheet_names = []  # Track all unique sheet types
        
        for file_idx, file_path in enumerate(file_paths):
            logger.info(f"[ASYNC TASK] Analyzing file {file_idx + 1}/{file_count}: {original_filenames[file_idx]}")
            
            # Determine engine based on file extension (.xlsb requires pyxlsb)
            engine = 'pyxlsb' if file_path.lower().endswith('.xlsb') else None
            with pd.ExcelFile(file_path, engine=engine) as excel_file:
                sheet_names = excel_file.sheet_names
            
            logger.info(f"[ASYNC TASK] Found {len(sheet_names)} sheets in workbook")
            
            all_sheet_info[file_path] = {}
            for sheet_name in sheet_names:
                row_count = get_sheet_row_count(file_path, sheet_name)
                all_sheet_info[file_path][sheet_name] = {
                    'row_count': row_count,
                    'use_chunking': row_count > LARGE_SHEET_THRESHOLD,
                    'cleaned_name': clean_sheet_name(sheet_name)
                }
                combined_sheet_names.append(sheet_name)
                
                if row_count > LARGE_SHEET_THRESHOLD:
                    logger.info(f"[ASYNC TASK] Large sheet detected: '{sheet_name}' ({row_count:,} rows) - will use chunked processing")
                else:
                    logger.info(f"[ASYNC TASK] Normal sheet: '{sheet_name}' ({row_count:,} rows)")
        
        upload_session.update_progress('upload_validation', 10, 'Excel file(s) analyzed, loading sheets...')
        
        # ========================================================================
        # MULTI-FILE MERGE: Collect all sheets by type, preprocess, then concatenate
        # ========================================================================
        # Step 1: Collect ALL sheets from all files, grouped by type
        sheets_by_type = {}  # {cleaned_name: [(file_path, sheet_name, info), ...]}
        
        for file_path in file_paths:
            file_info = all_sheet_info[file_path]
            
            for sheet_name, info in file_info.items():
                cleaned_name = info['cleaned_name']
                
                # Collect ALL sheets by type (don't skip duplicates - we'll merge them!)
                if cleaned_name not in sheets_by_type:
                    sheets_by_type[cleaned_name] = []
                
                sheets_by_type[cleaned_name].append({
                    'file_path': file_path,
                    'sheet_name': sheet_name,
                    'info': info
                })
        
        # Log multi-file detection
        for cleaned_name, entries in sheets_by_type.items():
            if len(entries) > 1:
                file_names = [os.path.basename(e['file_path']) for e in entries]
                logger.info(f"[MULTI-FILE] Sheet type '{cleaned_name}' found in {len(entries)} files: {file_names} - will merge after preprocessing")
        
        # Step 2: Load, preprocess (fuzzy map), and concatenate same sheet types
        xds = {}  # Will contain preprocessed & merged DataFrames
        processing_stats = []
        preprocessed_flags = {}  # Track which sheets have been preprocessed
        large_sheet_entries = {}  # Track large sheet entries for later processing
        
        for cleaned_name, entries in sheets_by_type.items():
            preprocessed_dfs = []
            total_rows = 0
            first_sheet_name = entries[0]['sheet_name']  # Use first sheet's name as key
            has_large_sheet = False
            large_entries_for_type = []  # Collect all large sheet entries for this type
            
            for entry in entries:
                file_path = entry['file_path']
                sheet_name = entry['sheet_name']
                info = entry['info']
                total_rows += info['row_count']
                
                if info['use_chunking']:
                    # Collect large sheet entries for later processing
                    has_large_sheet = True
                    large_entries_for_type.append(entry)
                    logger.info(f"[ASYNC TASK] Large sheet '{sheet_name}' from {os.path.basename(file_path)} - will process in chunks")
                else:
                    # Load the sheet (normal-sized)
                    header_row = find_header_row(file_path, sheet_name)
                    logger.info(f"[ASYNC TASK] Loading '{sheet_name}' from {os.path.basename(file_path)} with header at row {header_row}")
                    
                    # Determine engine based on file extension (.xlsb requires pyxlsb)
                    engine = 'pyxlsb' if file_path.lower().endswith('.xlsb') else None
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row, na_filter=False, dtype=object, engine=engine)
                    
                    # Convert all columns to string and clean nulls
                    for col in df.columns:
                        df[col] = df[col].astype(str)
                        df[col] = df[col].replace({'nan': '', 'None': '', 'NaN': ''})
                    
                    # Preprocess: normalize columns with fuzzy mapping
                    preprocessed_df = preprocess_sheet_columns(df, cleaned_name)
                    preprocessed_dfs.append(preprocessed_df)
                    logger.info(f"[MULTI-FILE] Preprocessed '{sheet_name}': {len(preprocessed_df)} rows, {len(preprocessed_df.columns)} cols")
            
            # Handle large sheets: process each and merge
            if has_large_sheet:
                logger.info(f"[MULTI-FILE] Processing {len(large_entries_for_type)} large sheet(s) for '{cleaned_name}'")
                
                for entry in large_entries_for_type:
                    file_path = entry['file_path']
                    sheet_name = entry['sheet_name']
                    
                    # Process large sheet in chunks (includes preprocessing via process_single_sheet)
                    # Note: process_large_sheet_chunked calls process_single_sheet which does preprocessing
                    # We need to use preprocess_sheet_columns only for consistency with normal sheets
                    logger.info(f"[MULTI-FILE] Chunked processing '{sheet_name}' from {os.path.basename(file_path)}")
                    
                    # We'll store entries for later processing in the main loop
                    # to maintain correct cutoff_date handling
                large_sheet_entries[cleaned_name] = {
                    'entries': large_entries_for_type,
                    'first_sheet_name': first_sheet_name,
                    'normal_preprocessed': preprocessed_dfs  # Include any normal-sized sheets
                }
                xds[first_sheet_name] = None  # Placeholder for chunked processing
                preprocessed_flags[cleaned_name] = False  # Will be set True after chunked processing
            
            # Concatenate all preprocessed DataFrames for this sheet type (normal sheets only)
            elif preprocessed_dfs:
                if len(preprocessed_dfs) == 1:
                    # Single file - no concatenation needed
                    xds[first_sheet_name] = preprocessed_dfs[0]
                    preprocessed_flags[cleaned_name] = True
                else:
                    # Multiple files - concatenate!
                    merged_df = pd.concat(preprocessed_dfs, ignore_index=True)
                    xds[first_sheet_name] = merged_df
                    preprocessed_flags[cleaned_name] = True
                    logger.info(f"[MULTI-FILE] ✓ Merged {len(preprocessed_dfs)} '{cleaned_name}' sheets: {len(merged_df)} total rows")
            
            # Initialize processing stats (aggregate across all files)
            processing_stats.append({
                'sheet_name': first_sheet_name,
                'initial_columns': 0,  # Will be updated during processing
                'initial_records': total_rows,
                'processed_columns': None,
                'valid_records': 0
            })
        
        # Ensure all required sheets exist
        processed_sheets = ensure_all_sheets_exist(xds)
        
        upload_session.update_progress('data_mapping', 15, 'Sheets validated, starting processing...')
        
        # Calculate cutoff_date for LOANEFFECTIVEDATE/ACCOUNTSTATUSDATE validation
        # Cutoff = last day of reporting month (or current month if not specified)
        cutoff_date = None
        if reporting_month and reporting_year:
            from calendar import monthrange
            from datetime import datetime
            last_day = monthrange(reporting_year, reporting_month)[1]
            cutoff_date = datetime(reporting_year, reporting_month, last_day, 23, 59, 59)
            logger.info(f"[DATE VALIDATION] Using cutoff date: {cutoff_date.strftime('%Y-%m-%d')} (end of reporting period)")
        else:
            # Default to end of current month if no reporting period specified
            from datetime import datetime
            from calendar import monthrange
            today = datetime.now()
            last_day = monthrange(today.year, today.month)[1]
            cutoff_date = datetime(today.year, today.month, last_day, 23, 59, 59)
            logger.info(f"[DATE VALIDATION] No reporting period specified, using current month end: {cutoff_date.strftime('%Y-%m-%d')}")
        
        # Process each sheet using appropriate method (chunked vs normal)
        # Use xds.keys() since we now combine sheets from multiple files
        sheet_list = list(xds.keys())
        total_sheets = len(sheet_list)
        for idx, sheet_name in enumerate(sheet_list):
            progress = 15 + (idx * 40 // total_sheets)  # Progress from 15% to 55%
            cleaned_name = clean_sheet_name(sheet_name)
            
            # Get the file path for this sheet (from multi-file tracking - use first file with this type)
            sheet_entries = sheets_by_type.get(cleaned_name, [])
            sheet_file_path = sheet_entries[0]['file_path'] if sheet_entries else file_paths[0]
            
            # Check if this sheet needs chunked processing
            needs_chunking = False
            for fp, fp_info in all_sheet_info.items():
                if sheet_name in fp_info and fp_info[sheet_name]['use_chunking']:
                    needs_chunking = True
                    chunk_info = fp_info[sheet_name]
                    break
            
            # Check if this sheet has large entries tracked for multi-file processing
            if cleaned_name in large_sheet_entries:
                # Multi-file large sheet processing
                large_info = large_sheet_entries[cleaned_name]
                entries_list = large_info['entries']
                normal_preprocessed = large_info['normal_preprocessed']
                
                total_large_rows = sum(e['info']['row_count'] for e in entries_list)
                upload_session.update_progress(
                    'data_mapping' if progress < 30 else 'data_cleaning',
                    progress,
                    f'Processing {len(entries_list)} large sheet(s) for {sheet_name} ({total_large_rows:,} total rows)...'
                )
                
                # Process each large file's sheet and collect preprocessed results
                all_preprocessed = list(normal_preprocessed)  # Start with any normal-sized sheets
                
                for entry in entries_list:
                    entry_file_path = entry['file_path']
                    entry_sheet_name = entry['sheet_name']
                    
                    logger.info(f"[MULTI-FILE CHUNKED] Processing large sheet '{entry_sheet_name}' from {os.path.basename(entry_file_path)}")
                    
                    # Process large sheet in chunks
                    processed_chunk_df = process_large_sheet_chunked(
                        entry_file_path, 
                        entry_sheet_name, 
                        cleaned_name, 
                        chunk_size=CHUNK_SIZE,
                        cutoff_date=cutoff_date
                    )
                    
                    if not processed_chunk_df.empty:
                        all_preprocessed.append(processed_chunk_df)
                        logger.info(f"[MULTI-FILE CHUNKED] ✓ Processed '{entry_sheet_name}': {len(processed_chunk_df)} rows")
                    
                    # Force garbage collection between files
                    gc.collect()
                
                # Merge all preprocessed DataFrames (from normal + large sheets)
                if all_preprocessed:
                    if len(all_preprocessed) == 1:
                        merged_df = all_preprocessed[0]
                    else:
                        merged_df = pd.concat(all_preprocessed, ignore_index=True)
                        logger.info(f"[MULTI-FILE CHUNKED] ✓ Merged {len(all_preprocessed)} sheets for '{cleaned_name}': {len(merged_df)} total rows")
                    
                    processed_sheets[cleaned_name] = merged_df
                    preprocessed_flags[cleaned_name] = True
                    
                    # Update stats
                    for stat in processing_stats:
                        if stat['sheet_name'] == sheet_name:
                            stat['processed_columns'] = len(merged_df.columns)
                            stat['valid_records'] = len(merged_df)
                            break
                else:
                    logger.warning(f"[MULTI-FILE CHUNKED] No data processed for '{cleaned_name}'")
                    processed_sheets[cleaned_name] = pd.DataFrame()
                
                # Clean up
                del all_preprocessed
                gc.collect()
                
            elif needs_chunking:
                # Single-file large sheet processing (legacy path)
                upload_session.update_progress(
                    'data_mapping' if progress < 30 else 'data_cleaning',
                    progress,
                    f'Processing large sheet {sheet_name} ({chunk_info["row_count"]:,} rows) in chunks...'
                )
                
                processed_df = process_large_sheet_chunked(
                    sheet_file_path, 
                    sheet_name, 
                    cleaned_name, 
                    chunk_size=CHUNK_SIZE,
                    cutoff_date=cutoff_date
                )
                
                processed_sheets[cleaned_name] = processed_df
                
                # Update stats
                for stat in processing_stats:
                    if stat['sheet_name'] == sheet_name:
                        stat['processed_columns'] = len(processed_df.columns)
                        stat['valid_records'] = len(processed_df)
                        break
                
                # Force garbage collection after large sheet
                gc.collect()
                
            else:
                # Process normal-sized sheet
                upload_session.update_progress(
                    'data_mapping' if progress < 30 else 'data_cleaning',
                    progress,
                    f'Processing sheet {sheet_name}...'
                )
                
                sheet_data = xds[sheet_name]
                if sheet_data is not None:
                    # Check if this sheet was already preprocessed during multi-file loading
                    already_preprocessed = preprocessed_flags.get(cleaned_name, False)
                    
                    processed_df = process_single_sheet(
                        sheet_data, 
                        cleaned_name, 
                        cutoff_date=cutoff_date,
                        skip_preprocessing=already_preprocessed
                    )
                    processed_sheets[cleaned_name] = processed_df
                    
                    # Update stats
                    for stat in processing_stats:
                        if stat['sheet_name'] == sheet_name:
                            stat['initial_columns'] = len(sheet_data.columns)
                            stat['processed_columns'] = len(processed_df.columns)
                            stat['valid_records'] = len(processed_df)
                            break
        
        # ========================================================================
        # VALIDATE REQUIRED COLUMNS (After Renaming, Before Processing)
        # ========================================================================
        upload_session.update_progress('data_cleaning', 54, 'Validating required columns...')
        
        logger.info(f"[[VALIDATION] Starting column validation for {len(processed_sheets)} sheets")
        is_valid, validation_message = validate_required_columns(processed_sheets)
        
        if not is_valid:
            # Validation failed - mark session as failed and stop processing
            logger.info(f"[[VALIDATION] ❌ Validation failed: {validation_message}")
            
            upload_session.update_progress(
                'upload_validation', 
                0, 
                'Validation Failed: Required columns are empty'
            )
            
            # Log detailed error to activity log
           
            upload_session.activity_log = json.dumps({
                'validation_error': validation_message,
                'timestamp': dt.now().isoformat()
            })
            
            upload_session.mark_failed(validation_message)
            upload_session.save()
            
            # Clean up the uploaded file
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    logger.info(f"[[CLEANUP] Removed uploaded file: {file_path}")
                except Exception as e:
                    logger.info(f"[[CLEANUP] Failed to remove file: {e}")
            
            return  # Stop processing
        
        logger.info(f"[[VALIDATION] ✓ {validation_message}")
        upload_session.update_progress('credit_matching', 55, 'Validation successful, saving to temporary storage...')
        
        # ========================================================================
        # SAVE PROCESSED SHEETS TO PARQUET (Clear memory)
        # ========================================================================
        
        # Track which sheets were actually saved to Parquet
        saved_to_parquet = []
        
        # Save each processed sheet to Parquet and clear from memory
        # IMPORTANT: Also save empty auto-generated sheets to preserve headers for merging
        for sheet_key, sheet_df in processed_sheets.items():
            if sheet_df is not None:
                # Save both non-empty processed sheets AND empty auto-generated sheets (with headers)
                # Empty auto-generated sheets are crucial for merge operations to have proper columns
                save_to_parquet(sheet_df, upload_session_id, sheet_key)
                saved_to_parquet.append(sheet_key)
                processed_sheets[sheet_key] = None  # Clear from memory
                
        # Force garbage collection after clearing all sheets
        gc.collect()
        logger.info(f"[[MEMORY] Cleared all processed sheets from memory, saved to Parquet")
        logger.info(f"[[MEMORY] Sheets saved to Parquet (including auto-generated with headers): {saved_to_parquet}")
        
        # Initialize credit DataFrame (will be empty for merged sheets)
        credit = pd.DataFrame()
        
        upload_session.update_progress('credit_matching', 58, 'Matching borrowers with credits...')
        
        # Handle merged sheets or perform merging
        if 'consumermerged' in saved_to_parquet or 'commercialmerged' in saved_to_parquet:
            # Load merged sheets from Parquet (only when needed)
            indi = load_from_parquet(upload_session_id, 'consumermerged') if 'consumermerged' in saved_to_parquet else pd.DataFrame()
            corpo = load_from_parquet(upload_session_id, 'commercialmerged') if 'commercialmerged' in saved_to_parquet else pd.DataFrame()
            individual_credit_unmerged = {'unmerged_count': 0, 'total_valid_credit': 0, 'matched_count': 0}
            corporate_credit_unmerged = {'unmerged_count': 0, 'total_valid_credit': 0, 'matched_count': 0}
        else:
            # Load only the sheets that were actually saved to Parquet
            logger.info("[MEMORY] Loading sheets from Parquet for merging...")
            
            consu = load_from_parquet(upload_session_id, 'individualborrowertemplate') if 'individualborrowertemplate' in saved_to_parquet else pd.DataFrame()
            comm = load_from_parquet(upload_session_id, 'corporateborrowertemplate') if 'corporateborrowertemplate' in saved_to_parquet else pd.DataFrame()
            credit = load_from_parquet(upload_session_id, 'creditinformation') if 'creditinformation' in saved_to_parquet else pd.DataFrame()
            guar = load_from_parquet(upload_session_id, 'guarantorsinformation') if 'guarantorsinformation' in saved_to_parquet else pd.DataFrame()
            prin = load_from_parquet(upload_session_id, 'principalofficerstemplate') if 'principalofficerstemplate' in saved_to_parquet else pd.DataFrame()
            
            individual_credit_unmerged = {'unmerged_count': 0, 'total_valid_credit': 0, 'matched_count': 0}
            corporate_credit_unmerged = {'unmerged_count': 0, 'total_valid_credit': 0, 'matched_count': 0}
            
            # Perform merging and track matched credit IDs
            upload_session.update_progress('credit_matching', 65, 'Merging individual borrowers with credits...')
            indi, individual_matched_credit_ids = merge_individual_borrowers(consu, credit, guar)
            
            # Clear merged source data from memory (keep credit for now to identify unmatched)
            del consu, guar
            gc.collect()
            
            upload_session.update_progress('credit_matching', 70, 'Merging corporate borrowers with credits...')
            corpo, corporate_matched_credit_ids = merge_corporate_borrowers(comm, credit, prin)
            
            # Clear merged source data from memory (keep credit for now to identify unmatched)
            del comm, prin
            gc.collect()
            
            # Identify truly unmatched credits (didn't match either individual or corporate)
            upload_session.update_progress('credit_matching', 73, 'Identifying unmatched credits...')
            
            # Combine all matched credit IDs from both merges
            all_matched_credit_ids = individual_matched_credit_ids | corporate_matched_credit_ids
            
            # Find credits that were not matched in either merge
            # A credit is considered "matched" if EITHER:
            # 1. Its CUSTOMERID matches a borrower's CUSTOMERID (primary merge), OR
            # 2. Its ACCOUNTNUMBER matches a borrower's CUSTOMERID (fallback merge)
            actual_unmatched_count = 0  # Initialize the ground truth count
            if not credit.empty and 'CUSTOMERID' in credit.columns:
                # Primary match: credit.CUSTOMERID in matched IDs
                primary_match_mask = credit['CUSTOMERID'].isin(all_matched_credit_ids)
                
                # Fallback match: credit.ACCOUNTNUMBER in matched IDs
                # This handles the case where merge matched borrower.CUSTOMERID -> credit.ACCOUNTNUMBER
                fallback_match_mask = pd.Series(False, index=credit.index)
                if 'ACCOUNTNUMBER' in credit.columns:
                    fallback_match_mask = credit['ACCOUNTNUMBER'].isin(all_matched_credit_ids)
                
                # Combined: credit is matched if EITHER condition is true
                matched_mask = primary_match_mask | fallback_match_mask
                unmatched_credits = credit[~matched_mask].copy()
                
                # Keep only the 5 required columns for the report
                required_columns = ['CUSTOMERID', 'ACCOUNTNUMBER', 'ACCOUNTSTATUS', 'CREDITLIMIT', 'AVAILEDLIMIT']
                available_columns = [col for col in required_columns if col in unmatched_credits.columns]
                
                unmatched_credits_report = unmatched_credits[available_columns].copy()
                
                # Store the actual count (ground truth)
                actual_unmatched_count = len(unmatched_credits_report)
                
                logger.info(f"[[UNMATCHED CREDITS] Total credits: {len(credit)}")
                logger.info(f"[[UNMATCHED CREDITS] Matched to individual: {len(individual_matched_credit_ids)}")
                logger.info(f"[[UNMATCHED CREDITS] Matched to corporate: {len(corporate_matched_credit_ids)}")
                logger.info(f"[[UNMATCHED CREDITS] Truly unmatched (GROUND TRUTH): {actual_unmatched_count}")
                
                # Save unmatched credits to Parquet (will be preserved for Excel report)
                if not unmatched_credits_report.empty:
                    save_to_parquet(unmatched_credits_report, upload_session_id, 'unmatched_credits')
                    logger.info(f"[[UNMATCHED CREDITS] Saved {actual_unmatched_count} unmatched credits to Parquet")
                else:
                    # Save empty DataFrame with correct columns to Parquet
                    empty_unmatched = pd.DataFrame(columns=available_columns)
                    save_to_parquet(empty_unmatched, upload_session_id, 'unmatched_credits')
                    logger.info(f"[[UNMATCHED CREDITS] No unmatched credits - saved empty DataFrame with headers")
            else:
                # No credit data available
                empty_unmatched = pd.DataFrame(columns=['CUSTOMERID', 'ACCOUNTNUMBER', 'ACCOUNTSTATUS', 'CREDITLIMIT', 'AVAILEDLIMIT'])
                save_to_parquet(empty_unmatched, upload_session_id, 'unmatched_credits')
                logger.info(f"[[UNMATCHED CREDITS] No credit data - saved empty DataFrame")
            
            # Store the actual unmatched count in upload session for result page display
            upload_session.unmatched_credit_records = actual_unmatched_count
            upload_session.save()
            logger.info(f"[[UNMATCHED CREDITS] Stored ground truth count in upload_session: {actual_unmatched_count}")
            
            # Now safe to delete credit DataFrame
            del credit
            gc.collect()
        
        upload_session.update_progress('credit_matching', 75, 'Credit matching complete, identifying split candidates...')
        
        # Save merged data to Parquet before splitting (needed for post-verification)
        save_to_parquet(indi, upload_session_id, 'merged_individual')
        save_to_parquet(corpo, upload_session_id, 'merged_corporate')
        
        # Split commercial/consumer entities for human verification
        split_indi, split_candidates_commercial = split_commercial_entities(indi)
        split_corpo, split_candidates_consumer = split_consumer_entities(corpo)
        
        # Clear original merged data from memory (keep only split versions)
        del indi, corpo
        gc.collect()
        
        # Save split data to Parquet (needed for post-verification final output)
        save_to_parquet(split_indi, upload_session_id, 'split_individual')
        save_to_parquet(split_corpo, upload_session_id, 'split_corporate')
        
        # Reorder columns for display
        reordered_commercial_columns = reorder_commercial_columns(split_candidates_commercial.columns)
        split_candidates_commercial = split_candidates_commercial.reindex(columns=reordered_commercial_columns)
        
        reordered_consumer_columns = reorder_consumer_columns(split_candidates_consumer.columns)
        split_candidates_consumer = split_candidates_consumer.reindex(columns=reordered_consumer_columns)
        
        # Prepare verification data (lightweight - only candidates for UI)
        commercial_records = json.loads(split_candidates_commercial.to_json(orient='records'))
        consumer_records = json.loads(split_candidates_consumer.to_json(orient='records'))
        
        # Determine final reporting period with priority system
        # Priority: 1) User-selected, 2) Filename extraction, 3) Previous month default
        from datetime import datetime
        extracted_month, extracted_year = extract_date_from_filename(original_filenames[0])
        
        now = datetime.now()
        default_month = now.month - 1 if now.month > 1 else 12
        default_year = now.year if now.month > 1 else now.year - 1
        
        final_month = reporting_month or extracted_month or default_month
        final_year = reporting_year or extracted_year or default_year
        
        logger.info(f"[DATE DETERMINATION] User-selected: {reporting_month}/{reporting_year}")
        logger.info(f"[DATE DETERMINATION] Filename-extracted: {extracted_month}/{extracted_year}")
        logger.info(f"[DATE DETERMINATION] Final decision: {final_month}/{final_year}")
        
        # Prepare processing data (common to both paths)
        processing_data_dict = {
            # Candidates for verification UI
            'commercial_candidates': commercial_records,
            'consumer_candidates': consumer_records,
            'columns_commercial': list(reordered_commercial_columns),
            'columns_consumer': list(reordered_consumer_columns),
            
            # Metadata
            'processing_stats': json.loads(json.dumps(processing_stats, default=convert_numpy)),
            'original_filename': original_filenames[0],
            'subscriber_id': subscriber_id,
            'subscriber_name': subscriber_name,
            
            # Reporting period (final decision)
            'final_month': final_month,
            'final_year': final_year,
            
            # Credit stats (small)
            'individual_credit_unmerged': {
                k: int(v) if isinstance(v, (np.int64, np.int32)) else v 
                for k, v in individual_credit_unmerged.items()
            },
            'corporate_credit_unmerged': {
                k: int(v) if isinstance(v, (np.int64, np.int32)) else v 
                for k, v in corporate_credit_unmerged.items()
            },
            
            # Flag indicating Parquet files are available
            'using_parquet': True
        }
        
        upload_session.processing_data = json.dumps(processing_data_dict)
        
        # Check if verification should be auto-skipped (no candidates found)
        if len(commercial_records) == 0 and len(consumer_records) == 0:
            # Auto-skip verification - no candidates to review
            logger.info(f"[[AUTO-SKIP] No verification candidates found for UploadSession {upload_session_id}")
            logger.info(f"[[AUTO-SKIP] Skipping verification and proceeding to post-verification processing")
            
            # Save processing_data first (mark_verification_skipped uses update_fields, doesn't include processing_data)
            upload_session.save(update_fields=['processing_data'])
            
            upload_session.mark_verification_skipped()
            
            # Immediately trigger post-verification processing
            process_post_verification(
                upload_session_id=upload_session_id,
                use_parquet=True,
                processing_stats=processing_stats,
                original_filename=original_filenames[0],
                subscriber_id=subscriber_id,
                subscriber_name=subscriber_name,
                total_individual_records=len(split_indi),
                total_corporate_records=len(split_corpo),
                final_month=final_month,
                final_year=final_year
            )
            
            logger.info(f"[[AUTO-SKIP] Post-verification processing triggered for UploadSession {upload_session_id}")
        else:
            # Mark as awaiting verification (original behavior)
            upload_session.update_progress('awaiting_verification', 80, 'Processing complete, awaiting human verification...')
            
            # Mark as awaiting verification
            upload_session.mark_awaiting_verification(
                commercial_count=len(commercial_records),
                consumer_count=len(consumer_records)
            )
            
            upload_session.save()
            
            logger.info(f"[[ASYNC TASK] Processing complete for UploadSession {upload_session_id}")
            logger.info(f"[[ASYNC TASK] Awaiting human verification: {len(commercial_records)} commercial, {len(consumer_records)} consumer candidates")
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.info(f"[ASYNC TASK ERROR] {error_details}")
        
        # Mark upload session as failed
        try:
            upload_session = UploadSession.objects.get(id=upload_session_id)
            upload_session.mark_failed(error_message=str(e))
            
            # Cleanup temp Parquet files on error
            cleanup_temp_files(upload_session_id)
        except Exception:
            pass
        
        raise
    
    finally:
        # Clean up all temp uploaded files (supports multi-file upload)
        for fp in file_paths:
            if os.path.exists(fp):
                try:
                    os.remove(fp)
                    logger.info(f"[ASYNC TASK] Removed temp file: {fp}")
                except Exception as e:
                    logger.info(f"[ASYNC TASK] Failed to remove temp file: {e}")


def process_post_verification(upload_session_id, use_parquet=True,
                               processing_stats=None, original_filename=None,
                               subscriber_id=None, subscriber_name=None,
                               total_individual_records=0, total_corporate_records=0,
                               final_month=None, final_year=None):
    """
    Asynchronous task to process post-verification file generation
    
    Args:
        upload_session_id: ID of the UploadSession record
        use_parquet: Whether to load data from Parquet files (True) or JSON (False - legacy)
        processing_stats: Processing statistics
        original_filename: Original filename for naming outputs
        subscriber_id: Subscriber ID
        subscriber_name: Subscriber name
        total_individual_records: Count of individual records (may be recalculated)
        total_corporate_records: Count of corporate records (may be recalculated)
    """
    from .views import (
        modify_middle_names, clean_for_output, remove_duplicates,
        generate_filename_from_user, extract_date_from_filename,
        save_excel_as_text, enforce_string_columns, remove_titles,
        transform_to_commercial, transform_to_consumer
    )
    from .map import guarantor_columns_to_clear, principal_officer_columns_to_clear
    
    try:
        logger.info(f"[POST-VERIFICATION TASK] Starting post-verification processing for UploadSession {upload_session_id}")
        
        upload_session = UploadSession.objects.get(id=upload_session_id)
        
        # Update progress to post-verification stage
        upload_session.update_progress('post_verification', 85, 'Processing verification results...')
        
        # Load DataFrames from Parquet (memory efficient) or JSON (legacy)
        if use_parquet:
            logger.info("[POST-VERIFICATION] Loading final data from Parquet files")
            indi = load_from_parquet(upload_session_id, 'split_individual')
            corpo = load_from_parquet(upload_session_id, 'split_corporate')
            indi = enforce_string_columns(indi)
            corpo = enforce_string_columns(corpo)
        else:
            # Legacy: Load from JSON (for backward compatibility)
            logger.info("[POST-VERIFICATION] Loading final data from JSON (legacy mode)")
            raise NotImplementedError("JSON loading not implemented - use Parquet")
        
        # ========================================================================
        # APPLY VERIFICATION DECISIONS (if submitted by user)
        # ========================================================================
        processing_data = json.loads(upload_session.processing_data or '{}')
        verification_submitted = processing_data.get('verification_submitted', False)
        
        if verification_submitted:
            logger.info("[POST-VERIFICATION] Applying user verification decisions...")
            upload_session.update_progress('post_verification', 86, 'Applying verification decisions...')
            
            commercial_moves = processing_data.get('commercial_moves', [])
            consumer_moves = processing_data.get('consumer_moves', [])
            
            # Load verification candidates from processing_data
            commercial_candidates = processing_data.get('commercial_candidates', [])
            consumer_candidates = processing_data.get('consumer_candidates', [])
            
            split_candidates_commercial = pd.DataFrame(commercial_candidates)
            split_candidates_consumer = pd.DataFrame(consumer_candidates)
            
            if not split_candidates_commercial.empty:
                split_candidates_commercial = enforce_string_columns(split_candidates_commercial)
            if not split_candidates_consumer.empty:
                split_candidates_consumer = enforce_string_columns(split_candidates_consumer)
            
            logger.info(f"[POST-VERIFICATION] Commercial candidates: {len(split_candidates_commercial)}, Consumer candidates: {len(split_candidates_consumer)}")
            logger.info(f"[POST-VERIFICATION] Commercial moves: {len(commercial_moves)}, Consumer moves: {len(consumer_moves)}")
            
            # For commercial candidates: checked = move to corpo, unchecked = stay in indi
            move_to_corp_idx = [i for i, move in enumerate(commercial_moves) if move]
            stay_in_indi_idx = [i for i, move in enumerate(commercial_moves) if not move]
            
            # Validate indices are within bounds for commercial candidates
            if len(split_candidates_commercial) > 0:
                commercial_max_idx = len(split_candidates_commercial) - 1
                move_to_corp_idx = [i for i in move_to_corp_idx if i <= commercial_max_idx]
                stay_in_indi_idx = [i for i in stay_in_indi_idx if i <= commercial_max_idx]
            else:
                move_to_corp_idx = []
                stay_in_indi_idx = []
            
            # Separate checked vs unchecked commercial candidates
            checked_commercial = split_candidates_commercial.iloc[move_to_corp_idx].copy() if move_to_corp_idx else pd.DataFrame()
            unchecked_commercial = split_candidates_commercial.iloc[stay_in_indi_idx].copy() if stay_in_indi_idx else pd.DataFrame()
            
            # Debug logging for "no candidates checked" scenario
            logger.info(f"[POST-VERIFICATION] Commercial: checked={len(checked_commercial)}, unchecked={len(unchecked_commercial)}")
            
            # For consumer candidates: checked = move to indi, unchecked = stay in corpo
            move_to_indi_idx = [i for i, move in enumerate(consumer_moves) if move]
            stay_in_corp_idx = [i for i, move in enumerate(consumer_moves) if not move]
            
            # Validate indices are within bounds for consumer candidates
            if len(split_candidates_consumer) > 0:
                consumer_max_idx = len(split_candidates_consumer) - 1
                move_to_indi_idx = [i for i in move_to_indi_idx if i <= consumer_max_idx]
                stay_in_corp_idx = [i for i in stay_in_corp_idx if i <= consumer_max_idx]
            else:
                move_to_indi_idx = []
                stay_in_corp_idx = []
            
            logger.info(f"[POST-VERIFICATION] Move to corp: {len(move_to_corp_idx)}, Stay in indi: {len(stay_in_indi_idx)}")
            logger.info(f"[POST-VERIFICATION] Move to indi: {len(move_to_indi_idx)}, Stay in corp: {len(stay_in_corp_idx)}")
            
            # Separate checked vs unchecked consumer candidates
            checked_consumer = split_candidates_consumer.iloc[move_to_indi_idx].copy() if move_to_indi_idx else pd.DataFrame()
            unchecked_consumer = split_candidates_consumer.iloc[stay_in_corp_idx].copy() if stay_in_corp_idx else pd.DataFrame()
            
            # Debug logging for "no candidates checked" scenario
            logger.info(f"[POST-VERIFICATION] Consumer: checked={len(checked_consumer)}, unchecked={len(unchecked_consumer)}")
            
            # Return unchecked records to original DataFrames (no transformation)
            if not unchecked_commercial.empty:
                # Restore original individual name structure for unchecked commercial candidates
                if 'ORIGINAL_BUSINESSNAME' in unchecked_commercial.columns:
                    for idx, row in unchecked_commercial.iterrows():
                        if pd.notna(row['ORIGINAL_BUSINESSNAME']):
                            cleaned_business_name = remove_titles(str(row['ORIGINAL_BUSINESSNAME']))
                            name_parts = cleaned_business_name.split(maxsplit=2)
                            unchecked_commercial.at[idx, 'SURNAME'] = name_parts[0] if len(name_parts) > 0 else ''
                            unchecked_commercial.at[idx, 'FIRSTNAME'] = name_parts[1] if len(name_parts) > 1 else ''
                            unchecked_commercial.at[idx, 'MIDDLENAME'] = name_parts[2] if len(name_parts) > 2 else ''
                    unchecked_commercial = unchecked_commercial.drop(columns=['ORIGINAL_BUSINESSNAME'], errors='ignore')
                indi = pd.concat([indi, unchecked_commercial], ignore_index=True)
            
            if not unchecked_consumer.empty:
                # Restore original business name for unchecked consumer records
                if 'ORIGINAL_BUSINESSNAME' in unchecked_consumer.columns:
                    unchecked_consumer['BUSINESSNAME'] = unchecked_consumer['ORIGINAL_BUSINESSNAME']
                    columns_to_drop = ['ORIGINAL_BUSINESSNAME', 'SURNAME', 'FIRSTNAME', 'MIDDLENAME', 'DEPENDANTS']
                    unchecked_consumer = unchecked_consumer.drop(columns=[col for col in columns_to_drop if col in unchecked_consumer.columns], errors='ignore')
                corpo = pd.concat([corpo, unchecked_consumer], ignore_index=True)
            
            # Transform ONLY checked records
            confirmed_commercial = pd.DataFrame()
            confirmed_consumer = pd.DataFrame()
            
            if not checked_commercial.empty:
                # Moving individual to commercial
                confirmed_commercial = transform_to_commercial(
                    checked_commercial, 
                    columns_to_clear=guarantor_columns_to_clear
                )
                logger.info(f"[POST-VERIFICATION] Transformed {len(confirmed_commercial)} records to commercial")
                
            if not checked_consumer.empty:
                # Moving corporate to consumer
                confirmed_consumer = transform_to_consumer(
                    checked_consumer, 
                    columns_to_clear=principal_officer_columns_to_clear
                )
                logger.info(f"[POST-VERIFICATION] Transformed {len(confirmed_consumer)} records to consumer")
            
            # Concatenate transformed records
            if not confirmed_consumer.empty:
                indi = pd.concat([indi, confirmed_consumer], ignore_index=True)
            if not confirmed_commercial.empty:
                corpo = pd.concat([corpo, confirmed_commercial], ignore_index=True)
            
            logger.info(f"[POST-VERIFICATION] After applying decisions: indi={len(indi)}, corpo={len(corpo)}")
        else:
            logger.info("[POST-VERIFICATION] No verification decisions to apply (auto-skipped or legacy)")
        
        # ========================================================================
        # APPLY FINAL TRANSFORMATIONS
        # ========================================================================
        upload_session.update_progress('post_verification', 88, 'Applying final data transformations...')
        indi = modify_middle_names(indi)
        corpo = modify_middle_names(corpo)
        
        indi = clean_for_output(indi)
        corpo = clean_for_output(corpo)
        
        # Drop name and dependant columns from corpo
        columns_to_remove = ['SURNAME', 'FIRSTNAME', 'MIDDLENAME', 'DEPENDANTS']
        corpo = corpo.drop(columns=[col for col in columns_to_remove if col in corpo.columns], errors='ignore')
        
        indi = remove_duplicates(indi)
        corpo = remove_duplicates(corpo)
        
        # ========================================================================
        # REMOVE BLANK ROWS BEFORE VALIDATION
        # ========================================================================
        # Filter out completely empty rows BEFORE validation - these are garbage
        # data (trailing empty rows in Excel, formatting artifacts) and should not
        # be counted as valid OR excluded records
        indi_before_blank_removal = len(indi)
        corpo_before_blank_removal = len(corpo)
        
        indi = indi.replace('', pd.NA).dropna(how='all').reset_index(drop=True)
        corpo = corpo.replace('', pd.NA).dropna(how='all').reset_index(drop=True)
        # Replace NA back to empty strings for processing
        indi = indi.fillna('')
        corpo = corpo.fillna('')
        
        blank_indi_removed = indi_before_blank_removal - len(indi)
        blank_corpo_removed = corpo_before_blank_removal - len(corpo)
        
        if blank_indi_removed > 0 or blank_corpo_removed > 0:
            logger.info(f"[POST-VERIFICATION] Removed blank rows: {blank_indi_removed} individual, {blank_corpo_removed} corporate")
        
        # ========================================================================
        # DATA QUALITY VALIDATION (After duplicates and blank rows removed)
        # ========================================================================
        upload_session.update_progress('post_verification', 89, 'Applying data quality validation...')
        
        indi, corpo, excluded_indi, excluded_corpo = apply_data_quality_validation(indi, corpo)
        
        excluded_indi_count = len(excluded_indi)
        excluded_corpo_count = len(excluded_corpo)
        
        logger.info(f"[POST-VERIFICATION] Data quality validation complete:")
        logger.info(f"[POST-VERIFICATION]   Individual: {len(indi)} valid, {excluded_indi_count} excluded")
        logger.info(f"[POST-VERIFICATION]   Corporate: {len(corpo)} valid, {excluded_corpo_count} excluded")
        
        # Safety check: ensure no blank rows in excluded DataFrames
        # (This should be a no-op since blank rows are filtered before validation,
        #  but kept as defense-in-depth in case any slip through)
        if not excluded_indi.empty:
            original_count = len(excluded_indi)
            excluded_indi = excluded_indi.replace('', pd.NA).dropna(how='all').reset_index(drop=True)
            excluded_indi = excluded_indi.fillna('')
            if len(excluded_indi) != original_count:
                logger.warning(f"[POST-VERIFICATION] Safety check removed {original_count - len(excluded_indi)} blank excluded individual rows")
                excluded_indi_count = len(excluded_indi)
        
        if not excluded_corpo.empty:
            original_count = len(excluded_corpo)
            excluded_corpo = excluded_corpo.replace('', pd.NA).dropna(how='all').reset_index(drop=True)
            excluded_corpo = excluded_corpo.fillna('')
            if len(excluded_corpo) != original_count:
                logger.warning(f"[POST-VERIFICATION] Safety check removed {original_count - len(excluded_corpo)} blank excluded corporate rows")
                excluded_corpo_count = len(excluded_corpo)
        
        # Save excluded records to Parquet for Excel report generation
        if not excluded_indi.empty:
            save_to_parquet(excluded_indi, upload_session_id, 'excluded_individual')
            logger.info(f"[POST-VERIFICATION] Saved {excluded_indi_count} excluded individual records to Parquet")
        if not excluded_corpo.empty:
            save_to_parquet(excluded_corpo, upload_session_id, 'excluded_corporate')
            logger.info(f"[POST-VERIFICATION] Saved {excluded_corpo_count} excluded corporate records to Parquet")
        
        # Note: unmatched_credit_records was already calculated and stored during process_uploaded_file()
        # We preserve that ground truth value instead of recalculating here
        upload_session.update_progress('output_generation', 90, 'Generating output files...')
        
        # ========================================================================
        # GET REPORTING PERIOD (prefer passed parameters, fallback to processing_data)
        # ========================================================================
        # Use passed parameters if available (from auto-skip flow), otherwise fallback to DB
        if final_month is not None and final_year is not None:
            extracted_month = final_month
            extracted_year = final_year
            logger.info(f"[[OUTPUT] Using reporting period from parameters: {extracted_month}/{extracted_year}")
        else:
            # Fallback: retrieve from stored processing_data
            upload_session_obj = UploadSession.objects.get(id=upload_session_id)
            processing_data_json = json.loads(upload_session_obj.processing_data or '{}')
            extracted_month = processing_data_json.get('final_month')
            extracted_year = processing_data_json.get('final_year')
            logger.info(f"[[OUTPUT] Using reporting period from processing_data: {extracted_month}/{extracted_year}")
        
        # ========================================================================
        # SMART OUTPUT FORMAT SELECTION
        # ========================================================================
        # Configuration
        EXCEL_ROW_THRESHOLD = 300000  # Switch to TXT if sheet exceeds this
        
        # Generate output files with smart format selection
        upload_session.update_progress('output_generation', 92, 'Generating outputs with smart format selection...')
        
        # Create subscriber-specific output directory
        subscriber_name_clean = subscriber_name.replace(' ', '')
        subscriber_output_dir = os.path.join(settings.MEDIA_ROOT, 'outputs', subscriber_name_clean)
        os.makedirs(subscriber_output_dir, exist_ok=True)
        
        fs = FileSystemStorage()
        
        # ========================================================================
        # INDIVIDUAL SHEET - Smart Format Selection (skip if empty)
        # ========================================================================
        individual_row_count = len(indi)
        indi_output_format = None
        indi_processed_file_url = None
        indi_output_filename = None
        
        # Only generate file if there are actual data rows
        if individual_row_count > 0 and not indi.empty:
            if individual_row_count > EXCEL_ROW_THRESHOLD:
                # Generate TXT for large individual sheet
                logger.info(f"[[OUTPUT] Individual sheet has {individual_row_count} rows - generating TXT (threshold: {EXCEL_ROW_THRESHOLD})")
                indi_output_filename = generate_filename_from_user(subscriber_id, subscriber_name, 'txt', 'consumer', extracted_month, extracted_year)
                indi_output_path = os.path.join(subscriber_output_dir, indi_output_filename)
                
                # Re-upload detection: save as _reuploaded_N if file already exists
                if os.path.exists(indi_output_path):
                    base, ext = os.path.splitext(indi_output_filename)
                    reupload_num = 1
                    while True:
                        indi_output_filename = f"{base}_reuploaded_{reupload_num}{ext}"
                        indi_output_path = os.path.join(subscriber_output_dir, indi_output_filename)
                        if not os.path.exists(indi_output_path):
                            break
                        reupload_num += 1
                    logger.info(f"[[OUTPUT] Existing file detected, saving as: {indi_output_filename}")
                
                indi.to_csv(indi_output_path, sep='\t', index=False, encoding='utf-8')
                indi_output_format = 'txt'
                logger.info(f"[[OUTPUT] Individual TXT generated: {indi_output_filename}")
            else:
                # Generate Excel for small individual sheet
                logger.info(f"[[OUTPUT] Individual sheet has {individual_row_count} rows - generating Excel (threshold: {EXCEL_ROW_THRESHOLD})")
                indi_output_filename = generate_filename_from_user(subscriber_id, subscriber_name, 'excel', 'consumer', extracted_month, extracted_year)
                indi_output_path = os.path.join(subscriber_output_dir, indi_output_filename)
                
                # Re-upload detection: save as _reuploaded_N if file already exists
                if os.path.exists(indi_output_path):
                    base, ext = os.path.splitext(indi_output_filename)
                    reupload_num = 1
                    while True:
                        indi_output_filename = f"{base}_reuploaded_{reupload_num}{ext}"
                        indi_output_path = os.path.join(subscriber_output_dir, indi_output_filename)
                        if not os.path.exists(indi_output_path):
                            break
                        reupload_num += 1
                    logger.info(f"[[OUTPUT] Existing file detected, saving as: {indi_output_filename}")
                
                save_excel_as_text(indi, indi_output_path)
                indi_output_format = 'xlsx'
                logger.info(f"[[OUTPUT] Individual Excel generated: {indi_output_filename}")
            
            # Store relative path for URL generation
            indi_output_relative = os.path.join('outputs', subscriber_name_clean, indi_output_filename)
            indi_processed_file_url = fs.url(indi_output_relative)
        else:
            logger.info(f"[[OUTPUT] Skipping individual output - no records (len={individual_row_count}, empty={indi.empty})")
        
        # ========================================================================
        # CORPORATE SHEET - Smart Format Selection (skip if empty)
        # ========================================================================
        corporate_row_count = len(corpo)
        corpo_output_format = None
        corpo_processed_file_url = None
        corpo_output_filename = None
        
        # Only generate file if there are actual data rows
        if corporate_row_count > 0 and not corpo.empty:
            if corporate_row_count > EXCEL_ROW_THRESHOLD:
                # Generate TXT for large corporate sheet
                logger.info(f"[[OUTPUT] Corporate sheet has {corporate_row_count} rows - generating TXT (threshold: {EXCEL_ROW_THRESHOLD})")
                corpo_output_filename = generate_filename_from_user(subscriber_id, subscriber_name, 'txt', 'commercial', extracted_month, extracted_year)
                corpo_output_path = os.path.join(subscriber_output_dir, corpo_output_filename)
                
                # Re-upload detection: save as _reuploaded_N if file already exists
                if os.path.exists(corpo_output_path):
                    base, ext = os.path.splitext(corpo_output_filename)
                    reupload_num = 1
                    while True:
                        corpo_output_filename = f"{base}_reuploaded_{reupload_num}{ext}"
                        corpo_output_path = os.path.join(subscriber_output_dir, corpo_output_filename)
                        if not os.path.exists(corpo_output_path):
                            break
                        reupload_num += 1
                    logger.info(f"[[OUTPUT] Existing file detected, saving as: {corpo_output_filename}")
                
                corpo.to_csv(corpo_output_path, sep='\t', index=False, encoding='utf-8')
                corpo_output_format = 'txt'
                logger.info(f"[[OUTPUT] Corporate TXT generated: {corpo_output_filename}")
            else:
                # Generate Excel for small corporate sheet
                logger.info(f"[[OUTPUT] Corporate sheet has {corporate_row_count} rows - generating Excel (threshold: {EXCEL_ROW_THRESHOLD})")
                corpo_output_filename = generate_filename_from_user(subscriber_id, subscriber_name, 'excel', 'commercial', extracted_month, extracted_year)
                corpo_output_path = os.path.join(subscriber_output_dir, corpo_output_filename)
                
                # Re-upload detection: save as _reuploaded_N if file already exists
                if os.path.exists(corpo_output_path):
                    base, ext = os.path.splitext(corpo_output_filename)
                    reupload_num = 1
                    while True:
                        corpo_output_filename = f"{base}_reuploaded_{reupload_num}{ext}"
                        corpo_output_path = os.path.join(subscriber_output_dir, corpo_output_filename)
                        if not os.path.exists(corpo_output_path):
                            break
                        reupload_num += 1
                    logger.info(f"[[OUTPUT] Existing file detected, saving as: {corpo_output_filename}")
                
                save_excel_as_text(corpo, corpo_output_path)
                corpo_output_format = 'xlsx'
                logger.info(f"[[OUTPUT] Corporate Excel generated: {corpo_output_filename}")
            
            # Store relative path for URL generation
            corpo_output_relative = os.path.join('outputs', subscriber_name_clean, corpo_output_filename)
            corpo_processed_file_url = fs.url(corpo_output_relative)
        else:
            logger.info(f"[[OUTPUT] Skipping corporate output - no records (len={corporate_row_count}, empty={corpo.empty})")
        
        # ========================================================================
        # UPDATE UPLOAD SESSION
        # ========================================================================
        upload_session.update_progress('output_generation', 98, 'Finalizing upload session...')
        
        upload_session.individual_file_path = indi_processed_file_url
        upload_session.corporate_file_path = corpo_processed_file_url
        # Store output formats for download UI (None if skipped)
        upload_session.individual_output_format = indi_output_format
        upload_session.corporate_output_format = corpo_output_format
        # Note: unmatched_credit_records already set during process_uploaded_file() - don't overwrite
        upload_session.individual_credit_matched = total_individual_records
        upload_session.corporate_credit_matched = total_corporate_records
        # Store excluded record counts for reporting
        upload_session.excluded_individual_records = excluded_indi_count
        upload_session.excluded_corporate_records = excluded_corpo_count
        
        # IMPORTANT: Save file paths and output data BEFORE calling mark_completed()
        # mark_completed() uses save(update_fields=[...]) which doesn't include file paths,
        # so we must persist them first to ensure download links appear for multi-subscriber users
        upload_session.save()
        
        # Recalculate actual record counts from final DataFrames
        # (async calls may pass 0 for these values, so we need to use actual counts)
        actual_individual_count = len(indi) if not indi.empty else 0
        actual_corporate_count = len(corpo) if not corpo.empty else 0
        
        logger.info(f"[POST-VERIFICATION] Final record counts: individual={actual_individual_count}, corporate={actual_corporate_count}")
        
        # Mark as completed (uses update_fields to only update status-related fields)
        upload_session.mark_completed(
            individual_count=actual_individual_count,
            corporate_count=actual_corporate_count,
            processing_time=None  # Can calculate if needed
        )
        
        # ========================================================================
        # EMAIL REPORT: Send analysis report to subscriber email
        # ========================================================================
        try:
            from .models import SubscriberEmail
            from .views import build_excel_report
            from django.core.mail import EmailMessage
            from django.conf import settings as django_settings
            
            subscriber_email = SubscriberEmail.get_email(upload_session.subscriber_id)
            if subscriber_email:
                logger.info(f"[EMAIL REPORT] Sending report to {subscriber_email} for session {upload_session_id}")
                
                # Build report (reuses same function as portal download)
                buffer, report_filename = build_excel_report(upload_session, user=upload_session.user)
                
                # Compose email
                subject = f"FCB Processing Report - {subscriber_name} - {final_month}/{final_year}" if final_month and final_year else f"FCB Processing Report - {subscriber_name}"
                body = (
                    f"Dear {subscriber_name},\n\n"
                    f"Your data processing has been completed successfully.\n\n"
                    f"Processing Summary:\n"
                    f"  - Individual Records: {actual_individual_count:,}\n"
                    f"  - Corporate Records: {actual_corporate_count:,}\n"
                    f"  - Unmatched Credits: {upload_session.unmatched_credit_records or 0:,}\n\n"
                    f"The detailed analysis report is attached as an Excel file.\n\n"
                    f"Best regards,\n"
                    f"FCB Auto Processing System"
                )
                
                email = EmailMessage(
                    subject=subject,
                    body=body,
                    from_email=django_settings.DEFAULT_FROM_EMAIL,
                    to=[subscriber_email],
                )
                
                # Attach report (check size < 20MB)
                report_bytes = buffer.getvalue()
                if len(report_bytes) < 20 * 1024 * 1024:  # 20MB limit
                    email.attach(report_filename, report_bytes, 
                                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    email.send(fail_silently=False)
                    logger.info(f"[EMAIL REPORT] Report sent successfully to {subscriber_email}")
                else:
                    logger.info(f"[EMAIL REPORT] Report too large ({len(report_bytes)} bytes), skipping attachment")
            else:
                logger.info(f"[EMAIL REPORT] No email configured for subscriber {upload_session.subscriber_id}, skipping")
        except Exception as email_error:
            # Email failure should NEVER break the processing pipeline
            logger.info(f"[EMAIL REPORT ERROR] Failed to send email: {email_error}")
        
        # ========================================================================
        # CLEANUP: Delete temporary Parquet files after successful completion
        # ========================================================================
        cleanup_temp_files(upload_session_id)
        
        logger.info(f"[POST-VERIFICATION TASK] Processing complete for UploadSession {upload_session_id}")
        logger.info(f"[POST-VERIFICATION TASK] Files generated: individual={indi_output_filename or 'None (skipped)'}, corporate={corpo_output_filename or 'None (skipped)'}")
        
        # Log download links for debugging multi-subscriber issues
        if indi_processed_file_url or corpo_processed_file_url:
            logger.info(f"[DOWNLOAD LINKS] Generated for session {upload_session_id} (subscriber: {subscriber_name})")
            if indi_processed_file_url:
                logger.info(f"[DOWNLOAD LINKS]   Individual: {indi_processed_file_url}")
            if corpo_processed_file_url:
                logger.info(f"[DOWNLOAD LINKS]   Corporate: {corpo_processed_file_url}") 
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        logger.info(f"[POST-VERIFICATION TASK ERROR] {error_details}")
        
        # Mark upload session as failed
        try:
            upload_session = UploadSession.objects.get(id=upload_session_id)
            upload_session.mark_failed(error_message=str(e))
            
            # Cleanup temp files even on error
            cleanup_temp_files(upload_session_id)
        except Exception:
            pass
        
        raise
